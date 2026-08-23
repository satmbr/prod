import io
import os
import re
import tempfile
import unittest
import uuid
from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import MagicMock, patch

from PIL import Image
from flask import request
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from werkzeug.datastructures import FileStorage
from werkzeug.datastructures import MultiDict

os.environ.setdefault("SECRET_KEY", "test-secret-key")

from app import ProdRequest, create_app
from routes.financeiro_novo.services.anexos import (
    AnexoInvalido,
    nome_objeto_pdf,
    normalizar_anexo,
)
from routes.financeiro_novo.cadastros import TIPOS, _normalizar
from routes.financeiro_novo.services.valores import ValorInvalido, decimal_br
from routes.financeiro_novo.services.exportacao_om import gerar_excel_om, gerar_pdf_om
from routes.financeiro_novo.services.empresas import empresa_valida, nome_empresa
from routes.financeiro_novo.homologacao import diagnosticar_armazenamento
from routes.financeiro_novo.missoes import _ler_linhas_om_excel, _linhas_om_formulario
from routes.financeiro_novo.reembolsos import _vincular_anexo


class FinanceiroNovoIsolamentoTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.raiz = Path(__file__).resolve().parents[1]

    def test_fundacao_nao_referencia_tabelas_do_financeiro_anterior(self):
        arquivos = [
            *list((self.raiz / "migrations").glob("*_financeiro_novo_*.sql")),
            *list((self.raiz / "routes" / "financeiro_novo").rglob("*.py")),
            *list((self.raiz / "templates" / "financeiro_novo").rglob("*.html")),
        ]
        conteudo = "\n".join(item.read_text(encoding="utf-8") for item in arquivos)
        self.assertNotIn("financeiro2_", conteudo.lower())
        self.assertIn("financeiro3_", conteudo.lower())

    def test_permissao_do_financeiro_atual_nao_abre_o_novo(self):
        app = create_app()
        app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        with app.test_client() as client:
            with client.session_transaction() as sessao:
                sessao["usuario_id"] = 1
                sessao["permissoes"] = ["financeiro:visualizar"]
                sessao["ultimo_acesso"] = 9999999999
            resposta = client.get("/financeiro-novo/")
        self.assertEqual(resposta.status_code, 403)

    def test_permissao_propria_abre_dashboard_sem_dados_antigos(self):
        resultado_config = MagicMock()
        resultado_config.mappings.return_value.one.return_value = {
            "nome_modulo": "Financeiro Novo",
            "ambiente": "HOMOLOGACAO",
            "versao_schema": 1,
        }
        resultado_totais = MagicMock()
        resultado_totais.mappings.return_value.one.return_value = {
            "despesas": 0,
            "missoes": 0,
            "acertos_pendentes": 0,
            "notas_debito": 0,
            "reembolsos": 0,
            "conciliacoes": 0,
            "cadastros": 0,
            "arquivos": 0,
            "anexos": 0,
            "eventos_auditoria": 0,
        }
        conexao = MagicMock()
        conexao.execute.side_effect = [resultado_config, resultado_totais]
        contexto = MagicMock()
        contexto.__enter__.return_value = conexao
        engine = MagicMock()
        engine.connect.return_value = contexto

        app = create_app()
        app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        with patch("routes.financeiro_novo.views.get_engine", return_value=engine):
            with app.test_client() as client:
                with client.session_transaction() as sessao:
                    sessao["usuario_id"] = 1
                    sessao["permissoes"] = ["financeiro_novo:visualizar"]
                    sessao["ultimo_acesso"] = 9999999999
                resposta = client.get("/financeiro-novo/")

        self.assertEqual(resposta.status_code, 200)
        self.assertIn("Base independente e vazia", resposta.get_data(as_text=True))
        sql_executado = " ".join(str(chamada.args[0]) for chamada in conexao.execute.call_args_list)
        self.assertNotIn("financeiro2_", sql_executado.lower())

    def test_cadastros_exigem_permissoes_proprias_para_escrita(self):
        app = create_app()
        app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        with app.test_client() as client:
            with client.session_transaction() as sessao:
                sessao["usuario_id"] = 1
                sessao["permissoes"] = ["financeiro_novo:visualizar"]
                sessao["ultimo_acesso"] = 9999999999
            resposta = client.post(
                "/financeiro-novo/cadastros/moedas/novo",
                data={"codigo": "BRL", "nome": "Real", "simbolo": "R$", "casas_decimais": "2"},
            )
        self.assertEqual(resposta.status_code, 403)

    def test_migration_cria_tabelas_relacionais_sem_exclusao_fisica(self):
        migration = (self.raiz / "migrations" / "004_financeiro_novo_cadastros.sql").read_text(encoding="utf-8")
        for tabela in ("pessoas", "centros_custo", "categorias", "moedas", "contas"):
            self.assertIn(f"financeiro3_{tabela}", migration)
        cadastros = (self.raiz / "routes" / "financeiro_novo" / "cadastros.py").read_text(encoding="utf-8")
        self.assertNotIn("DELETE FROM", cadastros.upper())

    def test_tela_de_cadastros_inicia_vazia_e_renderiza(self):
        resultado_registros = MagicMock()
        resultado_registros.mappings.return_value.all.return_value = []
        resultado_moedas = MagicMock()
        resultado_moedas.mappings.return_value.all.return_value = []
        conexao = MagicMock()
        conexao.execute.side_effect = [resultado_registros, resultado_moedas]
        contexto = MagicMock()
        contexto.__enter__.return_value = conexao
        engine = MagicMock()
        engine.connect.return_value = contexto

        app = create_app()
        app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        with patch("routes.financeiro_novo.cadastros.get_engine", return_value=engine):
            with app.test_client() as client:
                with client.session_transaction() as sessao:
                    sessao["usuario_id"] = 1
                    sessao["permissoes"] = ["financeiro_novo:visualizar"]
                    sessao["ultimo_acesso"] = 9999999999
                resposta = client.get("/financeiro-novo/cadastros?tipo=pessoas")

        self.assertEqual(resposta.status_code, 200)
        pagina = resposta.get_data(as_text=True)
        self.assertIn("Este cadastro começa vazio", pagina)
        self.assertNotIn("Novo cadastro", pagina)

    def test_despesas_usam_tabelas_novas_e_total_calculado_no_banco(self):
        migration = (self.raiz / "migrations" / "005_financeiro_novo_despesas.sql").read_text(encoding="utf-8")
        for tabela in ("despesas", "despesa_itens", "despesa_decisoes", "despesa_pagamentos"):
            self.assertIn(f"financeiro3_{tabela}", migration)
        self.assertIn("GENERATED ALWAYS AS", migration)
        self.assertIn("financeiro3_atualizar_total_despesa", migration)
        self.assertIn("NUMERIC(18, 2)", migration)
        self.assertNotIn("FLOAT", migration.upper())

    def test_permissao_de_edicao_nao_autoriza_aprovar_ou_pagar(self):
        app = create_app()
        app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        with app.test_client() as client:
            with client.session_transaction() as sessao:
                sessao["usuario_id"] = 1
                sessao["permissoes"] = ["financeiro_novo:visualizar", "financeiro_novo:editar"]
                sessao["ultimo_acesso"] = 9999999999
            aprovacao = client.post("/financeiro-novo/despesas/1/aprovar")
            pagamento = client.post("/financeiro-novo/despesas/1/pagar")
        self.assertEqual(aprovacao.status_code, 403)
        self.assertEqual(pagamento.status_code, 403)

    def test_om_rd_e_acertos_tem_modelo_independente_e_total_derivado(self):
        migration = (self.raiz / "migrations" / "006_financeiro_novo_om_rd.sql").read_text(encoding="utf-8")
        for tabela in ("oms", "om_decisoes", "rds", "rd_itens", "rd_decisoes", "rd_acertos"):
            self.assertIn(f"financeiro3_{tabela}", migration)
        self.assertIn("om_id BIGINT NOT NULL UNIQUE", migration)
        self.assertIn("financeiro3_atualizar_total_rd", migration)
        self.assertIn("'REEMBOLSO','DEVOLUCAO'", migration)

    def test_edicao_nao_autoriza_decidir_om_ou_rd(self):
        app = create_app()
        app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        with app.test_client() as client:
            with client.session_transaction() as sessao:
                sessao["usuario_id"] = 1
                sessao["permissoes"] = ["financeiro_novo:visualizar", "financeiro_novo:editar"]
                sessao["ultimo_acesso"] = 9999999999
            respostas = (
                client.post("/financeiro-novo/oms/1/aprovar"),
                client.post("/financeiro-novo/rds/1/aprovar"),
                client.post("/financeiro-novo/rds/1/liquidar"),
            )
        self.assertTrue(all(resposta.status_code == 403 for resposta in respostas))

    def test_notas_relatorios_e_conciliacao_usam_modelo_novo(self):
        migration = (self.raiz / "migrations" / "007_financeiro_novo_nd_relatorios.sql").read_text(encoding="utf-8")
        for tabela in ("clientes", "notas_debito", "nd_itens", "nd_decisoes", "nd_recebimentos", "conciliacoes"):
            self.assertIn(f"financeiro3_{tabela}", migration)
        self.assertIn("financeiro3_atualizar_total_nd", migration)
        self.assertIn("UNIQUE (origem_tipo,origem_id)", migration)

    def test_reembolsos_e_previsao_usam_apenas_modelo_novo(self):
        migration = (self.raiz / "migrations" / "008_financeiro_novo_reembolsos_previsao.sql").read_text(encoding="utf-8")
        for tabela in ("reembolsos", "reembolso_itens", "reembolso_decisoes", "reembolso_pagamentos"):
            self.assertIn(f"financeiro3_{tabela}", migration)
        self.assertIn("financeiro3_atualizar_total_reembolso", migration)
        self.assertNotIn("FLOAT", migration.upper())
        previsao = (self.raiz / "routes" / "financeiro_novo" / "previsao.py").read_text(encoding="utf-8")
        for origem in ("financeiro3_despesas", "financeiro3_reembolsos", "financeiro3_rd_acertos", "financeiro3_notas_debito"):
            self.assertIn(origem, previsao)
        relatorios = (self.raiz / "routes" / "financeiro_novo" / "relatorios.py").read_text(encoding="utf-8")
        self.assertIn("PAGAMENTO_REEMBOLSO", relatorios)
        self.assertNotIn("ADIANTAMENTO_OM", relatorios)

    def test_om_tem_numero_matricula_linhas_e_total_derivado(self):
        migration = (self.raiz / "migrations" / "009_financeiro_novo_om_linhas.sql").read_text(encoding="utf-8")
        for campo in ("numero_om", "matricula_favorecido", "valor_total"):
            self.assertIn(campo, migration)
        self.assertIn("financeiro3_om_itens", migration)
        self.assertIn("financeiro3_atualizar_total_om", migration)
        self.assertIn("removido_em", migration)
        self.assertNotIn("FLOAT", migration.upper())

    def test_formulario_om_nao_exibe_campos_descontinuados(self):
        campos = (self.raiz / "templates" / "financeiro_novo" / "_om_campos.html").read_text(encoding="utf-8")
        self.assertIn('name="numero_om"', campos)
        self.assertIn('name="matricula_favorecido"', campos)
        for nome in ("objetivo", "valor_adiantamento", "origem", "destino", "data_inicio", "data_fim"):
            self.assertNotIn(f'name="{nome}"', campos)
        detalhe = (self.raiz / "templates" / "financeiro_novo" / "om_detalhe.html").read_text(encoding="utf-8")
        self.assertIn("om_item_novo", detalhe)
        self.assertNotIn("justificativa_sem_comprovante", detalhe)
        self.assertNotIn("Recibo ou justificativa", detalhe)
        self.assertIn("<th>Recibo</th>", detalhe)

    def test_linha_om_pode_ser_salva_sem_recibo(self):
        app = create_app()
        with app.test_request_context("/", method="POST", data={
            "data_despesa": "2026-08-11",
            "centro_custo_id": "2",
            "categoria_id": "7",
            "descricao": "Despesa sem recibo",
            "valor": "50,00",
            "arquivo": (io.BytesIO(b""), ""),
        }):
            linhas = _linhas_om_formulario()
        self.assertEqual(len(linhas), 1)
        self.assertFalse(linhas[0]["arquivo"].filename)

    def test_om_em_lote_e_rd_independente(self):
        migration = (self.raiz / "migrations" / "010_financeiro_novo_om_lote_rd_independente.sql").read_text(encoding="utf-8")
        self.assertIn("centro_custo_id", migration)
        self.assertIn("numero_rd", migration)
        self.assertIn("matricula_responsavel", migration)
        self.assertIn("ALTER COLUMN om_id DROP NOT NULL", migration)
        detalhe = (self.raiz / "templates" / "financeiro_novo" / "om_detalhe.html").read_text(encoding="utf-8")
        self.assertIn("om-batch-form", detalhe)
        self.assertIn("Adicionar linha", detalhe)
        self.assertIn("om_verificar_duplicidades", detalhe)
        self.assertNotIn('name="fornecedor_id"', detalhe)
        self.assertNotIn('name="numero_documento"', detalhe)
        rd_form = (self.raiz / "templates" / "financeiro_novo" / "rd_form.html").read_text(encoding="utf-8")
        self.assertIn('name="numero_rd"', rd_form)
        self.assertIn('name="matricula_responsavel"', rd_form)

    def test_alerta_duplicidade_mantem_data_valor_e_exibe_detalhes(self):
        rotas = (self.raiz / "routes" / "financeiro_novo" / "missoes.py").read_text(encoding="utf-8")
        self.assertGreaterEqual(rotas.count("i.data_despesa=:data AND i.valor=:valor"), 3)
        for campo in ("descricao", "categoria", "centro", "moeda", "comprovante_url", "documento_url"):
            self.assertIn(f'"{campo}"', rotas)
        detalhe = (self.raiz / "templates" / "financeiro_novo" / "om_detalhe.html").read_text(encoding="utf-8")
        for coluna in ("Origem", "Descrição", "Categoria", "Centro de custo", "Comprovante"):
            self.assertIn(coluna, detalhe)
        self.assertIn('id="duplicate-dialog"', detalhe)
        self.assertIn("Salvar mesmo assim", detalhe)
        self.assertNotIn("URLSearchParams", detalhe)
        self.assertIn("method:'POST'", detalhe)
        self.assertIn("JSON.stringify({linhas})", detalhe)
        self.assertIn("'Content-Type':'application/json'", detalhe)
        self.assertIn("cache:'no-store'", detalhe)
        self.assertIn("credentials:'same-origin'", detalhe)
        self.assertIn('@bp.post("/oms/<int:om_id>/verificar-duplicidades")', rotas)
        self.assertIn("request.get_json(silent=True)", rotas)
        self.assertIn("jsonb_to_recordset", rotas)
        self.assertNotIn("Analisei os possíveis lançamentos duplicados", detalhe)
        self.assertNotIn('type="checkbox" name="confirmar_duplicidade"', detalhe)

    def test_verificacao_duplicidades_funciona_com_csrf_ativado(self):
        resultado_om = MagicMock()
        resultado_om.mappings.return_value.first.return_value = {"id": 1}
        resultado_duplicidades = MagicMock()
        resultado_duplicidades.mappings.return_value.all.return_value = []
        conexao = MagicMock()
        conexao.execute.side_effect = [resultado_om, resultado_duplicidades]
        contexto = MagicMock()
        contexto.__enter__.return_value = conexao
        engine = MagicMock()
        engine.connect.return_value = contexto

        app = create_app()
        app.config.update(TESTING=True)
        with patch("routes.financeiro_novo.missoes.get_engine", return_value=engine):
            with app.test_client() as client:
                token = re.search(r'<meta name="csrf-token" content="([^"]+)"',
                                  client.get("/auth/login").get_data(as_text=True)).group(1)
                with client.session_transaction() as sessao:
                    sessao["usuario_id"] = 1
                    sessao["permissoes"] = ["financeiro_novo:editar"]
                    sessao["ultimo_acesso"] = 9999999999
                resposta = client.post(
                    "/financeiro-novo/oms/1/verificar-duplicidades",
                    json={"linhas": [{"data": "2026-08-06", "valor": "123,45"}]},
                    headers={"X-CSRFToken": token},
                )

        self.assertEqual(resposta.status_code, 200)
        self.assertEqual(resposta.get_json(), {"duplicidades": []})

    def test_verificacao_duplicidades_aceita_lote_grande_em_json(self):
        resultado_om = MagicMock()
        resultado_om.mappings.return_value.first.return_value = {"id": 1}
        resultado_duplicidades = MagicMock()
        resultado_duplicidades.mappings.return_value.all.return_value = []
        conexao = MagicMock()
        conexao.execute.side_effect = [resultado_om, resultado_duplicidades]
        contexto = MagicMock(); contexto.__enter__.return_value = conexao
        engine = MagicMock(); engine.connect.return_value = contexto
        linhas = [
            {"data": "2026-08-06", "valor": f"{indice},00"}
            for indice in range(1, 249)
        ]

        app = create_app(); app.config.update(TESTING=True)
        with patch("routes.financeiro_novo.missoes.get_engine", return_value=engine):
            with app.test_client() as client:
                token = re.search(r'<meta name="csrf-token" content="([^"]+)"',
                                  client.get("/auth/login").get_data(as_text=True)).group(1)
                with client.session_transaction() as sessao:
                    sessao["usuario_id"] = 1
                    sessao["permissoes"] = ["financeiro_novo:editar"]
                    sessao["ultimo_acesso"] = 9999999999
                resposta = client.post(
                    "/financeiro-novo/oms/1/verificar-duplicidades", json={"linhas": linhas},
                    headers={"X-CSRFToken": token},
                )

        self.assertEqual(resposta.status_code, 200)
        self.assertEqual(resposta.get_json(), {"duplicidades": []})
        self.assertEqual(conexao.execute.call_count, 2)

    def test_formulario_suporta_mil_linhas_de_om_sem_liberar_limite(self):
        self.assertGreaterEqual(ProdRequest.max_form_parts, 6002)
        self.assertLessEqual(ProdRequest.max_form_parts, 7000)
        self.assertEqual(ProdRequest.max_form_memory_size, 2 * 1024 * 1024)

        app = create_app()
        app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        app.add_url_rule(
            "/__teste_partes", "teste_partes",
            lambda: str(len(request.form.getlist("campo"))), methods=["POST"],
        )
        dados = MultiDict([("campo", str(indice)) for indice in range(1490)])
        dados.add("arquivo", (io.BytesIO(b"teste"), "teste.txt"))
        with app.test_client() as client:
            with client.session_transaction() as sessao:
                sessao["usuario_id"] = 1
                sessao["ultimo_acesso"] = 9999999999
            resposta = client.post("/__teste_partes", data=dados)
        self.assertEqual(resposta.status_code, 200)
        self.assertEqual(resposta.get_data(as_text=True), "1490")

    def test_erro_de_lote_grande_tem_mensagem_compreensivel(self):
        app = create_app()
        app.config.update(TESTING=True, WTF_CSRF_ENABLED=False, MAX_CONTENT_LENGTH=10)
        with app.test_client() as client:
            resposta = client.post("/auth/login", data={"campo": "x" * 100})
        self.assertEqual(resposta.status_code, 413)
        self.assertIn("excedem o limite permitido", resposta.get_data(as_text=True))

    def test_rd_aplica_a_mesma_verificacao_de_duplicidades(self):
        rotas = (self.raiz / "routes" / "financeiro_novo" / "missoes.py").read_text(encoding="utf-8")
        detalhe = (self.raiz / "templates" / "financeiro_novo" / "rd_detalhe.html").read_text(encoding="utf-8")
        self.assertIn('@bp.get("/rds/<int:rd_id>/verificar-duplicidades")', rotas)
        self.assertIn('request.form.get("forcar_salvamento") != "1"', rotas)
        self.assertIn('id="rd-duplicate-dialog"', detalhe)
        self.assertIn("rd_verificar_duplicidades", detalhe)
        self.assertIn("Salvar mesmo assim", detalhe)
        self.assertIn("mesma data e o mesmo valor", detalhe)

        resultado_rd = MagicMock()
        resultado_rd.mappings.return_value.first.return_value = {"id": 1}
        resultado_duplicidades = MagicMock()
        resultado_duplicidades.mappings.return_value.all.return_value = []
        conexao = MagicMock()
        conexao.execute.side_effect = [resultado_rd, resultado_duplicidades]
        contexto = MagicMock()
        contexto.__enter__.return_value = conexao
        engine = MagicMock()
        engine.connect.return_value = contexto

        app = create_app()
        app.config.update(TESTING=True)
        with patch("routes.financeiro_novo.missoes.get_engine", return_value=engine):
            with app.test_client() as client:
                with client.session_transaction() as sessao:
                    sessao["usuario_id"] = 1
                    sessao["permissoes"] = ["financeiro_novo:editar"]
                    sessao["ultimo_acesso"] = 9999999999
                resposta = client.get(
                    "/financeiro-novo/rds/1/verificar-duplicidades"
                    "?data=2026-08-06&valor=123%2C45"
                )

        self.assertEqual(resposta.status_code, 200)
        self.assertEqual(resposta.get_json(), {"duplicidades": []})

    def test_fluxo_documental_novo_elimina_conversao_om_rd(self):
        migration = (self.raiz / "migrations" / "011_financeiro_novo_fluxo_documentos_previsao.sql").read_text(encoding="utf-8")
        rotas = (self.raiz / "routes" / "financeiro_novo" / "missoes.py").read_text(encoding="utf-8")
        self.assertIn("financeiro3_rds_independente_ck", migration)
        self.assertIn("CHECK (om_id IS NULL)", migration)
        self.assertNotIn("def rd_criar", rotas)
        self.assertIn("financeiro3_om_pagamentos", migration)
        self.assertIn("financeiro3_rd_pagamentos", migration)

    def test_previsao_distingue_data_prevista_e_data_realizada(self):
        previsao = (self.raiz / "routes" / "financeiro_novo" / "previsao.py").read_text(encoding="utf-8")
        for trecho in ("pg.data_prevista_pagamento", "pg.data_pagamento", "'REALIZADO'", "'PREVISTO'"):
            self.assertIn(trecho, previsao)
        self.assertIn("NOT d.paga_na_origem", previsao)
        self.assertIn("r.forma_liquidacao='DIRETO'", previsao)

    def test_nota_debito_nasce_de_linha_de_despesa_e_nao_de_rd(self):
        migration = (self.raiz / "migrations" / "011_financeiro_novo_fluxo_documentos_previsao.sql").read_text(encoding="utf-8")
        rotas = (self.raiz / "routes" / "financeiro_novo" / "notas_debito.py").read_text(encoding="utf-8")
        self.assertIn("despesa_item_id", migration)
        self.assertIn("CHECK (rd_id IS NULL)", migration)
        self.assertIn('origem not in {"MANUAL","DESPESA_ITEM"}', rotas)
        self.assertNotIn('origem == "RD"', rotas)

    def test_despesas_sao_separadas_por_empresa_sem_reclassificar_dados_atuais(self):
        migration = (self.raiz / "migrations" / "012_financeiro_novo_empresas.sql").read_text(encoding="utf-8")
        despesas = (self.raiz / "routes" / "financeiro_novo" / "despesas.py").read_text(encoding="utf-8")
        pagina = (self.raiz / "templates" / "financeiro_novo" / "despesas.html").read_text(encoding="utf-8")
        self.assertGreaterEqual(migration.count("DEFAULT 'MATISA'"), 2)
        for empresa in ("MATISA", "PRUMO", "PRUMAT"):
            self.assertIn(empresa, migration)
        self.assertIn("d.empresa = :empresa", despesas)
        self.assertIn("'MATISA')", despesas)
        self.assertIn("company-tabs", pagina)
        self.assertIn("Nova despesa {{ empresa|title }}", pagina)

    def test_despesa_sem_favorecido_tipifica_parametro_nulo_no_postgresql(self):
        despesas = (self.raiz / "routes" / "financeiro_novo" / "despesas.py").read_text(encoding="utf-8")
        self.assertIn("CAST(:favorecido_id AS BIGINT) IS NULL", despesas)
        self.assertNotIn("(:favorecido_id IS NULL", despesas)

    def test_transicoes_nao_reutilizam_status_com_tipos_ambiguos_no_postgresql(self):
        arquivos = [
            self.raiz / "routes" / "financeiro_novo" / "despesas.py",
            self.raiz / "routes" / "financeiro_novo" / "notas_debito.py",
            self.raiz / "routes" / "financeiro_novo" / "reembolsos.py",
        ]
        conteudo = "\n".join(arquivo.read_text(encoding="utf-8") for arquivo in arquivos)
        self.assertNotIn("CASE WHEN :status", conteudo)
        for parametro in (":aprovada", ":quitada", ":emitida", ":aprovado"):
            self.assertIn(parametro, conteudo)

    def test_pagamento_de_despesa_nao_exige_conta_nem_motivo_para_cancelar(self):
        migration = (self.raiz / "migrations" / "013_financeiro_novo_pagamento_despesa_sem_conta.sql").read_text(encoding="utf-8")
        rotas = (self.raiz / "routes" / "financeiro_novo" / "despesas.py").read_text(encoding="utf-8")
        pagina = (self.raiz / "templates" / "financeiro_novo" / "despesa_detalhe.html").read_text(encoding="utf-8")
        self.assertIn("ALTER COLUMN conta_id DROP NOT NULL", migration)
        self.assertNotIn('request.form.get("conta_id")', rotas)
        self.assertNotIn("despesa_id, conta_id, data_pagamento", rotas)
        self.assertIn("LEFT JOIN financeiro3_contas", rotas)
        self.assertNotIn("Conta na mesma moeda", pagina)
        self.assertNotIn("Motivo obrigatório", pagina)
        self.assertNotIn('name="motivo"', pagina)

    def test_listagem_de_despesas_exibe_numero_do_documento(self):
        pagina = (self.raiz / "templates" / "financeiro_novo" / "despesas.html").read_text(encoding="utf-8")
        self.assertIn("<th>Documento</th>", pagina)
        self.assertIn("item.numero_documento or '—'", pagina)
        self.assertIn('colspan="8"', pagina)

    def test_nota_debito_aceita_somente_despesa_da_mesma_empresa(self):
        migration = (self.raiz / "migrations" / "012_financeiro_novo_empresas.sql").read_text(encoding="utf-8")
        rotas = (self.raiz / "routes" / "financeiro_novo" / "notas_debito.py").read_text(encoding="utf-8")
        self.assertIn("financeiro3_validar_empresa_nd_item", migration)
        self.assertIn("d.empresa=:empresa", rotas)
        self.assertIn('origem_item["empresa"] != nd["empresa"]', rotas)
        self.assertIn("mesma empresa", rotas)

    def test_empresas_financeiras_validas_sao_fechadas(self):
        self.assertEqual(empresa_valida(None), "MATISA")
        self.assertEqual(empresa_valida("prumo"), "PRUMO")
        self.assertEqual(nome_empresa("PRUMAT"), "Prumat")
        with self.assertRaises(ValorInvalido):
            empresa_valida("OUTRA")

    def test_reembolso_vinculado_nao_duplica_pagamento(self):
        migration = (self.raiz / "migrations" / "011_financeiro_novo_fluxo_documentos_previsao.sql").read_text(encoding="utf-8")
        despesas = (self.raiz / "routes" / "financeiro_novo" / "despesas.py").read_text(encoding="utf-8")
        self.assertIn("forma_liquidacao", migration)
        self.assertIn("paga_na_origem", migration)
        self.assertIn("incluindo reembolsos vinculados", despesas)
        self.assertIn("sem duplicar o pagamento", despesas)

    def test_excel_da_om_e_convertido_em_linhas_editaveis_sem_salvar(self):
        arquivo = io.BytesIO()
        pasta = Workbook()
        aba = pasta.active
        aba.append(["Data", "Centro de custo", "Categoria", "Descrição", "Valor"])
        aba.append(["10/08/2026", "02 - PRUMAT", "MAT - Material", "Dormentes", 1234.56])
        pasta.save(arquivo)

        linhas = _ler_linhas_om_excel(
            arquivo.getvalue(),
            [{"id": 2, "codigo": "02", "nome": "PRUMAT"}],
            [{"id": 7, "codigo": "MAT", "nome": "Material"}],
        )

        self.assertEqual(linhas, [{
            "data": "2026-08-10", "centro_custo_id": 2, "categoria_id": 7,
            "descricao": "Dormentes", "valor": "1234,56",
        }])

    def test_excel_da_om_rejeita_cadastro_desconhecido(self):
        arquivo = io.BytesIO()
        pasta = Workbook()
        aba = pasta.active
        aba.append(["Data", "Centro de custo", "Categoria", "Descrição", "Valor"])
        aba.append(["10/08/2026", "Centro inexistente", "MAT", "Teste", 10])
        pasta.save(arquivo)
        with self.assertRaisesRegex(ValorInvalido, "Centro inexistente"):
            _ler_linhas_om_excel(
                arquivo.getvalue(),
                [{"id": 2, "codigo": "02", "nome": "PRUMAT"}],
                [{"id": 7, "codigo": "MAT", "nome": "Material"}],
            )

    def test_upload_excel_da_om_retorna_previa_sem_gravar(self):
        arquivo = io.BytesIO()
        pasta = Workbook()
        aba = pasta.active
        aba.append(["Data", "Centro de custo", "Categoria", "Descrição", "Valor"])
        aba.append(["10/08/2026", "2", "MAT", "Teste", 50])
        pasta.save(arquivo)

        resultado_om = MagicMock(); resultado_om.mappings.return_value.first.return_value = {"id": 1, "status": "RASCUNHO"}
        resultado_centros = MagicMock(); resultado_centros.mappings.return_value.all.return_value = [{"id": 2, "codigo": "02", "nome": "PRUMAT"}]
        resultado_categorias = MagicMock(); resultado_categorias.mappings.return_value.all.return_value = [{"id": 7, "codigo": "MAT", "nome": "Material"}]
        conexao = MagicMock(); conexao.execute.side_effect = [resultado_om, resultado_centros, resultado_categorias]
        contexto = MagicMock(); contexto.__enter__.return_value = conexao
        engine = MagicMock(); engine.connect.return_value = contexto

        app = create_app(); app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        with patch("routes.financeiro_novo.missoes.get_engine", return_value=engine):
            with app.test_client() as client:
                with client.session_transaction() as sessao:
                    sessao["usuario_id"] = 1
                    sessao["permissoes"] = ["financeiro_novo:editar"]
                    sessao["ultimo_acesso"] = 9999999999
                resposta = client.post("/financeiro-novo/oms/1/itens/carregar-excel",
                    data={"arquivo_excel": (io.BytesIO(arquivo.getvalue()), "linhas.xlsx")})

        self.assertEqual(resposta.status_code, 200)
        self.assertEqual(resposta.get_json()["quantidade"], 1)
        sql_executado = " ".join(str(chamada.args[0]) for chamada in conexao.execute.call_args_list)
        self.assertNotIn("INSERT", sql_executado.upper())

    def test_botao_carregar_excel_fica_entre_adicionar_e_salvar(self):
        detalhe = (self.raiz / "templates" / "financeiro_novo" / "om_detalhe.html").read_text(encoding="utf-8")
        self.assertLess(detalhe.index("+ Adicionar linha"), detalhe.index("Carregar Excel"))
        self.assertLess(detalhe.index("Carregar Excel"), detalhe.index("Salvar todas as linhas"))
        self.assertIn("om_itens_carregar_excel", detalhe)

    def test_pagamento_om_tem_botao_de_salvar_destacado(self):
        detalhe = (self.raiz / "templates" / "financeiro_novo" / "om_detalhe.html").read_text(encoding="utf-8")
        self.assertIn('type="submit" class="btn btn-primary full">Registrar adiantamento ou quitação', detalhe)
        self.assertIn('name="data_pagamento"', detalhe)
        self.assertNotIn('name="data_prevista_pagamento"', detalhe)
        self.assertIn("om_pagamento_programar", detalhe)

    def test_resumo_om_mostra_diferenca_entre_total_e_pago(self):
        detalhe = (self.raiz / "templates" / "financeiro_novo" / "om_detalhe.html").read_text(encoding="utf-8")
        rotas = (self.raiz / "routes" / "financeiro_novo" / "missoes.py").read_text(encoding="utf-8")
        self.assertIn("<span>Diferença</span>", detalhe)
        self.assertIn("format(diferenca)", detalhe)
        self.assertNotIn("<span>Previsto</span>", detalhe)
        self.assertIn('diferenca = om["valor_total"] + om["valor_reembolsos"] - valor_pago', rotas)

    def test_listagem_om_mostra_diferenca_entre_despesas_e_pago(self):
        listagem = (self.raiz / "templates" / "financeiro_novo" / "missoes.html").read_text(encoding="utf-8")
        self.assertIn('<th class="right">Diferença</th>', listagem)
        self.assertIn("format(om.valor_total+om.valor_reembolsos-om.valor_pago)", listagem)
        self.assertNotIn('<th class="right">Previsto</th>', listagem)

    def test_railway_volume_define_automaticamente_diretorio_de_upload(self):
        with tempfile.TemporaryDirectory() as volume:
            with patch.dict(os.environ, {"RAILWAY_VOLUME_MOUNT_PATH": volume, "UPLOAD_ROOT": ""}):
                app = create_app()
        self.assertEqual(app.config["UPLOAD_ROOT"], os.path.abspath(os.path.join(volume, "uploads")))

    def test_om_permite_substituir_recibo_sem_alterar_linha(self):
        detalhe = (self.raiz / "templates" / "financeiro_novo" / "om_detalhe.html").read_text(encoding="utf-8")
        self.assertIn("om_item_recibo_substituir", detalhe)
        self.assertIn("Substituir", detalhe)
        rotas = (self.raiz / "routes" / "financeiro_novo" / "missoes.py").read_text(encoding="utf-8")
        self.assertIn("RECIBO_SUBSTITUIDO", rotas)

    def test_recibo_ausente_redireciona_com_aviso_em_vez_de_404(self):
        arquivo_id = "47d668e8-0877-4fde-8230-5c20a653d5a4"
        resultado = MagicMock()
        resultado.mappings.return_value.first.return_value = {
            "object_key": f"financeiro_novo/47/{arquivo_id}.pdf", "nome_original": "recibo.pdf",
        }
        conexao = MagicMock(); conexao.execute.return_value = resultado
        contexto = MagicMock(); contexto.__enter__.return_value = conexao
        engine = MagicMock(); engine.connect.return_value = contexto
        with tempfile.TemporaryDirectory() as upload_root:
            app = create_app(); app.config.update(TESTING=True, UPLOAD_ROOT=upload_root)
            with patch("routes.financeiro_novo.missoes.get_engine", return_value=engine):
                with app.test_client() as client:
                    with client.session_transaction() as sessao:
                        sessao["usuario_id"] = 1
                        sessao["permissoes"] = ["financeiro_novo:visualizar"]
                        sessao["ultimo_acesso"] = 9999999999
                    resposta = client.get(f"/financeiro-novo/oms/1/itens/1/anexos/{arquivo_id}")
        self.assertEqual(resposta.status_code, 302)
        self.assertIn("/financeiro-novo/oms/1", resposta.location)

    def test_pagamento_om_e_registrado_diretamente_com_data_real(self):
        resultado_om = MagicMock()
        resultado_om.mappings.return_value.first.return_value = {"id": 1, "status": "RASCUNHO"}
        resultado_pagamento = MagicMock()
        resultado_pagamento.mappings.return_value.one.return_value = {
            "id": 9, "om_id": 1, "tipo": "ADIANTAMENTO", "status": "PAGO",
            "data_pagamento": "2026-08-10", "valor": Decimal("100.00"),
        }
        conexao = MagicMock()
        conexao.execute.side_effect = [resultado_om, resultado_pagamento, MagicMock()]
        contexto = MagicMock(); contexto.__enter__.return_value = conexao
        engine = MagicMock(); engine.begin.return_value = contexto

        app = create_app(); app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        with patch("routes.financeiro_novo.missoes.get_engine", return_value=engine):
            with app.test_client() as client:
                with client.session_transaction() as sessao:
                    sessao["usuario_id"] = 1
                    sessao["permissoes"] = ["financeiro_novo:pagar"]
                    sessao["ultimo_acesso"] = 9999999999
                resposta = client.post("/financeiro-novo/oms/1/pagamentos", data={
                    "tipo_pagamento": "ADIANTAMENTO", "data_pagamento": "2026-08-10",
                    "valor": "100,00", "observacoes": "Teste",
                })

        self.assertEqual(resposta.status_code, 302)
        sql_executado = " ".join(str(chamada.args[0]) for chamada in conexao.execute.call_args_list)
        self.assertIn("'PAGO'", sql_executado)
        self.assertIn("data_pagamento", sql_executado)

    def test_edicao_nao_autoriza_registrar_pagamento_da_om(self):
        app = create_app(); app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        with app.test_client() as client:
            with client.session_transaction() as sessao:
                sessao["usuario_id"] = 1
                sessao["permissoes"] = ["financeiro_novo:editar"]
                sessao["ultimo_acesso"] = 9999999999
            resposta = client.post("/financeiro-novo/oms/1/pagamentos")
        self.assertEqual(resposta.status_code, 403)

    def test_reembolso_separa_edicao_aprovacao_e_pagamento(self):
        app = create_app(); app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        with app.test_client() as client:
            with client.session_transaction() as sessao:
                sessao["usuario_id"] = 1
                sessao["permissoes"] = ["financeiro_novo:visualizar", "financeiro_novo:editar"]
                sessao["ultimo_acesso"] = 9999999999
            aprovacao = client.post("/financeiro-novo/reembolsos/1/aprovar")
            pagamento = client.post("/financeiro-novo/reembolsos/1/pagar")
        self.assertEqual(aprovacao.status_code, 403)
        self.assertEqual(pagamento.status_code, 403)

    def test_conciliacao_exige_permissao_administrativa_do_modulo(self):
        app = create_app(); app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        with app.test_client() as client:
            with client.session_transaction() as sessao:
                sessao["usuario_id"] = 1
                sessao["permissoes"] = ["financeiro_novo:visualizar", "financeiro_novo:editar"]
                sessao["ultimo_acesso"] = 9999999999
            resposta = client.get("/financeiro-novo/conciliacao")
        self.assertEqual(resposta.status_code, 403)

    def test_homologacao_exige_permissao_administrativa_do_modulo(self):
        app = create_app(); app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        with app.test_client() as client:
            with client.session_transaction() as sessao:
                sessao["usuario_id"] = 1
                sessao["permissoes"] = ["financeiro_novo:visualizar"]
                sessao["ultimo_acesso"] = 9999999999
            resposta = client.get("/financeiro-novo/homologacao")
        self.assertEqual(resposta.status_code, 403)

    def test_diagnostico_de_armazenamento_nao_cria_diretorios(self):
        app = create_app(); app.config.update(TESTING=True, UPLOAD_ROOT=str(self.raiz / "diretorio-que-nao-existe"))
        with app.app_context():
            diagnostico = diagnosticar_armazenamento()
        self.assertFalse(diagnostico["existe"])
        self.assertFalse((self.raiz / "diretorio-que-nao-existe").exists())


class FinanceiroNovoCadastrosTests(unittest.TestCase):
    def test_pessoa_exige_papel_e_normaliza_documento(self):
        dados, erros = _normalizar(
            TIPOS["pessoas"],
            {
                "tipo_pessoa": "JURIDICA",
                "nome_razao": "Fornecedor Teste",
                "documento": "12.345.678/0001-90",
                "fornecedor": "on",
            },
        )
        self.assertEqual(erros, [])
        self.assertEqual(dados["documento"], "12345678000190")
        self.assertTrue(dados["fornecedor"])

        _, erros_sem_papel = _normalizar(
            TIPOS["pessoas"],
            {"tipo_pessoa": "FISICA", "nome_razao": "Favorecido"},
        )
        self.assertTrue(any("fornecedor" in erro.lower() for erro in erros_sem_papel))

    def test_moeda_exige_codigo_iso_de_tres_letras(self):
        dados, erros = _normalizar(
            TIPOS["moedas"],
            {"codigo": "br", "nome": "Real", "simbolo": "R$", "casas_decimais": "2"},
        )
        self.assertEqual(dados["codigo"], "BR")
        self.assertTrue(any("3 letras" in erro for erro in erros))

    def test_tipo_de_conta_fora_da_lista_e_rejeitado(self):
        _, erros = _normalizar(
            TIPOS["contas"],
            {"nome": "Conta", "tipo": "CRIPTO", "moeda_id": "1"},
        )
        self.assertTrue(any("Tipo é inválido" in erro for erro in erros))


class FinanceiroNovoValoresTests(unittest.TestCase):
    def test_decimal_aceita_formato_brasileiro_sem_usar_float(self):
        self.assertEqual(decimal_br("1.234,56"), Decimal("1234.56"))
        self.assertEqual(decimal_br("10,999", casas=2), Decimal("11.00"))

    def test_decimal_bloqueia_zero_em_pagamentos_e_itens(self):
        with self.assertRaises(ValorInvalido):
            decimal_br("0,00", positivo=True)


class FinanceiroNovoAnexosTests(unittest.TestCase):
    def test_foto_e_convertida_para_pdf_canonico(self):
        origem = io.BytesIO()
        Image.new("RGB", (1600, 1000), "white").save(origem, format="BMP")
        origem.seek(0)
        arquivo = FileStorage(stream=origem, filename="recibo.bmp", content_type="image/bmp")

        anexo = normalizar_anexo(arquivo)

        self.assertTrue(anexo.conteudo.startswith(b"%PDF-"))
        self.assertEqual(anexo.mime_canonico, "application/pdf")
        self.assertEqual(anexo.paginas, 1)
        self.assertEqual(len(anexo.sha256_canonico), 64)
        self.assertLess(anexo.tamanho_canonico, anexo.tamanho_original)

    def test_nome_do_objeto_usa_pdf_em_diretorio_exclusivo(self):
        nome = nome_objeto_pdf("abcdef12-3456-7890-abcd-ef1234567890")
        self.assertEqual(
            nome,
            "financeiro_novo/ab/abcdef12-3456-7890-abcd-ef1234567890.pdf",
        )

    def test_pdf_e_regravado_e_validado(self):
        origem = io.BytesIO()
        documento = canvas.Canvas(origem)
        documento.drawString(50, 800, "Recibo de teste")
        documento.save()
        origem.seek(0)
        arquivo = FileStorage(stream=origem, filename="recibo.pdf", content_type="application/pdf")

        anexo = normalizar_anexo(arquivo)

        self.assertTrue(anexo.conteudo.startswith(b"%PDF-"))
        self.assertEqual(anexo.paginas, 1)

    def test_pdf_de_scanner_tem_imagens_internas_compactadas(self):
        imagem = Image.effect_noise((1600, 1200), 80).convert("RGB")
        imagem_png = io.BytesIO(); imagem.save(imagem_png, format="PNG")
        origem = io.BytesIO()
        documento = canvas.Canvas(origem, pagesize=(1600, 1200), pageCompression=1)
        documento.drawImage(ImageReader(io.BytesIO(imagem_png.getvalue())), 0, 0, 1600, 1200)
        documento.save(); origem.seek(0)
        arquivo = FileStorage(stream=origem, filename="scanner.pdf", content_type="application/pdf")

        anexo = normalizar_anexo(arquivo)

        self.assertTrue(anexo.compressao_aplicada)
        self.assertLess(anexo.tamanho_canonico, anexo.tamanho_original * 0.6)
        self.assertEqual(len(PdfReader(io.BytesIO(anexo.conteudo)).pages), 1)

    def test_otimizacao_nunca_aumenta_pdf_original(self):
        origem = io.BytesIO()
        documento = canvas.Canvas(origem)
        documento.drawString(50, 800, "PDF pequeno já otimizado")
        documento.save(); origem.seek(0)
        original = origem.getvalue()

        anexo = normalizar_anexo(FileStorage(
            stream=io.BytesIO(original), filename="pequeno.pdf", content_type="application/pdf"
        ))

        self.assertLessEqual(anexo.tamanho_canonico, len(original))

    def test_arquivo_canonico_identico_reutiliza_copia_fisica(self):
        origem = io.BytesIO(); Image.new("RGB", (800, 600), "white").save(origem, format="JPEG"); origem.seek(0)
        anexo = normalizar_anexo(FileStorage(
            stream=origem, filename="duplicado.jpg", content_type="image/jpeg"
        ))
        with tempfile.TemporaryDirectory() as upload_root:
            raiz = Path(upload_root)
            existente = raiz / "financeiro_novo" / "aa" / "existente.pdf"
            novo = raiz / "financeiro_novo" / "bb" / "novo.pdf"
            existente.parent.mkdir(parents=True); novo.parent.mkdir(parents=True)
            existente.write_bytes(anexo.conteudo); novo.write_bytes(anexo.conteudo)
            resultado_existente = MagicMock()
            resultado_existente.mappings.return_value.first.return_value = {
                "id": uuid.UUID("aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"),
                "object_key": "financeiro_novo/aa/existente.pdf",
            }
            resultado_vinculo = MagicMock(); resultado_vinculo.scalar.return_value = 77
            conexao = MagicMock(); conexao.execute.side_effect = [resultado_existente, resultado_vinculo]
            app = create_app(); app.config.update(TESTING=True, UPLOAD_ROOT=upload_root)
            with app.test_request_context("/"):
                vinculo = _vincular_anexo(conexao,
                    (anexo, uuid.uuid4(), "financeiro_novo/bb/novo.pdf", novo), "OM_ITEM", 1, "COMPROVANTE")

            self.assertEqual(vinculo, 77)
            self.assertFalse(novo.exists())
            sql = " ".join(str(chamada.args[0]) for chamada in conexao.execute.call_args_list)
            self.assertNotIn("INSERT INTO financeiro3_arquivos", sql)

    def test_arquivo_que_nao_e_foto_nem_pdf_e_rejeitado(self):
        arquivo = FileStorage(
            stream=io.BytesIO(b"conteudo executavel"),
            filename="recibo.exe",
            content_type="application/octet-stream",
        )
        with self.assertRaises(AnexoInvalido):
            normalizar_anexo(arquivo)


class FinanceiroNovoExportacaoOmTests(unittest.TestCase):
    def setUp(self):
        self.om = {
            "numero_om": "OM/2026-001", "solicitante": "Favorecido Teste",
            "matricula_favorecido": "12345", "centro_codigo": "02",
            "centro_nome": "PRUMAT", "status": "APROVADA", "moeda": "BRL",
            "observacoes": "Relatório de teste", "valor_reembolsos": Decimal("25.00"),
        }

    def _item(self, numero, descricao, valor, arquivo_id=None, caminho=None):
        return {
            "numero_linha": numero, "data_despesa": date(2026, 8, numero),
            "centro_codigo": f"0{numero}", "centro_nome": f"Centro {numero}",
            "categoria": "Material", "descricao": descricao, "valor": Decimal(valor),
            "arquivo_id": arquivo_id, "nome_original": f"recibo-{numero}.pdf" if arquivo_id else None,
            "caminho_recibo": caminho,
        }

    def test_excel_mantem_ordem_das_linhas_e_total_por_formula(self):
        itens = [self._item(1, "Primeira", "100.50"), self._item(2, "Segunda", "20.00")]
        arquivo = gerar_excel_om(self.om, itens, [])
        pasta = load_workbook(arquivo, data_only=False)
        resumo = pasta["Resumo OM"]

        self.assertEqual(pasta.sheetnames, ["Resumo OM", "Pagamentos"])
        self.assertEqual(resumo["A11"].value, 1)
        self.assertEqual(resumo["E11"].value, "Primeira")
        self.assertEqual(resumo["A12"].value, 2)
        self.assertEqual(resumo["E12"].value, "Segunda")
        self.assertEqual(resumo["G13"].value, "=SUM(G11:G12)")
        self.assertEqual(resumo.freeze_panes, "A11")

    def test_pdf_apresenta_resumo_e_recibos_na_mesma_ordem(self):
        with tempfile.TemporaryDirectory() as pasta:
            caminhos = []
            for numero in (1, 2):
                caminho = Path(pasta) / f"recibo-{numero}.pdf"
                pdf = canvas.Canvas(str(caminho))
                pdf.drawString(50, 750, f"RECIBO ORDEM {numero}")
                pdf.save()
                caminhos.append(caminho)
            itens = [
                self._item(1, "Primeira", "10", "a", caminhos[0]),
                self._item(2, "Segunda", "20", "b", caminhos[1]),
            ]
            leitor = PdfReader(gerar_pdf_om(self.om, itens))

        textos = [pagina.extract_text() for pagina in leitor.pages]
        self.assertIn("Resumo das despesas", textos[0])
        self.assertEqual(textos[1].strip(), "RECIBO ORDEM 1")
        self.assertEqual(textos[2].strip(), "RECIBO ORDEM 2")

    def test_pdf_sinaliza_recibo_vinculado_ausente_sem_interromper_exportacao(self):
        item = self._item(1, "Arquivo perdido", "10", "a", None)
        leitor = PdfReader(gerar_pdf_om(self.om, [item]))
        self.assertEqual(len(leitor.pages), 2)
        self.assertIn("Arquivo indisponível", leitor.pages[1].extract_text())

    def test_tela_expoe_as_duas_acoes_de_exportacao(self):
        raiz = Path(__file__).resolve().parents[1]
        detalhe = (raiz / "templates" / "financeiro_novo" / "om_detalhe.html").read_text(encoding="utf-8")
        self.assertIn("Exportar Excel", detalhe)
        self.assertIn("Exportar OM", detalhe)
        self.assertIn("om_exportar_excel", detalhe)
        self.assertIn("om_exportar_pdf", detalhe)


if __name__ == "__main__":
    unittest.main()
