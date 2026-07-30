from collections import defaultdict
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import text


CODIGOS_EH = (
    "RENOVACAO",
    "CARREGAMENTO_NOVO",
    "REMOCAO_GRAMPOS",
    "REMOCAO_GALOCHAS",
    "DESCARREGAMENTO_VELHO",
    "APLICACAO_GRAMPOS",
    "DESCARREGAMENTO_NOVO",
)

SALDOS = (
    ("PULMAO_CHAO", "Pulmão no chão"),
    ("NOVOS_VAGOES", "Novos nos vagões"),
    ("GRAMPOS_ABERTOS", "Grampos removidos em aberto"),
    ("GALOCHAS_ABERTAS", "Galochas removidas em aberto"),
    ("PREGACAO_ABERTA", "Pregação em aberto"),
    ("VELHOS_VAGOES", "Velhos nos vagões"),
)

CATEGORIAS_IMPACTO = (
    ("CLIMA", "Clima"),
    ("INTERFERENCIA_OPERACIONAL", "Interferência operacional"),
    ("EQUIPAMENTO", "Falha de equipamento"),
    ("LOGISTICA_MATERIAL", "Logística / material"),
    ("MAO_DE_OBRA", "Mão de obra"),
    ("SEGURANCA", "Segurança"),
    ("MANUTENCAO", "Manutenção"),
    ("OUTRO", "Outro"),
)

STATUS_IMPACTO = (
    ("ABERTO", "Aberto"),
    ("EM_TRATAMENTO", "Em tratamento"),
    ("RESOLVIDO", "Resolvido"),
)


def _float(valor):
    return float(valor or 0)


def _iso(valor):
    return valor.isoformat() if hasattr(valor, "isoformat") else str(valor)


def calcular_saldos(estado_anterior, producao_dia):
    """Aplica as movimentações físicas do dia e devolve uma nova posição."""
    estado = dict(estado_anterior)
    estado["PULMAO_CHAO"] += producao_dia["DESCARREGAMENTO_NOVO"] - producao_dia["CARREGAMENTO_NOVO"]
    estado["NOVOS_VAGOES"] += producao_dia["CARREGAMENTO_NOVO"] - producao_dia["RENOVACAO"]
    estado["GRAMPOS_ABERTOS"] += producao_dia["REMOCAO_GRAMPOS"] - producao_dia["RENOVACAO"]
    estado["GALOCHAS_ABERTAS"] += producao_dia["REMOCAO_GALOCHAS"] - producao_dia["RENOVACAO"]
    estado["PREGACAO_ABERTA"] += producao_dia["RENOVACAO"] - producao_dia["APLICACAO_GRAMPOS"]
    estado["VELHOS_VAGOES"] += producao_dia["RENOVACAO"] - producao_dia["DESCARREGAMENTO_VELHO"]
    return estado


def _periodo(conn, eh_id, inicio, fim):
    limites = conn.execute(
        text(
            """
            SELECT MIN(data)::date AS inicio, MAX(data)::date AS fim
            FROM (
                SELECT data FROM producao_realizada WHERE eh_id = :eh_id
                UNION ALL
                SELECT data FROM producao_planejada WHERE eh_id = :eh_id
            ) dados
            """
        ),
        {"eh_id": eh_id},
    ).mappings().first()

    hoje = date.today()
    inicio = inicio or (limites["inicio"].isoformat() if limites and limites["inicio"] else hoje.isoformat())
    fim = fim or (limites["fim"].isoformat() if limites and limites["fim"] else hoje.isoformat())
    if inicio > fim:
        inicio, fim = fim, inicio
    return inicio, fim


def _configuracao(conn, eh_id):
    row = conn.execute(
        text(
            """
            SELECT meta_total, produtividade_dia, data_inicio, data_fim_planejada, observacao
            FROM operacao_eh_config
            WHERE eh_id = :eh_id
            """
        ),
        {"eh_id": eh_id},
    ).mappings().first()
    if not row:
        return {
            "meta_total": 0.0,
            "produtividade_dia": 850.0,
            "data_inicio": None,
            "data_fim_planejada": None,
            "observacao": "",
        }
    return {
        "meta_total": _float(row["meta_total"]),
        "produtividade_dia": _float(row["produtividade_dia"]) or 850.0,
        "data_inicio": _iso(row["data_inicio"]) if row["data_inicio"] else None,
        "data_fim_planejada": _iso(row["data_fim_planejada"]) if row["data_fim_planejada"] else None,
        "observacao": row["observacao"] or "",
    }


def _saldos_iniciais(conn, eh_id, inicio):
    referencia = conn.execute(
        text(
            """
            SELECT MAX(data_referencia) AS data_referencia
            FROM operacao_saldo_inicial
            WHERE eh_id = :eh_id AND data_referencia <= :inicio
            """
        ),
        {"eh_id": eh_id, "inicio": inicio},
    ).scalar()
    valores = {codigo: 0.0 for codigo, _ in SALDOS}
    if referencia:
        rows = conn.execute(
            text(
                """
                SELECT saldo_codigo, quantidade
                FROM operacao_saldo_inicial
                WHERE eh_id = :eh_id AND data_referencia = :referencia
                """
            ),
            {"eh_id": eh_id, "referencia": referencia},
        ).mappings().all()
        for row in rows:
            valores[row["saldo_codigo"]] = _float(row["quantidade"])
    return {"data_referencia": _iso(referencia) if referencia else inicio, "valores": valores}


def _producao_diaria(conn, eh_id, inicio, fim, saldos):
    inicio_calculo = min(inicio, saldos["data_referencia"])
    acumulados_anteriores = conn.execute(
        text(
            """
            SELECT
                COALESCE((
                    SELECT SUM(p.planejado)
                    FROM producao_planejada p
                    JOIN frente_equipe f ON f.id = p.frente_id
                    WHERE p.eh_id = :eh_id AND p.data < :inicio
                      AND f.codigo = 'RENOVACAO'
                ), 0) AS planejado,
                COALESCE((
                    SELECT SUM(r.realizado)
                    FROM producao_realizada r
                    JOIN frente_equipe f ON f.id = r.frente_id
                    WHERE r.eh_id = :eh_id AND r.data < :inicio
                      AND f.codigo = 'RENOVACAO'
                ), 0) AS renovacao
            """
        ),
        {"eh_id": eh_id, "inicio": inicio},
    ).mappings().one()
    rows = conn.execute(
        text(
            """
            WITH datas AS (
                SELECT data::date
                FROM producao_realizada
                WHERE eh_id = :eh_id AND data BETWEEN :inicio_calculo AND :fim
                UNION
                SELECT data::date
                FROM producao_planejada
                WHERE eh_id = :eh_id AND data BETWEEN :inicio_calculo AND :fim
            ),
            realizado AS (
                SELECT r.data::date, f.codigo, SUM(r.realizado)::numeric AS valor
                FROM producao_realizada r
                JOIN frente_equipe f ON f.id = r.frente_id
                WHERE r.eh_id = :eh_id AND r.data BETWEEN :inicio_calculo AND :fim
                GROUP BY r.data, f.codigo
            ),
            planejado AS (
                SELECT p.data::date, SUM(p.planejado)::numeric AS valor
                FROM producao_planejada p
                JOIN frente_equipe f ON f.id = p.frente_id
                WHERE p.eh_id = :eh_id
                  AND p.data BETWEEN :inicio_calculo AND :fim
                  AND f.codigo = 'RENOVACAO'
                GROUP BY p.data
            )
            SELECT d.data, COALESCE(p.valor, 0) AS planejado,
                   r.codigo, COALESCE(r.valor, 0) AS realizado
            FROM datas d
            LEFT JOIN planejado p ON p.data = d.data
            LEFT JOIN realizado r ON r.data = d.data
            ORDER BY d.data, r.codigo
            """
        ),
        {"eh_id": eh_id, "inicio_calculo": inicio_calculo, "fim": fim},
    ).mappings().all()

    dias = {}
    for row in rows:
        chave = _iso(row["data"])
        item = dias.setdefault(
            chave,
            {"data": chave, "planejado": _float(row["planejado"]), **{codigo: 0.0 for codigo in CODIGOS_EH}},
        )
        if row["codigo"] in CODIGOS_EH:
            item[row["codigo"]] = _float(row["realizado"])

    acumulado_planejado = _float(acumulados_anteriores["planejado"])
    acumulado_renovacao = _float(acumulados_anteriores["renovacao"])
    estado = dict(saldos["valores"])
    resultado = []
    for chave in sorted(dias):
        dia = dias[chave]
        estado = calcular_saldos(estado, dia)
        if chave < inicio:
            continue
        acumulado_planejado += dia["planejado"]
        acumulado_renovacao += dia["RENOVACAO"]
        negativos = [codigo for codigo, valor in estado.items() if valor < 0]
        resultado.append(
            {
                **dia,
                "planejado_acumulado": acumulado_planejado,
                "renovacao_acumulada": acumulado_renovacao,
                "desvio": acumulado_renovacao - acumulado_planejado,
                "saldos": dict(estado),
                "inconsistencias": negativos,
            }
        )
    return resultado, estado


def carregar_dashboard(conn, eh_id, inicio=None, fim=None, data_parte_diaria=None):
    ehs = conn.execute(text("SELECT id, eh AS nome FROM entre_house ORDER BY eh")).mappings().all()
    frentes = conn.execute(
        text(
            """
            SELECT id, frente, codigo
            FROM frente_equipe
            WHERE escopo = 'EH'
            ORDER BY ordem, frente
            """
        )
    ).mappings().all()
    if eh_id is None and ehs:
        eh_id = int(ehs[0]["id"])
    if eh_id is None:
        return {"ehs": [], "eh_id": None, "frentes": frentes}

    inicio, fim = _periodo(conn, eh_id, inicio, fim)
    data_parte_diaria = data_parte_diaria or fim
    config = _configuracao(conn, eh_id)
    saldos = _saldos_iniciais(conn, eh_id, inicio)
    diario, saldos_finais = _producao_diaria(conn, eh_id, inicio, fim, saldos)

    impactos = conn.execute(
        text(
            """
            SELECT i.id, i.data, i.hora_inicio, i.hora_fim, i.minutos_perdidos,
                   i.categoria, i.descricao, i.responsavel, i.providencia, i.status,
                   f.frente AS frente
            FROM operacao_impacto i
            LEFT JOIN frente_equipe f ON f.id = i.frente_id
            WHERE i.eh_id = :eh_id AND i.data BETWEEN :inicio AND :fim
            ORDER BY i.data DESC, i.hora_inicio DESC NULLS LAST, i.id DESC
            """
        ),
        {"eh_id": eh_id, "inicio": inicio, "fim": fim},
    ).mappings().all()

    patio = conn.execute(
        text(
            """
            SELECT p.id, p.data, p.patio, p.classificacao, p.quantidade,
                   p.observacao, e.eh AS origem_eh
            FROM operacao_patio p
            LEFT JOIN entre_house e ON e.id = p.origem_eh_id
            WHERE p.data BETWEEN :inicio AND :fim
            ORDER BY p.data DESC, p.id DESC
            """
        ),
        {"inicio": inicio, "fim": fim},
    ).mappings().all()

    parte_diaria = conn.execute(
        text(
            """
            SELECT
                pd.id,
                a.nome AS atividade,
                m.tag,
                pd.data,
                to_char(pd.hora_inicio, 'HH24:MI') AS hora_inicio,
                to_char(pd.hora_fim, 'HH24:MI') AS hora_fim,
                pd.obs,
                EXTRACT(
                    EPOCH FROM (
                        CASE
                            WHEN pd.hora_fim >= pd.hora_inicio
                                THEN pd.hora_fim - pd.hora_inicio
                            ELSE pd.hora_fim - pd.hora_inicio + INTERVAL '24 hours'
                        END
                    )
                ) / 60 AS duracao_minutos
            FROM parte_diaria pd
            JOIN maquina m ON m.id = pd.maquina_id
            JOIN atividade a ON a.id = pd.atividade_id
            WHERE pd.data = :data_ref
              AND m.tag = 'P190-66001'
            ORDER BY pd.hora_inicio, pd.id
            """
        ),
        {"data_ref": data_parte_diaria},
    ).mappings().all()

    movimentos = defaultdict(float)
    for item in parte_diaria:
        movimentos[item["atividade"]] += _float(item["duracao_minutos"])

    finais = saldos_finais
    totais = conn.execute(
        text(
            """
            SELECT
                COALESCE((
                    SELECT SUM(r.realizado)
                    FROM producao_realizada r
                    JOIN frente_equipe f ON f.id = r.frente_id
                    WHERE r.eh_id = :eh_id AND r.data <= :fim
                      AND f.codigo = 'RENOVACAO'
                ), 0) AS renovado,
                COALESCE((
                    SELECT SUM(p.planejado)
                    FROM producao_planejada p
                    JOIN frente_equipe f ON f.id = p.frente_id
                    WHERE p.eh_id = :eh_id AND p.data <= :fim
                      AND f.codigo = 'RENOVACAO'
                ), 0) AS planejado
            """
        ),
        {"eh_id": eh_id, "fim": fim},
    ).mappings().one()
    renovado = _float(totais["renovado"])
    planejado = _float(totais["planejado"])
    produtividade = config["produtividade_dia"]
    atraso_dias = max(0.0, (planejado - renovado) / produtividade) if produtividade else 0.0
    meta = config["meta_total"]
    percentual = min(100.0, renovado / meta * 100.0) if meta else 0.0
    alertas = sum(len(item["inconsistencias"]) for item in diario)
    minutos_impacto = sum(int(item["minutos_perdidos"] or 0) for item in impactos)

    impacto_por_categoria = defaultdict(int)
    for item in impactos:
        impacto_por_categoria[item["categoria"]] += int(item["minutos_perdidos"] or 0)

    patio_totais = {"BOM": 0.0, "RUIM": 0.0}
    for item in patio:
        patio_totais[item["classificacao"]] += _float(item["quantidade"])

    return {
        "ehs": ehs,
        "eh_id": eh_id,
        "frentes": frentes,
        "inicio": inicio,
        "fim": fim,
        "data_parte_diaria": data_parte_diaria,
        "config": config,
        "saldo_inicial": saldos,
        "diario": diario,
        "impactos": impactos,
        "patio": patio,
        "parte_diaria": parte_diaria,
        "patio_totais": patio_totais,
        "categorias_impacto": CATEGORIAS_IMPACTO,
        "status_impacto": STATUS_IMPACTO,
        "saldos_definicao": SALDOS,
        "kpis": {
            "renovado": renovado,
            "planejado": planejado,
            "meta": meta,
            "percentual": percentual,
            "desvio": renovado - planejado,
            "atraso_dias": atraso_dias,
            "minutos_impacto": minutos_impacto,
            "alertas": alertas,
            **finais,
        },
        "grafico": {
            "labels": [item["data"] for item in diario],
            "planejado": [item["planejado_acumulado"] for item in diario],
            "renovacao": [item["renovacao_acumulada"] for item in diario],
            "pulmao": [item["saldos"]["PULMAO_CHAO"] for item in diario],
            "vagoes": [item["saldos"]["NOVOS_VAGOES"] for item in diario],
            "grampos": [item["saldos"]["GRAMPOS_ABERTOS"] for item in diario],
            "galochas": [item["saldos"]["GALOCHAS_ABERTAS"] for item in diario],
            "pregacao": [item["saldos"]["PREGACAO_ABERTA"] for item in diario],
            "impacto_labels": [
                dict(CATEGORIAS_IMPACTO).get(codigo, codigo)
                for codigo in impacto_por_categoria
            ],
            "impacto_horas": [
                round(minutos / 60.0, 2) for minutos in impacto_por_categoria.values()
            ],
            "movimento_labels": list(movimentos.keys()),
            "movimento_horas": [
                round(minutos / 60.0, 2) for minutos in movimentos.values()
            ],
        },
    }


def gerar_relatorio_xlsx(dashboard, nome_eh):
    wb = Workbook()
    resumo = wb.active
    resumo.title = "Resumo"
    resumo.append(["RELATÓRIO DE PRODUÇÃO POR EH"])
    resumo.append(["EH", nome_eh])
    resumo.append(["Período", f'{dashboard["inicio"]} a {dashboard["fim"]}'])
    resumo.append([])
    resumo.append(["Indicador", "Valor"])
    indicadores = (
        ("Meta total", "meta"),
        ("Renovado no período", "renovado"),
        ("Planejado no período", "planejado"),
        ("Percentual da meta (%)", "percentual"),
        ("Desvio", "desvio"),
        ("Atraso estimado (dias)", "atraso_dias"),
        ("Horas impactadas", "minutos_impacto"),
        ("Pulmão no chão", "PULMAO_CHAO"),
        ("Novos nos vagões", "NOVOS_VAGOES"),
        ("Grampos abertos", "GRAMPOS_ABERTOS"),
        ("Galochas abertas", "GALOCHAS_ABERTAS"),
        ("Pregação aberta", "PREGACAO_ABERTA"),
        ("Velhos nos vagões", "VELHOS_VAGOES"),
    )
    for rotulo, chave in indicadores:
        valor = dashboard["kpis"][chave]
        if chave == "minutos_impacto":
            valor = round(valor / 60.0, 2)
        resumo.append([rotulo, valor])

    producao = wb.create_sheet("Produção diária")
    producao.append(
        [
            "Data", "Planejado", "Renovação", "Carregamento novo",
            "Remoção grampos", "Remoção galochas", "Aplicação grampos",
            "Desc. novo", "Desc. velho", "Pulmão", "Novos vagões",
            "Grampos abertos", "Galochas abertas", "Pregação aberta",
            "Velhos vagões", "Alertas",
        ]
    )
    for item in dashboard["diario"]:
        producao.append(
            [
                item["data"], item["planejado"], item["RENOVACAO"],
                item["CARREGAMENTO_NOVO"], item["REMOCAO_GRAMPOS"],
                item["REMOCAO_GALOCHAS"], item["APLICACAO_GRAMPOS"],
                item["DESCARREGAMENTO_NOVO"], item["DESCARREGAMENTO_VELHO"],
                item["saldos"]["PULMAO_CHAO"], item["saldos"]["NOVOS_VAGOES"],
                item["saldos"]["GRAMPOS_ABERTOS"], item["saldos"]["GALOCHAS_ABERTAS"],
                item["saldos"]["PREGACAO_ABERTA"], item["saldos"]["VELHOS_VAGOES"],
                ", ".join(item["inconsistencias"]),
            ]
        )

    impactos = wb.create_sheet("Impactos")
    impactos.append(
        ["Data", "Frente", "Início", "Fim", "Minutos", "Categoria", "Descrição", "Responsável", "Providência", "Status"]
    )
    for item in dashboard["impactos"]:
        impactos.append(
            [
                _iso(item["data"]), item["frente"] or "", str(item["hora_inicio"] or ""),
                str(item["hora_fim"] or ""), item["minutos_perdidos"], item["categoria"],
                item["descricao"], item["responsavel"] or "", item["providencia"] or "", item["status"],
            ]
        )

    parte_diaria = wb.create_sheet("Parte diária P190")
    parte_diaria.append(["Data", "Movimento / atividade", "Início", "Fim", "Duração (min)", "Observação"])
    for item in dashboard["parte_diaria"]:
        parte_diaria.append(
            [
                _iso(item["data"]),
                item["atividade"],
                item["hora_inicio"],
                item["hora_fim"],
                round(_float(item["duracao_minutos"]), 2),
                item["obs"] or "",
            ]
        )

    patio = wb.create_sheet("Pátio")
    patio.append(["Data", "Pátio", "Classificação", "Quantidade", "Origem EH", "Observação"])
    for item in dashboard["patio"]:
        patio.append(
            [_iso(item["data"]), item["patio"], item["classificacao"], item["quantidade"], item["origem_eh"] or "", item["observacao"] or ""]
        )

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="16324F")
            cell.alignment = Alignment(horizontal="center")
        for column in sheet.columns:
            largura = min(45, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = largura

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo
