from flask import Blueprint, render_template, session, url_for, abort, request, redirect, flash, send_file, current_app
from sqlalchemy import text
from db import get_engine
from routes.auth import login_required, permission_required
from io import BytesIO
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from pypdf import PdfReader, PdfWriter
from PIL import Image
import os
from reportlab.lib.utils import ImageReader
from datetime import date, datetime
import uuid
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash

bp = Blueprint("financeiro_dois", __name__, url_prefix="/financeiro-dois")

def _calcular_saldo_om(conn, om_id: int) -> float:
    saldo = conn.execute(text("""
        SELECT COALESCE(SUM(valor_brl), 0) AS saldo
        FROM financeiro2_om_linhas
        WHERE om_id = :om_id
          AND status = 'Ativo'
    """), {"om_id": om_id}).mappings().first()

    return float(saldo["saldo"] or 0)

def user_can(chave: str) -> bool:
    permissoes = session.get("permissoes", [])
    return chave in permissoes or "auth:administrar" in permissoes

def build_financeiro_dois_subnav(active: str | None):
    links = []

    if user_can("financeiro:visualizar"):
        links.append({"text": "Início", "href": url_for("financeiro_dois.index"), "active": active == "index"})
        links.append({"text": "OM", "href": url_for("financeiro_dois.om"), "active": active == "om"})
        links.append({"text": "RD", "href": url_for("financeiro_dois.rd"), "active": active == "rd"})
        links.append({"text": "Despesas", "href": url_for("financeiro_dois.despesas"), "active": active == "despesas"})
        links.append({"text": "Previsão", "href": url_for("financeiro_dois.previsao"), "active": active == "previsao"})
        links.append({"text": "Reembolsos", "href": url_for("financeiro_dois.reembolsos"), "active": active == "reembolsos"})
        links.append({"text": "Notas de Débito", "href": url_for("financeiro_dois.notas_debito"), "active": active == "nd"})
        links.append({"text": "Aprovações", "href": url_for("financeiro_dois.aprovacoes"), "active": active == "aprovacoes"})
        links.append({"text": "Cadastros", "href": url_for("financeiro_dois.cadastros"), "active": active == "cadastros"})

    return links

def _nome_preenchido(valor: str | None) -> str:
    return (valor or "").strip()


def _redirect_cadastros():
    return redirect(url_for("financeiro_dois.cadastros"))

def _toggle_status_generico(tabela: str, item_id: int, campo_nome: str = "nome"):
    engine = get_engine()
    with engine.begin() as conn:
        item = conn.execute(
            text(f"""
                SELECT id, status, {campo_nome}
                FROM {tabela}
                WHERE id = :id
            """),
            {"id": item_id}
        ).mappings().first()

        if not item:
            flash("Registro não encontrado.", "danger")
            return _redirect_cadastros()

        novo_status = "Inativo" if item["status"] == "Ativo" else "Ativo"

        conn.execute(
            text(f"""
                UPDATE {tabela}
                SET status = :status,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = :id
            """),
            {"status": novo_status, "id": item_id}
        )

    flash(f"Status alterado para {novo_status}.", "success")
    return _redirect_cadastros()
    
def _calcular_saldo_rd(conn, rd_id: int) -> float:
    saldo = conn.execute(text("""
        SELECT COALESCE(SUM(valor), 0) AS saldo
        FROM financeiro2_rd_linhas
        WHERE rd_id = :rd_id
          AND status = 'Ativo'
    """), {"rd_id": rd_id}).mappings().first()

    return float(saldo["saldo"] or 0)
    
def _status_nd_om(conn, om_id: int) -> str:
    totais = conn.execute(text("""
        SELECT
            COUNT(*) FILTER (
                WHERE COALESCE(status, 'Ativo') = 'Ativo'
                  AND COALESCE(valor_brl, 0) > 0
            ) AS total_linhas,
            COUNT(*) FILTER (
                WHERE COALESCE(status, 'Ativo') = 'Ativo'
                  AND COALESCE(valor_brl, 0) > 0
                  AND (
                        COALESCE(numero_nd, '') <> ''
                     OR COALESCE(desconsiderada_nd, FALSE) = TRUE
                  )
            ) AS total_resolvidas
        FROM financeiro2_om_linhas
        WHERE om_id = :om_id
    """), {"om_id": om_id}).mappings().first()

    total_linhas = int((totais or {}).get("total_linhas") or 0)
    total_resolvidas = int((totais or {}).get("total_resolvidas") or 0)

    if total_linhas <= 0 or total_resolvidas <= 0:
        return "NÃO VINCULADA"
    if total_resolvidas < total_linhas:
        return "PARCIAL"
    return "VINCULADA"


def _status_nd_rd(conn, rd_id: int) -> str:
    totais = conn.execute(text("""
        SELECT
            COUNT(*) FILTER (
                WHERE COALESCE(status, 'Ativo') = 'Ativo'
                  AND COALESCE(valor, 0) > 0
            ) AS total_linhas,
            COUNT(*) FILTER (
                WHERE COALESCE(status, 'Ativo') = 'Ativo'
                  AND COALESCE(valor, 0) > 0
                  AND (
                        COALESCE(numero_nd, '') <> ''
                     OR COALESCE(desconsiderada_nd, FALSE) = TRUE
                  )
            ) AS total_resolvidas
        FROM financeiro2_rd_linhas
        WHERE rd_id = :rd_id
    """), {"rd_id": rd_id}).mappings().first()

    total_linhas = int((totais or {}).get("total_linhas") or 0)
    total_resolvidas = int((totais or {}).get("total_resolvidas") or 0)

    if total_linhas <= 0 or total_resolvidas <= 0:
        return "NÃO VINCULADA"
    if total_resolvidas < total_linhas:
        return "PARCIAL"
    return "VINCULADA"    

def _resolver_caminho_anexo_om(nome_arquivo: str) -> str | None:
    if not nome_arquivo:
        return None

    nome_arquivo = str(nome_arquivo).strip().replace("\\", "/")
    if not nome_arquivo:
        return None

    candidatos = []

    # Se já vier um caminho relativo completo
    candidatos.append(os.path.join(current_app.root_path, nome_arquivo))

    # Se vier começando por static/
    candidatos.append(os.path.join(current_app.root_path, nome_arquivo.lstrip("/")))

    # Caminho atual
    candidatos.append(os.path.join(
        current_app.root_path, "static", "uploads", "financeiro2", "om_recibos", os.path.basename(nome_arquivo)
    ))

    # Legados possíveis
    candidatos.append(os.path.join(
        current_app.root_path, "static", "uploads", "financeiro2", os.path.basename(nome_arquivo)
    ))
    candidatos.append(os.path.join(
        current_app.root_path, "static", "uploads", "financeiro2", "recibos", os.path.basename(nome_arquivo)
    ))
    candidatos.append(os.path.join(
        current_app.root_path, "static", "uploads", "om_recibos", os.path.basename(nome_arquivo)
    ))
    candidatos.append(os.path.join(
        current_app.root_path, "static", "uploads", "recibos", os.path.basename(nome_arquivo)
    ))

    for caminho in candidatos:
        if caminho and os.path.exists(caminho):
            return caminho

    return None
        
def _recalcular_status_nd(conn, nd_id: int):
    nd = conn.execute(text("""
        SELECT
            id,
            UPPER(COALESCE(status, '')) AS status
        FROM financeiro2_notas_debito
        WHERE id = :id
    """), {"id": nd_id}).mappings().first()

    if not nd:
        return

    # Se foi definida manualmente como rejeitada/cancelada,
    # não sobrescrever automaticamente.
    if nd["status"] in ("REJEITADA", "CANCELADA"):
        return

    rel = conn.execute(text("""
        SELECT
            COUNT(*) AS total_despesas,
            SUM(CASE WHEN UPPER(COALESCE(d.status_nd, '')) = 'VINCULADA' THEN 1 ELSE 0 END) AS total_vinculadas,
            SUM(CASE WHEN UPPER(COALESCE(d.status_nd, '')) = 'PARCIAL' THEN 1 ELSE 0 END) AS total_parciais,
            SUM(CASE WHEN UPPER(COALESCE(d.status_nd, '')) = 'NÃO VINCULADA' THEN 1 ELSE 0 END) AS total_nao_vinculadas
        FROM financeiro2_notas_debito_despesas rel
        JOIN financeiro2_despesas d ON d.id = rel.despesa_id
        WHERE rel.nd_id = :nd_id
    """), {"nd_id": nd_id}).mappings().first()

    total_despesas = int(rel["total_despesas"] or 0)
    total_vinculadas = int(rel["total_vinculadas"] or 0)
    total_parciais = int(rel["total_parciais"] or 0)
    total_nao_vinculadas = int(rel["total_nao_vinculadas"] or 0)

    if total_despesas <= 0:
        novo_status = "ABERTA"
    elif total_vinculadas == total_despesas:
        novo_status = "VINCULADA"
    elif total_parciais > 0 or total_nao_vinculadas > 0:
        novo_status = "PARCIAL"
    else:
        novo_status = "ABERTA"

    conn.execute(text("""
        UPDATE financeiro2_notas_debito
        SET
            status = :status,
            atualizado_em = CURRENT_TIMESTAMP
        WHERE id = :id
    """), {
        "id": nd_id,
        "status": novo_status,
    })
    
def _usuario_eh_administrador() -> bool:
    return user_can("auth:administrar")

def _validar_senha_usuario_atual(conn, senha_digitada: str) -> bool:
    """
    Ajuste SOMENTE a query abaixo se sua autenticação usar outra tabela/colunas.
    Estou assumindo:
      - session['usuario_id']
      - tabela: usuarios
      - coluna hash: senha_hash
    """
    usuario_id = session.get("usuario_id")
    if not usuario_id or not senha_digitada:
        return False

    usuario = conn.execute(text("""
        SELECT senha_hash
        FROM usuarios
        WHERE id = :id
        LIMIT 1
    """), {"id": usuario_id}).mappings().first()

    if not usuario or not usuario.get("senha_hash"):
        return False

    return check_password_hash(usuario["senha_hash"], senha_digitada)


def _recalcular_status_despesa_origem(conn, origem_tipo: str, origem_id: int):
    origem_tipo = (origem_tipo or "").upper()

    if origem_tipo == "OM":
        totais = conn.execute(text("""
            SELECT
                COUNT(*) FILTER (
                    WHERE COALESCE(status, 'Ativo') = 'Ativo'
                      AND COALESCE(valor_brl, 0) > 0
                ) AS total_linhas,
                COUNT(*) FILTER (
                    WHERE COALESCE(status, 'Ativo') = 'Ativo'
                      AND COALESCE(valor_brl, 0) > 0
                      AND (
                            COALESCE(numero_nd, '') <> ''
                         OR COALESCE(desconsiderada_nd, FALSE) = TRUE
                      )
                ) AS total_resolvidas
            FROM financeiro2_om_linhas
            WHERE om_id = :origem_id
        """), {"origem_id": origem_id}).mappings().first()

    elif origem_tipo == "RD":
        totais = conn.execute(text("""
            SELECT
                COUNT(*) FILTER (
                    WHERE COALESCE(status, 'Ativo') = 'Ativo'
                      AND COALESCE(valor, 0) > 0
                ) AS total_linhas,
                COUNT(*) FILTER (
                    WHERE COALESCE(status, 'Ativo') = 'Ativo'
                      AND COALESCE(valor, 0) > 0
                      AND (
                            COALESCE(numero_nd, '') <> ''
                         OR COALESCE(desconsiderada_nd, FALSE) = TRUE
                      )
                ) AS total_resolvidas
            FROM financeiro2_rd_linhas
            WHERE rd_id = :origem_id
        """), {"origem_id": origem_id}).mappings().first()
    else:
        return

    total_linhas = int(totais["total_linhas"] or 0)
    total_resolvidas = int(totais["total_resolvidas"] or 0)

    if total_resolvidas <= 0:
        novo_status = "NÃO VINCULADA"
    elif total_resolvidas < total_linhas:
        novo_status = "PARCIAL"
    else:
        novo_status = "VINCULADA"

    conn.execute(text("""
        UPDATE financeiro2_despesas
        SET
            status_nd = :status_nd,
            atualizado_em = CURRENT_TIMESTAMP
        WHERE UPPER(COALESCE(origem_tipo, '')) = :origem_tipo
          AND origem_id = :origem_id
    """), {
        "status_nd": novo_status,
        "origem_tipo": origem_tipo,
        "origem_id": origem_id,
    })


def _recalcular_relacao_nd_despesa(conn, nd_numero: str, origem_tipo: str, origem_id: int):
    origem_tipo = (origem_tipo or "").upper()

    despesa = conn.execute(text("""
        SELECT id
        FROM financeiro2_despesas
        WHERE UPPER(COALESCE(origem_tipo, '')) = :origem_tipo
          AND origem_id = :origem_id
        LIMIT 1
    """), {
        "origem_tipo": origem_tipo,
        "origem_id": origem_id,
    }).mappings().first()

    if not despesa:
        return

    nd = conn.execute(text("""
        SELECT id
        FROM financeiro2_notas_debito
        WHERE UPPER(COALESCE(numero_nd, '')) = :numero_nd
        LIMIT 1
    """), {"numero_nd": (nd_numero or "").upper()}).mappings().first()

    if not nd:
        return

    if origem_tipo == "OM":
        uso = conn.execute(text("""
            SELECT COUNT(*) AS total
            FROM financeiro2_om_linhas
            WHERE om_id = :origem_id
              AND (
                    UPPER(COALESCE(numero_nd, '')) = :numero_nd
                 OR UPPER(COALESCE(numero_nd_desconsiderada, '')) = :numero_nd
              )
        """), {
            "origem_id": origem_id,
            "numero_nd": (nd_numero or "").upper(),
        }).mappings().first()
    else:
        uso = conn.execute(text("""
            SELECT COUNT(*) AS total
            FROM financeiro2_rd_linhas
            WHERE rd_id = :origem_id
              AND (
                    UPPER(COALESCE(numero_nd, '')) = :numero_nd
                 OR UPPER(COALESCE(numero_nd_desconsiderada, '')) = :numero_nd
              )
        """), {
            "origem_id": origem_id,
            "numero_nd": (nd_numero or "").upper(),
        }).mappings().first()

    if int(uso["total"] or 0) <= 0:
        conn.execute(text("""
            DELETE FROM financeiro2_notas_debito_despesas
            WHERE nd_id = :nd_id
              AND despesa_id = :despesa_id
        """), {
            "nd_id": nd["id"],
            "despesa_id": despesa["id"],
        })

    _recalcular_status_nd(conn, nd["id"])
    
@bp.route("/")
@login_required
@permission_required("financeiro", "visualizar")
def index():
    cards = [
        {"titulo": "OM", "descricao": "Ordens com saldo automático, despesas, adiantamentos e pagamentos.", "href": url_for("financeiro_dois.om"), "icone": "📄"},
        {"titulo": "RD", "descricao": "Relatórios de despesas por período, colaborador e centro de custo.", "href": url_for("financeiro_dois.rd"), "icone": "🧾"},
        {"titulo": "Despesas", "descricao": "Despesas avulsas ou importadas, com controle de pagamento e vínculo.", "href": url_for("financeiro_dois.despesas"), "icone": "💸"},
        {"titulo": "Previsão", "descricao": "Despesas não vinculadas, em espera ou rejeitadas para ND.", "href": url_for("financeiro_dois.previsao"), "icone": "📊"},
        {"titulo": "Reembolsos", "descricao": "Solicitação, aprovação e pagamento com comprovante.", "href": url_for("financeiro_dois.reembolsos"), "icone": "💳"},
        {"titulo": "Notas de Débito", "descricao": "Criação, edição, vínculo de despesas e exportação em PDF.", "href": url_for("financeiro_dois.notas_debito"), "icone": "🗂️"},
        {"titulo": "Aprovações", "descricao": "Fila de aprovações para exclusões e alterações sensíveis.", "href": url_for("financeiro_dois.aprovacoes"), "icone": "✅"},
        {"titulo": "Cadastros", "descricao": "Página única com abas para categorias, moedas, CC e parâmetros.", "href": url_for("financeiro_dois.cadastros"), "icone": "⚙️"},
    ]

    return render_template(
        "financeiro_dois/index.html",
        subnav_links=build_financeiro_dois_subnav("index"),
        cards=cards,
    )


@bp.route("/cadastros")
@login_required
@permission_required("financeiro", "visualizar")
def cadastros():
    abas = [
        {"id": "categorias", "titulo": "Categorias"},
        {"id": "descricoes", "titulo": "Descrições padrão"},
        {"id": "aplicacoes", "titulo": "Aplicações"},
        {"id": "moedas", "titulo": "Moedas"},
        {"id": "centros_custo", "titulo": "Centros de custo"},
        {"id": "status_despesa", "titulo": "Status despesa"},
        {"id": "status_nd", "titulo": "Status ND"},
        {"id": "tipos_documento", "titulo": "Tipos de documento"},
        {"id": "parametros", "titulo": "Parâmetros"},
        {"id": "empresas_nd", "titulo": "Empresas ND"},
    ]

    engine = get_engine()

    with engine.connect() as conn:
        categorias = conn.execute(text("""
            SELECT id, nome, status
            FROM financeiro2_cad_categorias
            ORDER BY id
        """)).mappings().all()

        descricoes = conn.execute(text("""
            SELECT d.id, d.nome, d.categoria_id, COALESCE(c.nome, '') AS categoria_nome, d.status
            FROM financeiro2_cad_descricoes d
            LEFT JOIN financeiro2_cad_categorias c ON c.id = d.categoria_id
            ORDER BY d.id
        """)).mappings().all()

        aplicacoes = conn.execute(text("""
            SELECT id, nome, status
            FROM financeiro2_cad_aplicacoes
            ORDER BY id
        """)).mappings().all()

        moedas = conn.execute(text("""
            SELECT id, codigo, nome, cambio_padrao, status
            FROM financeiro2_cad_moedas
            ORDER BY id
        """)).mappings().all()

        centros_custo = conn.execute(text("""
            SELECT id, nome, status
            FROM financeiro2_cad_centros_custo
            ORDER BY id
        """)).mappings().all()

        status_despesa = conn.execute(text("""
            SELECT id, nome, status
            FROM financeiro2_cad_status_despesa
            ORDER BY id
        """)).mappings().all()

        status_nd = conn.execute(text("""
            SELECT id, nome, status
            FROM financeiro2_cad_status_nd
            ORDER BY id
        """)).mappings().all()

        tipos_documento = conn.execute(text("""
            SELECT id, nome, status
            FROM financeiro2_cad_tipos_documento
            ORDER BY id
        """)).mappings().all()

        parametros = conn.execute(text("""
            SELECT id, chave, valor, COALESCE(descricao, '') AS descricao, status
            FROM financeiro2_cad_parametros
            ORDER BY id
        """)).mappings().all()

        empresas_nd = conn.execute(text("""
            SELECT id, nome, status
            FROM financeiro2_cad_empresas_nd
            ORDER BY id
        """)).mappings().all()

    return render_template(
        "financeiro_dois/cadastros.html",
        subnav_links=build_financeiro_dois_subnav("cadastros"),
        abas=abas,
        categorias=categorias,
        descricoes=descricoes,
        aplicacoes=aplicacoes,
        moedas=moedas,
        centros_custo=centros_custo,
        status_despesa=status_despesa,
        status_nd=status_nd,
        tipos_documento=tipos_documento,
        parametros=parametros,
        empresas_nd=empresas_nd,
    )
    
def _valor_decimal(txt: str) -> float:
    bruto = (txt or "").strip().replace("R$", "").replace(" ", "")
    if "," in bruto:
        bruto = bruto.replace(".", "").replace(",", ".")
    return float(bruto)

def _proximo_numero_despesa(conn, dt_ref: date | None = None) -> str:
    dt_ref = dt_ref or date.today()
    prefixo = f"DESP-{dt_ref.strftime('%Y%m')}-"

    linha = conn.execute(text("""
        SELECT numero_despesa
        FROM financeiro2_despesas
        WHERE numero_despesa LIKE :prefixo
        ORDER BY numero_despesa DESC
        LIMIT 1
    """), {"prefixo": f"{prefixo}%"}).mappings().first()

    seq = 1
    if linha and linha["numero_despesa"]:
        try:
            seq = int(str(linha["numero_despesa"]).split("-")[-1]) + 1
        except Exception:
            seq = 1

    return f"{prefixo}{seq:04d}"

def _salvar_anexo_despesa(arquivo):
    if not arquivo or not arquivo.filename:
        return None, None

    pasta = os.path.join("static", "uploads", "financeiro2", "despesas")
    os.makedirs(pasta, exist_ok=True)

    nome_original = secure_filename(arquivo.filename)
    extensao = os.path.splitext(nome_original)[1].lower()
    nome_salvo = f"{uuid.uuid4().hex}{extensao}"

    arquivo.save(os.path.join(pasta, nome_salvo))
    return nome_salvo, nome_original

# =========================
# CADASTROS
# =========================

@bp.route("/cadastros/categorias/nova", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def categoria_nova():
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome da categoria.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_categorias
                WHERE LOWER(nome) = LOWER(:nome)
                LIMIT 1
            """),
            {"nome": nome}
        ).fetchone()

        if existe:
            flash("Já existe uma categoria com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                INSERT INTO financeiro2_cad_categorias (nome, status)
                VALUES (:nome, 'Ativo')
            """),
            {"nome": nome}
        )

    flash("Categoria cadastrada com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/categorias/<int:item_id>/editar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def categoria_editar(item_id: int):
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome da categoria.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_categorias
                WHERE LOWER(nome) = LOWER(:nome)
                  AND id <> :id
                LIMIT 1
            """),
            {"nome": nome, "id": item_id}
        ).fetchone()

        if existe:
            flash("Já existe outra categoria com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                UPDATE financeiro2_cad_categorias
                SET nome = :nome,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = :id
            """),
            {"nome": nome, "id": item_id}
        )

    flash("Categoria atualizada com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/categorias/<int:item_id>/toggle-status", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def categoria_toggle_status(item_id: int):
    return _toggle_status_generico("financeiro2_cad_categorias", item_id)


@bp.route("/cadastros/descricoes/nova", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def descricao_nova():
    nome = _nome_preenchido(request.form.get("nome"))
    categoria_id = request.form.get("categoria_id")

    if not nome:
        flash("Informe o nome da descrição.", "warning")
        return _redirect_cadastros()

    categoria_id_int = int(categoria_id) if categoria_id and categoria_id.isdigit() else None

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_descricoes
                WHERE LOWER(nome) = LOWER(:nome)
                LIMIT 1
            """),
            {"nome": nome}
        ).fetchone()

        if existe:
            flash("Já existe uma descrição com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                INSERT INTO financeiro2_cad_descricoes (nome, categoria_id, status)
                VALUES (:nome, :categoria_id, 'Ativo')
            """),
            {"nome": nome, "categoria_id": categoria_id_int}
        )

    flash("Descrição cadastrada com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/descricoes/<int:item_id>/editar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def descricao_editar(item_id: int):
    nome = _nome_preenchido(request.form.get("nome"))
    categoria_id = request.form.get("categoria_id")

    if not nome:
        flash("Informe o nome da descrição.", "warning")
        return _redirect_cadastros()

    categoria_id_int = int(categoria_id) if categoria_id and categoria_id.isdigit() else None

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_descricoes
                WHERE LOWER(nome) = LOWER(:nome)
                  AND id <> :id
                LIMIT 1
            """),
            {"nome": nome, "id": item_id}
        ).fetchone()

        if existe:
            flash("Já existe outra descrição com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                UPDATE financeiro2_cad_descricoes
                SET nome = :nome,
                    categoria_id = :categoria_id,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = :id
            """),
            {"nome": nome, "categoria_id": categoria_id_int, "id": item_id}
        )

    flash("Descrição atualizada com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/descricoes/<int:item_id>/toggle-status", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def descricao_toggle_status(item_id: int):
    return _toggle_status_generico("financeiro2_cad_descricoes", item_id)


@bp.route("/cadastros/aplicacoes/nova", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def aplicacao_nova():
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome da aplicação.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_aplicacoes
                WHERE LOWER(nome) = LOWER(:nome)
                LIMIT 1
            """),
            {"nome": nome}
        ).fetchone()

        if existe:
            flash("Já existe uma aplicação com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                INSERT INTO financeiro2_cad_aplicacoes (nome, status)
                VALUES (:nome, 'Ativo')
            """),
            {"nome": nome}
        )

    flash("Aplicação cadastrada com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/aplicacoes/<int:item_id>/editar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def aplicacao_editar(item_id: int):
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome da aplicação.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_aplicacoes
                WHERE LOWER(nome) = LOWER(:nome)
                  AND id <> :id
                LIMIT 1
            """),
            {"nome": nome, "id": item_id}
        ).fetchone()

        if existe:
            flash("Já existe outra aplicação com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                UPDATE financeiro2_cad_aplicacoes
                SET nome = :nome,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = :id
            """),
            {"nome": nome, "id": item_id}
        )

    flash("Aplicação atualizada com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/aplicacoes/<int:item_id>/toggle-status", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def aplicacao_toggle_status(item_id: int):
    return _toggle_status_generico("financeiro2_cad_aplicacoes", item_id)


@bp.route("/cadastros/moedas/nova", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def moeda_nova():
    codigo = _nome_preenchido(request.form.get("codigo")).upper()
    nome = _nome_preenchido(request.form.get("nome"))
    cambio_padrao = _nome_preenchido(request.form.get("cambio_padrao")) or "1"

    if not codigo or not nome:
        flash("Informe código e nome da moeda.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_moedas
                WHERE UPPER(codigo) = UPPER(:codigo)
                LIMIT 1
            """),
            {"codigo": codigo}
        ).fetchone()

        if existe:
            flash("Já existe uma moeda com esse código.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                INSERT INTO financeiro2_cad_moedas (codigo, nome, cambio_padrao, status)
                VALUES (:codigo, :nome, :cambio_padrao, 'Ativo')
            """),
            {"codigo": codigo, "nome": nome, "cambio_padrao": cambio_padrao}
        )

    flash("Moeda cadastrada com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/moedas/<int:item_id>/editar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def moeda_editar(item_id: int):
    codigo = _nome_preenchido(request.form.get("codigo")).upper()
    nome = _nome_preenchido(request.form.get("nome"))
    cambio_padrao = _nome_preenchido(request.form.get("cambio_padrao")) or "1"

    if not codigo or not nome:
        flash("Informe código e nome da moeda.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_moedas
                WHERE UPPER(codigo) = UPPER(:codigo)
                  AND id <> :id
                LIMIT 1
            """),
            {"codigo": codigo, "id": item_id}
        ).fetchone()

        if existe:
            flash("Já existe outra moeda com esse código.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                UPDATE financeiro2_cad_moedas
                SET codigo = :codigo,
                    nome = :nome,
                    cambio_padrao = :cambio_padrao,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = :id
            """),
            {"codigo": codigo, "nome": nome, "cambio_padrao": cambio_padrao, "id": item_id}
        )

    flash("Moeda atualizada com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/moedas/<int:item_id>/toggle-status", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def moeda_toggle_status(item_id: int):
    return _toggle_status_generico("financeiro2_cad_moedas", item_id)


@bp.route("/cadastros/centros-custo/novo", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def centro_custo_novo():
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome do centro de custo.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_centros_custo
                WHERE LOWER(nome) = LOWER(:nome)
                LIMIT 1
            """),
            {"nome": nome}
        ).fetchone()

        if existe:
            flash("Já existe um centro de custo com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                INSERT INTO financeiro2_cad_centros_custo (nome, status)
                VALUES (:nome, 'Ativo')
            """),
            {"nome": nome}
        )

    flash("Centro de custo cadastrado com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/centros-custo/<int:item_id>/editar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def centro_custo_editar(item_id: int):
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome do centro de custo.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_centros_custo
                WHERE LOWER(nome) = LOWER(:nome)
                  AND id <> :id
                LIMIT 1
            """),
            {"nome": nome, "id": item_id}
        ).fetchone()

        if existe:
            flash("Já existe outro centro de custo com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                UPDATE financeiro2_cad_centros_custo
                SET nome = :nome,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = :id
            """),
            {"nome": nome, "id": item_id}
        )

    flash("Centro de custo atualizado com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/centros-custo/<int:item_id>/toggle-status", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def centro_custo_toggle_status(item_id: int):
    return _toggle_status_generico("financeiro2_cad_centros_custo", item_id)


@bp.route("/cadastros/status-despesa/novo", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def status_despesa_novo():
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome do status.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_status_despesa
                WHERE LOWER(nome) = LOWER(:nome)
                LIMIT 1
            """),
            {"nome": nome}
        ).fetchone()

        if existe:
            flash("Já existe um status de despesa com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                INSERT INTO financeiro2_cad_status_despesa (nome, status)
                VALUES (:nome, 'Ativo')
            """),
            {"nome": nome}
        )

    flash("Status de despesa cadastrado com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/status-despesa/<int:item_id>/editar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def status_despesa_editar(item_id: int):
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome do status.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_status_despesa
                WHERE LOWER(nome) = LOWER(:nome)
                  AND id <> :id
                LIMIT 1
            """),
            {"nome": nome, "id": item_id}
        ).fetchone()

        if existe:
            flash("Já existe outro status de despesa com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                UPDATE financeiro2_cad_status_despesa
                SET nome = :nome,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = :id
            """),
            {"nome": nome, "id": item_id}
        )

    flash("Status de despesa atualizado com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/status-despesa/<int:item_id>/toggle-status", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def status_despesa_toggle_status(item_id: int):
    return _toggle_status_generico("financeiro2_cad_status_despesa", item_id)


@bp.route("/cadastros/status-nd/novo", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def status_nd_novo():
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome do status ND.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_status_nd
                WHERE LOWER(nome) = LOWER(:nome)
                LIMIT 1
            """),
            {"nome": nome}
        ).fetchone()

        if existe:
            flash("Já existe um status ND com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                INSERT INTO financeiro2_cad_status_nd (nome, status)
                VALUES (:nome, 'Ativo')
            """),
            {"nome": nome}
        )

    flash("Status ND cadastrado com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/status-nd/<int:item_id>/editar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def status_nd_editar(item_id: int):
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome do status ND.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_status_nd
                WHERE LOWER(nome) = LOWER(:nome)
                  AND id <> :id
                LIMIT 1
            """),
            {"nome": nome, "id": item_id}
        ).fetchone()

        if existe:
            flash("Já existe outro status ND com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                UPDATE financeiro2_cad_status_nd
                SET nome = :nome,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = :id
            """),
            {"nome": nome, "id": item_id}
        )

    flash("Status ND atualizado com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/status-nd/<int:item_id>/toggle-status", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def status_nd_toggle_status(item_id: int):
    return _toggle_status_generico("financeiro2_cad_status_nd", item_id)


@bp.route("/cadastros/tipos-documento/novo", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def tipo_documento_novo():
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome do tipo de documento.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_tipos_documento
                WHERE LOWER(nome) = LOWER(:nome)
                LIMIT 1
            """),
            {"nome": nome}
        ).fetchone()

        if existe:
            flash("Já existe um tipo de documento com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                INSERT INTO financeiro2_cad_tipos_documento (nome, status)
                VALUES (:nome, 'Ativo')
            """),
            {"nome": nome}
        )

    flash("Tipo de documento cadastrado com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/tipos-documento/<int:item_id>/editar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def tipo_documento_editar(item_id: int):
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome do tipo de documento.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_tipos_documento
                WHERE LOWER(nome) = LOWER(:nome)
                  AND id <> :id
                LIMIT 1
            """),
            {"nome": nome, "id": item_id}
        ).fetchone()

        if existe:
            flash("Já existe outro tipo de documento com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                UPDATE financeiro2_cad_tipos_documento
                SET nome = :nome,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = :id
            """),
            {"nome": nome, "id": item_id}
        )

    flash("Tipo de documento atualizado com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/tipos-documento/<int:item_id>/toggle-status", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def tipo_documento_toggle_status(item_id: int):
    return _toggle_status_generico("financeiro2_cad_tipos_documento", item_id)


@bp.route("/cadastros/parametros/novo", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def parametro_novo():
    chave = _nome_preenchido(request.form.get("chave"))
    valor = _nome_preenchido(request.form.get("valor"))
    descricao = _nome_preenchido(request.form.get("descricao"))

    if not chave or not valor:
        flash("Informe a chave e o valor do parâmetro.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_parametros
                WHERE LOWER(chave) = LOWER(:chave)
                LIMIT 1
            """),
            {"chave": chave}
        ).fetchone()

        if existe:
            flash("Já existe um parâmetro com essa chave.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                INSERT INTO financeiro2_cad_parametros (chave, valor, descricao, status)
                VALUES (:chave, :valor, :descricao, 'Ativo')
            """),
            {"chave": chave, "valor": valor, "descricao": descricao}
        )

    flash("Parâmetro cadastrado com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/parametros/<int:item_id>/editar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def parametro_editar(item_id: int):
    chave = _nome_preenchido(request.form.get("chave"))
    valor = _nome_preenchido(request.form.get("valor"))
    descricao = _nome_preenchido(request.form.get("descricao"))

    if not chave or not valor:
        flash("Informe a chave e o valor do parâmetro.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_parametros
                WHERE LOWER(chave) = LOWER(:chave)
                  AND id <> :id
                LIMIT 1
            """),
            {"chave": chave, "id": item_id}
        ).fetchone()

        if existe:
            flash("Já existe outro parâmetro com essa chave.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                UPDATE financeiro2_cad_parametros
                SET chave = :chave,
                    valor = :valor,
                    descricao = :descricao,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = :id
            """),
            {"chave": chave, "valor": valor, "descricao": descricao, "id": item_id}
        )

    flash("Parâmetro atualizado com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/parametros/<int:item_id>/toggle-status", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def parametro_toggle_status(item_id: int):
    return _toggle_status_generico("financeiro2_cad_parametros", item_id, campo_nome="chave")


@bp.route("/cadastros/empresas-nd/nova", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def empresa_nd_nova():
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome da empresa ND.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_empresas_nd
                WHERE LOWER(nome) = LOWER(:nome)
                LIMIT 1
            """),
            {"nome": nome}
        ).fetchone()

        if existe:
            flash("Já existe uma empresa ND com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                INSERT INTO financeiro2_cad_empresas_nd (nome, status)
                VALUES (:nome, 'Ativo')
            """),
            {"nome": nome}
        )

    flash("Empresa ND cadastrada com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/empresas-nd/<int:item_id>/editar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def empresa_nd_editar(item_id: int):
    nome = _nome_preenchido(request.form.get("nome"))
    if not nome:
        flash("Informe o nome da empresa ND.", "warning")
        return _redirect_cadastros()

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(
            text("""
                SELECT id
                FROM financeiro2_cad_empresas_nd
                WHERE LOWER(nome) = LOWER(:nome)
                  AND id <> :id
                LIMIT 1
            """),
            {"nome": nome, "id": item_id}
        ).fetchone()

        if existe:
            flash("Já existe outra empresa ND com esse nome.", "warning")
            return _redirect_cadastros()

        conn.execute(
            text("""
                UPDATE financeiro2_cad_empresas_nd
                SET nome = :nome,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = :id
            """),
            {"nome": nome, "id": item_id}
        )

    flash("Empresa ND atualizada com sucesso.", "success")
    return _redirect_cadastros()


@bp.route("/cadastros/empresas-nd/<int:item_id>/toggle-status", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def empresa_nd_toggle_status(item_id: int):
    return _toggle_status_generico("financeiro2_cad_empresas_nd", item_id)


# =========================
# OM
# =========================

@bp.route("/om")
@login_required
@permission_required("financeiro", "visualizar")
def om():
    engine = get_engine()

    with engine.connect() as conn:
        oms = conn.execute(text("""
            SELECT
                o.id,
                o.numero_om AS numero,
                o.matricula_colaborador AS matricula,
                o.nome_colaborador AS colaborador,
                o.status,
                TO_CHAR(o.criado_em, 'DD/MM/YYYY') AS criada_em,
                COALESCE(SUM(
                    CASE
                        WHEN COALESCE(l.status, 'Ativo') = 'Ativo' THEN COALESCE(l.valor_brl, 0)
                        ELSE 0
                    END
                ), 0) AS saldo
            FROM financeiro2_om o
            LEFT JOIN financeiro2_om_linhas l ON l.om_id = o.id
            GROUP BY o.id, o.numero_om, o.matricula_colaborador, o.nome_colaborador, o.status, o.criado_em
            ORDER BY o.id
        """)).mappings().all()

        oms = [dict(x) for x in oms]
        for item in oms:
            item["status_nd"] = _status_nd_om(conn, item["id"])

    return render_template(
        "financeiro_dois/om.html",
        subnav_links=build_financeiro_dois_subnav("om"),
        oms=oms,
    )


@bp.route("/om/nova", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def om_nova():
    matricula = _nome_preenchido(request.form.get("matricula_colaborador")).upper()
    nome_colaborador = _nome_preenchido(request.form.get("nome_colaborador")).upper()
    observacao = _nome_preenchido(request.form.get("observacao")).upper()
    numero_om = _nome_preenchido(request.form.get("numero_om")).upper()

    if not numero_om or not matricula or not nome_colaborador:
        flash("Informe o número da OM, a matrícula e o nome do colaborador.", "warning")
        return redirect(url_for("financeiro_dois.om"))

    engine = get_engine()

    with engine.begin() as conn:
        existe = conn.execute(text("""
            SELECT id
            FROM financeiro2_om
            WHERE numero_om = :numero_om
            LIMIT 1
        """), {
            "numero_om": numero_om
        }).mappings().first()

        if existe:
            flash(f"Já existe uma OM com o número {numero_om}.", "warning")
            return redirect(url_for("financeiro_dois.om"))

        novo_id = conn.execute(text("""
            INSERT INTO financeiro2_om (
                numero_om,
                matricula_colaborador,
                nome_colaborador,
                status,
                observacao
            )
            VALUES (
                :numero_om,
                :matricula_colaborador,
                :nome_colaborador,
                'Aberta',
                :observacao
            )
            RETURNING id
        """), {
            "numero_om": numero_om,
            "matricula_colaborador": matricula,
            "nome_colaborador": nome_colaborador,
            "observacao": observacao,
        }).scalar()

    flash(f"OM {numero_om} criada com sucesso.", "success")
    return redirect(url_for("financeiro_dois.om_editar", om_id=novo_id))

@bp.route("/om/<int:om_id>")
@login_required
@permission_required("financeiro", "visualizar")
def om_editar(om_id: int):
    engine = get_engine()

    with engine.connect() as conn:
        om = conn.execute(text("""
            SELECT
                id,
                numero_om AS numero,
                matricula_colaborador AS matricula,
                nome_colaborador AS colaborador,
                status,
                TO_CHAR(criado_em, 'DD/MM/YYYY') AS criada_em,
                COALESCE(observacao, '') AS observacao
            FROM financeiro2_om
            WHERE id = :id
        """), {"id": om_id}).mappings().first()

        if not om:
            abort(404)

        linhas = conn.execute(text("""
            SELECT
                id,
                TO_CHAR(data_lancamento, 'YYYY-MM-DD') AS data_form,
                TO_CHAR(data_lancamento, 'DD/MM/YYYY') AS data,
                COALESCE(recibo, id) AS recibo,
                UPPER(COALESCE(tipo_linha, '')) AS descricao,
                UPPER(COALESCE(detalhes, COALESCE(descricao, ''))) AS detalhes,
                UPPER(COALESCE(categoria, '')) AS categoria,
                UPPER(COALESCE(aplicacao, '')) AS aplicacao,
                COALESCE(valor, 0) AS valor,
                UPPER(COALESCE(moeda_codigo, 'BRL')) AS moeda_codigo,
                COALESCE(cambio, 1) AS cambio,
                COALESCE(valor_brl, 0) AS valor_brl,
                COALESCE(anexo_recibo, '') AS anexo_recibo,
                COALESCE(status, 'Ativo') AS status,
                UPPER(COALESCE(numero_nd, '')) AS numero_nd
            FROM financeiro2_om_linhas
            WHERE om_id = :id
            ORDER BY recibo, id
        """), {"id": om_id}).mappings().all()

        descricoes = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_descricoes
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        categorias = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_categorias
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        aplicacoes = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_aplicacoes
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        moedas = conn.execute(text("""
            SELECT UPPER(codigo) AS codigo, UPPER(nome) AS nome, cambio_padrao
            FROM financeiro2_cad_moedas
            WHERE status = 'Ativo'
            ORDER BY codigo
        """)).mappings().all()
        
        status_nd = _status_nd_om(conn, om["id"])

    total_brl = sum(float(item["valor_brl"]) for item in linhas if str(item["status"]) == "Ativo")

    om = dict(om)
    om["saldo"] = total_brl
    om["linhas"] = linhas
    om["bloqueada"] = str(om["status"]).upper() == "PAGA"
    om["status_nd"] = status_nd

    return render_template(
        "financeiro_dois/om_editar.html",
        subnav_links=build_financeiro_dois_subnav("om"),
        om=om,
        total_brl=total_brl,
        descricoes=descricoes,
        categorias=categorias,
        aplicacoes=aplicacoes,
        moedas=moedas,
    )


@bp.route("/om/<int:om_id>/salvar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def om_salvar(om_id: int):
    numero = _nome_preenchido(request.form.get("numero")).upper()
    matricula = _nome_preenchido(request.form.get("matricula")).upper()
    colaborador = _nome_preenchido(request.form.get("colaborador")).upper()
    observacao = _nome_preenchido(request.form.get("observacao")).upper()
    status = (_nome_preenchido(request.form.get("status")) or "ABERTA").upper()

    if not numero or not matricula or not colaborador:
        flash("Preencha número, matrícula e colaborador.", "warning")
        return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

    engine = get_engine()
    with engine.begin() as conn:
        om = conn.execute(text("""
            SELECT id
            FROM financeiro2_om
            WHERE id = :id
        """), {"id": om_id}).mappings().first()

        if not om:
            flash("OM não encontrada.", "danger")
            return redirect(url_for("financeiro_dois.om"))

        existe = conn.execute(text("""
            SELECT id
            FROM financeiro2_om
            WHERE numero_om = :numero
              AND id <> :id
            LIMIT 1
        """), {"numero": numero, "id": om_id}).mappings().first()

        if existe:
            flash(f"Já existe outra OM com o número {numero}.", "warning")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        conn.execute(text("""
            UPDATE financeiro2_om
            SET numero_om = :numero,
                matricula_colaborador = :matricula,
                nome_colaborador = :colaborador,
                status = :status,
                observacao = :observacao,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :id
        """), {
            "numero": numero,
            "matricula": matricula,
            "colaborador": colaborador,
            "status": status,
            "observacao": observacao,
            "id": om_id
        })

    flash("OM atualizada com sucesso.", "success")
    return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

@bp.route("/om/<int:om_id>/linhas/nova", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def om_linha_nova(om_id: int):
    data_lancamento = _nome_preenchido(request.form.get("data_lancamento"))
    descricao = _nome_preenchido(request.form.get("descricao")).upper()
    detalhes = _nome_preenchido(request.form.get("detalhes")).upper()
    categoria = _nome_preenchido(request.form.get("categoria")).upper()
    aplicacao = _nome_preenchido(request.form.get("aplicacao")).upper()
    valor_txt = _nome_preenchido(request.form.get("valor")).replace(",", ".")
    moeda_codigo = _nome_preenchido(request.form.get("moeda_codigo")).upper() or "BRL"
    anexo_recibo_salvo = _nome_preenchido(request.form.get("anexo_recibo_salvo"))
    forcar_salvamento = request.form.get("forcar_salvamento") == "1"

    if not data_lancamento or not descricao or not categoria or not aplicacao or not valor_txt:
        flash("PREENCHA DATA, DESCRIÇÃO, CATEGORIA, APLICAÇÃO E VALOR.", "warning")
        return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

    try:
        valor = float(valor_txt)
    except ValueError:
        flash("VALOR INVÁLIDO.", "warning")
        return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

    if valor <= 0:
        flash("A LINHA NORMAL DA OM ACEITA APENAS VALOR POSITIVO. VALORES NEGATIVOS SOMENTE EM ADIANTAR OM OU PAGAR OM.", "warning")
        return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

    engine = get_engine()
    with engine.begin() as conn:
        om = conn.execute(text("""
            SELECT id, numero_om, status
            FROM financeiro2_om
            WHERE id = :id
        """), {"id": om_id}).mappings().first()

        if not om:
            flash("OM NÃO ENCONTRADA.", "danger")
            return redirect(url_for("financeiro_dois.om"))

        if str(om["status"]).upper() == "PAGA":
            flash("ESTA OM ESTÁ PAGA E BLOQUEADA PARA EDIÇÃO.", "warning")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        moeda = conn.execute(text("""
            SELECT codigo, cambio_padrao
            FROM financeiro2_cad_moedas
            WHERE UPPER(codigo) = UPPER(:codigo)
            LIMIT 1
        """), {"codigo": moeda_codigo}).mappings().first()

        if not moeda:
            flash("MOEDA INVÁLIDA.", "warning")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        cambio = float(moeda["cambio_padrao"] or 1)
        if cambio == 0:
            cambio = 1

        valor_brl = round(valor / cambio, 2)

        nome_arquivo = anexo_recibo_salvo or None
        arquivo = request.files.get("anexo_recibo")

        if arquivo and arquivo.filename:
            import os
            import uuid
            from werkzeug.utils import secure_filename

            pasta = os.path.join("static", "uploads", "financeiro2", "om_recibos")
            os.makedirs(pasta, exist_ok=True)

            nome_seguro = secure_filename(arquivo.filename)
            extensao = os.path.splitext(nome_seguro)[1].lower()
            nome_arquivo = f"{uuid.uuid4().hex}{extensao}"
            caminho = os.path.join(pasta, nome_arquivo)
            arquivo.save(caminho)

        if not forcar_salvamento:
            duplicadas = conn.execute(text("""
                SELECT *
                FROM (
                    SELECT
                        'OM' AS origem,
                        om.numero_om AS origem_numero,
                        l.id,
                        TO_CHAR(l.data_lancamento, 'DD/MM/YYYY') AS data,
                        COALESCE(l.recibo, l.id) AS recibo,
                        UPPER(COALESCE(l.tipo_linha, '')) AS descricao,
                        UPPER(COALESCE(l.detalhes, '')) AS detalhes,
                        UPPER(COALESCE(l.categoria, '')) AS categoria,
                        UPPER(COALESCE(l.aplicacao, '')) AS aplicacao,
                        COALESCE(l.valor, 0) AS valor,
                        COALESCE(l.anexo_recibo, '') AS anexo_recibo,
                        'om_recibos' AS pasta_recibo
                    FROM financeiro2_om_linhas l
                    JOIN financeiro2_om om ON om.id = l.om_id
                    WHERE l.data_lancamento = :data_lancamento
                      AND l.valor = :valor
                      AND COALESCE(l.status, 'Ativo') = 'Ativo'

                    UNION ALL

                    SELECT
                        'RD' AS origem,
                        rd.numero_rd AS origem_numero,
                        l.id,
                        TO_CHAR(l.data_lancamento, 'DD/MM/YYYY') AS data,
                        NULL AS recibo,
                        UPPER(COALESCE(l.descricao, '')) AS descricao,
                        '' AS detalhes,
                        UPPER(COALESCE(l.categoria, '')) AS categoria,
                        UPPER(COALESCE(l.aplicacao, '')) AS aplicacao,
                        COALESCE(l.valor, 0) AS valor,
                        COALESCE(l.anexo_recibo, '') AS anexo_recibo,
                        'rd_recibos' AS pasta_recibo
                    FROM financeiro2_rd_linhas l
                    JOIN financeiro2_rd rd ON rd.id = l.rd_id
                    WHERE l.data_lancamento = :data_lancamento
                      AND l.valor = :valor
                      AND COALESCE(l.status, 'Ativo') = 'Ativo'
                ) x
                ORDER BY origem, origem_numero, id
            """), {
                "data_lancamento": data_lancamento,
                "valor": valor
            }).mappings().all()

            if duplicadas:
                return render_template(
                    "financeiro_dois/om_confirmar_duplicidade.html",
                    subnav_links=build_financeiro_dois_subnav("om"),
                    om_id=om_id,
                    duplicadas=duplicadas,
                    form_data={
                        "data_lancamento": data_lancamento,
                        "descricao": descricao,
                        "detalhes": detalhes,
                        "categoria": categoria,
                        "aplicacao": aplicacao,
                        "valor": valor_txt,
                        "moeda_codigo": moeda_codigo,
                        "anexo_recibo_salvo": nome_arquivo or "",
                    }
                )

        proximo_recibo = conn.execute(text("""
            SELECT COALESCE(MAX(recibo), 0) + 1 AS proximo
            FROM financeiro2_om_linhas
            WHERE om_id = :om_id
        """), {"om_id": om_id}).mappings().first()["proximo"]

        conn.execute(text("""
            INSERT INTO financeiro2_om_linhas (
                om_id,
                recibo,
                data_lancamento,
                tipo_linha,
                descricao,
                detalhes,
                categoria,
                aplicacao,
                valor,
                sinal,
                moeda_codigo,
                cambio,
                valor_brl,
                anexo_recibo,
                status,
                criado_em,
                atualizado_em
            ) VALUES (
                :om_id,
                :recibo,
                :data_lancamento,
                :tipo_linha,
                :descricao_antiga,
                :detalhes,
                :categoria,
                :aplicacao,
                :valor,
                '+',
                :moeda_codigo,
                :cambio,
                :valor_brl,
                :anexo_recibo,
                'Ativo',
                CURRENT_TIMESTAMP,
                CURRENT_TIMESTAMP
            )
        """), {
            "om_id": om_id,
            "recibo": proximo_recibo,
            "data_lancamento": data_lancamento,
            "tipo_linha": descricao,
            "descricao_antiga": detalhes,
            "detalhes": detalhes,
            "categoria": categoria,
            "aplicacao": aplicacao,
            "valor": valor,
            "moeda_codigo": moeda_codigo,
            "cambio": cambio,
            "valor_brl": valor_brl,
            "anexo_recibo": nome_arquivo
        })

    flash("LINHA DA OM SALVA COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

@bp.route("/om/<int:om_id>/adiantar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def om_adiantar(om_id: int):
    data_lancamento = _nome_preenchido(request.form.get("data_lancamento"))
    aplicacao = _nome_preenchido(request.form.get("aplicacao")).upper()
    valor_txt = _nome_preenchido(request.form.get("valor")).replace(",", ".")
    moeda_codigo = _nome_preenchido(request.form.get("moeda_codigo")).upper() or "BRL"

    if not data_lancamento or not aplicacao or not valor_txt:
        flash("Preencha data, aplicação e valor do adiantamento.", "warning")
        return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

    try:
        valor = float(valor_txt)
    except ValueError:
        flash("Valor inválido para o adiantamento.", "warning")
        return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

    engine = get_engine()
    with engine.begin() as conn:
        om = conn.execute(text("""
            SELECT id, numero_om, status
            FROM financeiro2_om
            WHERE id = :id
        """), {"id": om_id}).mappings().first()

        if not om:
            flash("OM não encontrada.", "danger")
            return redirect(url_for("financeiro_dois.om"))

        if om["status"] == "Paga":
            flash("Esta OM está paga e bloqueada para edição.", "warning")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        moeda = conn.execute(text("""
            SELECT codigo, cambio_padrao
            FROM financeiro2_cad_moedas
            WHERE UPPER(codigo) = UPPER(:codigo)
            LIMIT 1
        """), {"codigo": moeda_codigo}).mappings().first()

        if not moeda:
            flash("Moeda inválida.", "warning")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        cambio = float(moeda["cambio_padrao"] or 1)
        if cambio == 0:
            cambio = 1

        valor_brl = round(valor / cambio, 2)

        proximo_recibo = conn.execute(text("""
            SELECT COALESCE(MAX(recibo), 0) + 1 AS proximo
            FROM financeiro2_om_linhas
            WHERE om_id = :om_id
        """), {"om_id": om_id}).mappings().first()["proximo"]

        conn.execute(text("""
            INSERT INTO financeiro2_om_linhas (
                om_id,
                data_lancamento,
                recibo,
                tipo_linha,
                descricao,
                detalhes,
                categoria,
                aplicacao,
                valor,
                moeda_codigo,
                cambio,
                valor_brl,
                anexo_recibo,
                sinal
            )
            VALUES (
                :om_id,
                :data_lancamento,
                :recibo,
                :tipo_linha,
                :descricao_antiga,
                :detalhes,
                :categoria,
                :aplicacao,
                :valor,
                :moeda_codigo,
                :cambio,
                :valor_brl,
                NULL,
                '-'
            )
        """), {
            "om_id": om_id,
            "data_lancamento": data_lancamento,
            "recibo": proximo_recibo,
            "tipo_linha": "PIX adiantado",
            "descricao_antiga": f"Adiantamento da OM ({om['numero_om']})",
            "detalhes": f"Adiantamento da OM ({om['numero_om']})",
            "categoria": "Adiantamento",
            "aplicacao": aplicacao,
            "valor": -abs(valor),
            "moeda_codigo": moeda_codigo,
            "cambio": cambio,
            "valor_brl": -abs(valor_brl),
        })

        conn.execute(text("""
            UPDATE financeiro2_om
            SET status = 'Parcial',
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :id
              AND status = 'Aberta'
        """), {"id": om_id})

    flash("Adiantamento registrado com sucesso.", "success")
    return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))
    
@bp.route("/om/<int:om_id>/pagar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def om_pagar(om_id: int):
    data_lancamento = _nome_preenchido(request.form.get("data_lancamento"))
    aplicacao = _nome_preenchido(request.form.get("aplicacao")).upper() or "GERAL"

    if not data_lancamento:
        flash("Informe a data do pagamento.", "warning")
        return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

    engine = get_engine()
    with engine.begin() as conn:
        om = conn.execute(text("""
            SELECT id, numero_om, status
            FROM financeiro2_om
            WHERE id = :id
        """), {"id": om_id}).mappings().first()

        if not om:
            flash("OM não encontrada.", "danger")
            return redirect(url_for("financeiro_dois.om"))

        if om["status"] == "Paga":
            flash("Esta OM já está paga.", "warning")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        saldo_atual = _calcular_saldo_om(conn, om_id)

        if saldo_atual <= 0:
            flash("A OM não possui saldo positivo para pagamento.", "warning")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        proximo_recibo = conn.execute(text("""
            SELECT COALESCE(MAX(recibo), 0) + 1 AS proximo
            FROM financeiro2_om_linhas
            WHERE om_id = :om_id
        """), {"om_id": om_id}).mappings().first()["proximo"]

        conn.execute(text("""
            INSERT INTO financeiro2_om_linhas (
                om_id,
                data_lancamento,
                recibo,
                tipo_linha,
                descricao,
                detalhes,
                categoria,
                aplicacao,
                valor,
                moeda_codigo,
                cambio,
                valor_brl,
                anexo_recibo,
                sinal,
                status
            )
            VALUES (
                :om_id,
                :data_lancamento,
                :recibo,
                :tipo_linha,
                :descricao_antiga,
                :detalhes,
                :categoria,
                :aplicacao,
                :valor,
                'BRL',
                1,
                :valor_brl,
                NULL,
                '-',
                'Ativo'
            )
        """), {
            "om_id": om_id,
            "data_lancamento": data_lancamento,
            "recibo": proximo_recibo,
            "tipo_linha": "Pagamento de reembolso",
            "descricao_antiga": f"Pagamento da OM ({om['numero_om']})",
            "detalhes": f"Pagamento da OM ({om['numero_om']})",
            "categoria": "Pagamento",
            "aplicacao": aplicacao,
            "valor": -abs(saldo_atual),
            "valor_brl": -abs(saldo_atual),
        })

        conn.execute(text("""
            UPDATE financeiro2_om
            SET status = 'Paga',
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :id
        """), {"id": om_id})

    flash("Pagamento da OM registrado com sucesso.", "success")
    return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))
    
@bp.route("/om/<int:om_id>/exportar/excel")
@login_required
@permission_required("financeiro", "visualizar")
def om_exportar_excel(om_id: int):
    engine = get_engine()

    with engine.connect() as conn:
        om = conn.execute(text("""
            SELECT
                id,
                numero_om,
                matricula_colaborador,
                nome_colaborador,
                status,
                COALESCE(observacao, '') AS observacao
            FROM financeiro2_om
            WHERE id = :id
        """), {"id": om_id}).mappings().first()

        if not om:
            abort(404)

        linhas = conn.execute(text("""
            SELECT
                TO_CHAR(data_lancamento, 'DD/MM/YYYY') AS data,
                COALESCE(recibo, id) AS recibo,
                COALESCE(tipo_linha, '') AS descricao,
                COALESCE(detalhes, '') AS detalhes,
                COALESCE(categoria, '') AS categoria,
                COALESCE(aplicacao, '') AS aplicacao,
                COALESCE(valor, 0) AS valor,
                COALESCE(moeda_codigo, 'BRL') AS moeda,
                COALESCE(cambio, 1) AS cambio,
                COALESCE(valor_brl, 0) AS valor_brl
            FROM financeiro2_om_linhas
            WHERE om_id = :id
              AND COALESCE(status, 'Ativo') = 'Ativo'
            ORDER BY recibo, id
        """), {"id": om_id}).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "OM"

    ws.append(["Número OM", om["numero_om"]])
    ws.append(["Matrícula", om["matricula_colaborador"]])
    ws.append(["Colaborador", om["nome_colaborador"]])
    ws.append(["Status", om["status"]])
    ws.append(["Observação", om["observacao"]])
    ws.append([])
    ws.append(["Data", "Recibo", "Descrição", "Detalhes", "Categoria", "Aplicação", "Valor", "Moeda", "Câmbio", "Valor BRL"])

    total_brl = 0
    for linha in linhas:
        ws.append([
            linha["data"],
            linha["recibo"],
            linha["descricao"],
            linha["detalhes"],
            linha["categoria"],
            linha["aplicacao"],
            float(linha["valor"]),
            linha["moeda"],
            float(linha["cambio"]),
            float(linha["valor_brl"]),
        ])
        total_brl += float(linha["valor_brl"])

    ws.append([])
    ws.append(["", "", "", "", "", "", "", "", "Saldo BRL", total_brl])

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"{om['numero_om']}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    

@bp.route("/om/anexo")
@login_required
@permission_required("financeiro", "visualizar")
def om_abrir_anexo():
    nome_arquivo = _nome_preenchido(request.args.get("arquivo"))
    caminho = _resolver_caminho_anexo_om(nome_arquivo)
    if not caminho:
        abort(404)
    return send_file(caminho, as_attachment=False)

@bp.route("/om/<int:om_id>/exportar/pdf")
@login_required
@permission_required("financeiro", "visualizar")
def om_exportar_pdf(om_id: int):
    engine = get_engine()

    with engine.connect() as conn:
        om = conn.execute(text("""
            SELECT
                id,
                numero_om,
                matricula_colaborador,
                nome_colaborador,
                status,
                COALESCE(observacao, '') AS observacao
            FROM financeiro2_om
            WHERE id = :id
        """), {"id": om_id}).mappings().first()

        if not om:
            abort(404)

        linhas = conn.execute(text("""
            SELECT
                id,
                TO_CHAR(data_lancamento, 'DD/MM/YYYY') AS data,
                COALESCE(recibo, id) AS recibo,
                COALESCE(tipo_linha, '') AS descricao,
                COALESCE(detalhes, '') AS detalhes,
                COALESCE(categoria, '') AS categoria,
                COALESCE(aplicacao, '') AS aplicacao,
                COALESCE(valor_brl, 0) AS valor_brl,
                COALESCE(anexo_recibo, '') AS anexo_recibo,
                COALESCE(status, 'Ativo') AS status
            FROM financeiro2_om_linhas
            WHERE om_id = :id
              AND COALESCE(status, 'Ativo') = 'Ativo'
            ORDER BY recibo, id
        """), {"id": om_id}).mappings().all()

    total_brl = sum(float(l["valor_brl"]) for l in linhas if str(l["status"]) == "Ativo")

    # PDF base da OM
    buffer_base = BytesIO()
    pdf = canvas.Canvas(buffer_base, pagesize=A4)
    largura, altura = A4
    y = altura - 40

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(40, y, f"OM {om['numero_om']}")
    y -= 20

    pdf.setFont("Helvetica", 10)
    pdf.drawString(40, y, f"Matrícula: {om['matricula_colaborador']}")
    y -= 15
    pdf.drawString(40, y, f"Colaborador: {om['nome_colaborador']}")
    y -= 15
    pdf.drawString(40, y, f"Status: {om['status']}")
    y -= 15
    pdf.drawString(40, y, f"Observação: {om['observacao']}")
    y -= 25

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(40, y, "Data")
    pdf.drawString(90, y, "Recibo")
    pdf.drawString(135, y, "Descrição")
    pdf.drawString(260, y, "Categoria")
    pdf.drawString(350, y, "Aplicação")
    pdf.drawString(470, y, "Valor BRL")
    y -= 15

    pdf.setFont("Helvetica", 8)
    for linha in linhas:
        if y < 50:
            pdf.showPage()
            y = altura - 40
            pdf.setFont("Helvetica-Bold", 9)
            pdf.drawString(40, y, "Data")
            pdf.drawString(90, y, "Recibo")
            pdf.drawString(135, y, "Descrição")
            pdf.drawString(260, y, "Categoria")
            pdf.drawString(350, y, "Aplicação")
            pdf.drawString(470, y, "Valor BRL")
            y -= 15
            pdf.setFont("Helvetica", 8)

        pdf.drawString(40, y, str(linha["data"]))
        pdf.drawString(90, y, str(linha["recibo"]))
        pdf.drawString(135, y, str(linha["descricao"])[:22])
        pdf.drawString(260, y, str(linha["categoria"])[:14])
        pdf.drawString(350, y, str(linha["aplicacao"])[:18])
        pdf.drawRightString(540, y, f"{float(linha['valor_brl']):.2f}")
        y -= 13

    y -= 10
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawRightString(540, y, f"Saldo BRL: {total_brl:.2f}")
    pdf.save()
    buffer_base.seek(0)

    writer = PdfWriter()

    # Páginas do PDF base
    base_reader = PdfReader(buffer_base)
    for page in base_reader.pages:
        writer.add_page(page)

    # Anexos/recibos das linhas
    for linha in linhas:
        nome_anexo = (linha["anexo_recibo"] or "").strip()
        if not nome_anexo:
            continue

        caminho = _resolver_caminho_anexo_om(nome_anexo)
        if not caminho:
            continue

        extensao = os.path.splitext(caminho)[1].lower()

        try:
            if extensao == ".pdf":
                with open(caminho, "rb") as f:
                    anexo_reader = PdfReader(f)
                    for page in anexo_reader.pages:
                        writer.add_page(page)

            elif extensao in (".jpg", ".jpeg", ".png", ".bmp", ".webp"):
                imagem = Image.open(caminho)
                if imagem.mode != "RGB":
                    imagem = imagem.convert("RGB")

                largura_img, altura_img = imagem.size
                proporcao = min((A4[0] - 60) / largura_img, (A4[1] - 60) / altura_img)
                nova_largura = largura_img * proporcao
                nova_altura = altura_img * proporcao

                buffer_img = BytesIO()
                c = canvas.Canvas(buffer_img, pagesize=A4)
                c.drawImage(
                    ImageReader(imagem),
                    30,
                    (A4[1] - nova_altura) / 2,
                    width=nova_largura,
                    height=nova_altura,
                    preserveAspectRatio=True,
                    anchor='c'
                )
                c.save()
                buffer_img.seek(0)

                img_reader = PdfReader(buffer_img)
                for page in img_reader.pages:
                    writer.add_page(page)

        except Exception:
            # Ignora anexos corrompidos ou formatos problemáticos sem quebrar o PDF principal
            continue

    arquivo_final = BytesIO()
    writer.write(arquivo_final)
    arquivo_final.seek(0)

    return send_file(
        arquivo_final,
        as_attachment=True,
        download_name=f"{om['numero_om']}.pdf",
        mimetype="application/pdf",
    )
    
@bp.route("/om/<int:om_id>/linhas/<int:linha_id>/editar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def om_linha_editar(om_id: int, linha_id: int):
    data_lancamento = _nome_preenchido(request.form.get("data_lancamento"))
    descricao = _nome_preenchido(request.form.get("descricao")).upper()
    detalhes = _nome_preenchido(request.form.get("detalhes")).upper()
    categoria = _nome_preenchido(request.form.get("categoria")).upper()
    aplicacao = _nome_preenchido(request.form.get("aplicacao")).upper()
    valor_txt = _nome_preenchido(request.form.get("valor")).replace(",", ".")
    moeda_codigo = (_nome_preenchido(request.form.get("moeda_codigo")) or "BRL").upper()

    if not data_lancamento or not descricao or not categoria or not aplicacao or not valor_txt:
        flash("Preencha data, descrição, categoria, aplicação e valor.", "warning")
        return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

    try:
        valor = float(valor_txt)
    except ValueError:
        flash("Valor inválido.", "warning")
        return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

    engine = get_engine()
    with engine.begin() as conn:
        om = conn.execute(text("""
            SELECT id, status
            FROM financeiro2_om
            WHERE id = :id
        """), {"id": om_id}).mappings().first()

        if not om:
            flash("OM não encontrada.", "danger")
            return redirect(url_for("financeiro_dois.om"))

        if om["status"] == "Paga":
            flash("Esta OM está paga e bloqueada para edição.", "warning")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        linha = conn.execute(text("""
            SELECT id, sinal, status
            FROM financeiro2_om_linhas
            WHERE id = :linha_id
              AND om_id = :om_id
        """), {"linha_id": linha_id, "om_id": om_id}).mappings().first()

        if not linha:
            flash("Linha não encontrada.", "danger")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        if linha["status"] != "Ativo":
            flash("A linha está inativa e não pode ser editada.", "warning")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        moeda = conn.execute(text("""
            SELECT codigo, cambio_padrao
            FROM financeiro2_cad_moedas
            WHERE codigo = :codigo
            LIMIT 1
        """), {"codigo": moeda_codigo}).mappings().first()

        if not moeda:
            flash("Moeda inválida.", "warning")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        cambio = float(moeda["cambio_padrao"] or 1)
        if cambio == 0:
            cambio = 1

        valor_brl = round(valor / cambio, 2)

        sinal = linha["sinal"] or "+"
        valor_final = valor if sinal == "+" else -abs(valor)
        valor_brl_final = valor_brl if sinal == "+" else -abs(valor_brl)

        conn.execute(text("""
            UPDATE financeiro2_om_linhas
            SET data_lancamento = :data_lancamento,
                tipo_linha = :tipo_linha,
                descricao = :descricao_antiga,
                detalhes = :detalhes,
                categoria = :categoria,
                aplicacao = :aplicacao,
                valor = :valor,
                moeda_codigo = :moeda_codigo,
                cambio = :cambio,
                valor_brl = :valor_brl,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :linha_id
              AND om_id = :om_id
        """), {
            "data_lancamento": data_lancamento,
            "tipo_linha": descricao,
            "descricao_antiga": detalhes,
            "detalhes": detalhes,
            "categoria": categoria,
            "aplicacao": aplicacao,
            "valor": valor_final,
            "moeda_codigo": moeda_codigo,
            "cambio": cambio,
            "valor_brl": valor_brl_final,
            "linha_id": linha_id,
            "om_id": om_id
        })

    flash("Linha atualizada com sucesso.", "success")
    return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))
    
@bp.route("/om/<int:om_id>/linhas/<int:linha_id>/toggle-status", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def om_linha_toggle_status(om_id: int, linha_id: int):
    engine = get_engine()
    with engine.begin() as conn:
        om = conn.execute(text("""
            SELECT id, status
            FROM financeiro2_om
            WHERE id = :id
        """), {"id": om_id}).mappings().first()

        if not om:
            flash("OM não encontrada.", "danger")
            return redirect(url_for("financeiro_dois.om"))

        if om["status"] == "Paga":
            flash("Esta OM está paga e bloqueada para edição.", "warning")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        linha = conn.execute(text("""
            SELECT id, status
            FROM financeiro2_om_linhas
            WHERE id = :linha_id
              AND om_id = :om_id
        """), {"linha_id": linha_id, "om_id": om_id}).mappings().first()

        if not linha:
            flash("Linha não encontrada.", "danger")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        novo_status = "Inativo" if linha["status"] == "Ativo" else "Ativo"

        conn.execute(text("""
            UPDATE financeiro2_om_linhas
            SET status = :status,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :linha_id
              AND om_id = :om_id
        """), {
            "status": novo_status,
            "linha_id": linha_id,
            "om_id": om_id
        })

    flash(f"Linha alterada para {novo_status}.", "success")
    return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

# =========================
# RD
# =========================

@bp.route("/rd")
@login_required
@permission_required("financeiro", "visualizar")
def rd():
    engine = get_engine()

    with engine.connect() as conn:
        rds = conn.execute(text("""
            SELECT
                r.id,
                r.numero_rd AS numero,
                r.periodo,
                r.matricula_colaborador AS matricula,
                r.nome_colaborador AS colaborador,
                r.centro_custo,
                r.status,
                TO_CHAR(r.criado_em, 'DD/MM/YYYY') AS criada_em,
                COALESCE(SUM(
                    CASE
                        WHEN COALESCE(l.status, 'Ativo') = 'Ativo' THEN COALESCE(l.valor, 0)
                        ELSE 0
                    END
                ), 0) AS saldo
            FROM financeiro2_rd r
            LEFT JOIN financeiro2_rd_linhas l ON l.rd_id = r.id
            GROUP BY
                r.id, r.numero_rd, r.periodo,
                r.matricula_colaborador, r.nome_colaborador,
                r.centro_custo, r.status, r.criado_em
            ORDER BY r.id
        """)).mappings().all()

        rds = [dict(x) for x in rds]
        for item in rds:
            item["status_nd"] = _status_nd_rd(conn, item["id"])

    return render_template(
        "financeiro_dois/rd.html",
        subnav_links=build_financeiro_dois_subnav("rd"),
        rds=rds,
    )


@bp.route("/rd/nova", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def rd_nova():
    numero_rd = _nome_preenchido(request.form.get("numero_rd")).upper()
    periodo = _nome_preenchido(request.form.get("periodo")).upper()
    matricula = _nome_preenchido(request.form.get("matricula_colaborador")).upper()
    nome_colaborador = _nome_preenchido(request.form.get("nome_colaborador")).upper()
    centro_custo = _nome_preenchido(request.form.get("centro_custo")).upper()
    observacao = _nome_preenchido(request.form.get("observacao")).upper()

    if not numero_rd or not periodo or not matricula or not nome_colaborador or not centro_custo:
        flash("INFORME NÚMERO DA RD, PERÍODO, MATRÍCULA, NOME DO COLABORADOR E CENTRO DE CUSTO.", "warning")
        return redirect(url_for("financeiro_dois.rd"))

    engine = get_engine()

    with engine.begin() as conn:
        existe = conn.execute(text("""
            SELECT id
            FROM financeiro2_rd
            WHERE numero_rd = :numero_rd
            LIMIT 1
        """), {
            "numero_rd": numero_rd
        }).mappings().first()

        if existe:
            flash(f"JÁ EXISTE UMA RD COM O NÚMERO {numero_rd}.", "warning")
            return redirect(url_for("financeiro_dois.rd"))

        novo_id = conn.execute(text("""
            INSERT INTO financeiro2_rd (
                numero_rd,
                periodo,
                matricula_colaborador,
                nome_colaborador,
                centro_custo,
                status,
                observacao
            )
            VALUES (
                :numero_rd,
                :periodo,
                :matricula_colaborador,
                :nome_colaborador,
                :centro_custo,
                'Aberta',
                :observacao
            )
            RETURNING id
        """), {
            "numero_rd": numero_rd,
            "periodo": periodo,
            "matricula_colaborador": matricula,
            "nome_colaborador": nome_colaborador,
            "centro_custo": centro_custo,
            "observacao": observacao,
        }).scalar()

    flash(f"RD {numero_rd} CRIADA COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.rd_editar", rd_id=novo_id))

@bp.route("/rd/<int:rd_id>")
@login_required
@permission_required("financeiro", "visualizar")
def rd_editar(rd_id: int):
    engine = get_engine()

    with engine.connect() as conn:
        rd = conn.execute(text("""
            SELECT
                id,
                numero_rd AS numero,
                periodo,
                matricula_colaborador AS matricula,
                nome_colaborador AS colaborador,
                centro_custo,
                status,
                TO_CHAR(criado_em, 'DD/MM/YYYY') AS criada_em,
                COALESCE(observacao, '') AS observacao
            FROM financeiro2_rd
            WHERE id = :id
        """), {"id": rd_id}).mappings().first()

        if not rd:
            abort(404)

        linhas = conn.execute(text("""
            SELECT
                id,
                TO_CHAR(data_lancamento, 'YYYY-MM-DD') AS data_form,
                TO_CHAR(data_lancamento, 'DD/MM/YYYY') AS data,
                UPPER(COALESCE(descricao, '')) AS descricao,
                UPPER(COALESCE(categoria, '')) AS categoria,
                UPPER(COALESCE(aplicacao, '')) AS aplicacao,
                COALESCE(valor, 0) AS valor,
                COALESCE(status, 'Ativo') AS status,
                COALESCE(anexo_recibo, '') AS anexo_recibo,
                UPPER(COALESCE(numero_nd, '')) AS numero_nd
            FROM financeiro2_rd_linhas
            WHERE rd_id = :id
            ORDER BY id
        """), {"id": rd_id}).mappings().all()

        descricoes = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_descricoes
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        categorias = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_categorias
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        aplicacoes = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_aplicacoes
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        centros_custo_lista = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_centros_custo
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()
        
        status_nd = _status_nd_rd(conn, rd["id"])
        total_valor = sum(float(item["valor"]) for item in linhas if item["status"] == "Ativo")

        rd = dict(rd)
        rd["saldo"] = total_valor
        rd["linhas"] = linhas
        rd["bloqueada"] = str(rd["status"]).upper() == "QUITADA"
        rd["status_nd"] = status_nd

        return render_template(
            "financeiro_dois/rd_editar.html",
            subnav_links=build_financeiro_dois_subnav("rd"),
            rd=rd,
            total_valor=total_valor,
            descricoes=descricoes,
            categorias=categorias,
            aplicacoes=aplicacoes,
            centros_custo_lista=centros_custo_lista,
        )


@bp.route("/rd/<int:rd_id>/salvar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def rd_salvar(rd_id: int):
    numero = _nome_preenchido(request.form.get("numero")).upper()
    periodo = _nome_preenchido(request.form.get("periodo")).upper()
    matricula = _nome_preenchido(request.form.get("matricula")).upper()
    colaborador = _nome_preenchido(request.form.get("colaborador")).upper()
    centro_custo = _nome_preenchido(request.form.get("centro_custo")).upper()
    status = (_nome_preenchido(request.form.get("status")) or "ABERTA").upper()
    observacao = _nome_preenchido(request.form.get("observacao")).upper()

    if not numero or not periodo or not matricula or not colaborador or not centro_custo:
        flash("PREENCHA NÚMERO, PERÍODO, MATRÍCULA, COLABORADOR E CENTRO DE CUSTO.", "warning")
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    engine = get_engine()
    with engine.begin() as conn:
        rd = conn.execute(text("""
            SELECT id
            FROM financeiro2_rd
            WHERE id = :id
        """), {"id": rd_id}).mappings().first()

        if not rd:
            flash("RD NÃO ENCONTRADA.", "danger")
            return redirect(url_for("financeiro_dois.rd"))

        existe = conn.execute(text("""
            SELECT id
            FROM financeiro2_rd
            WHERE numero_rd = :numero
              AND id <> :id
            LIMIT 1
        """), {"numero": numero, "id": rd_id}).mappings().first()

        if existe:
            flash(f"JÁ EXISTE OUTRA RD COM O NÚMERO {numero}.", "warning")
            return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

        conn.execute(text("""
            UPDATE financeiro2_rd
            SET numero_rd = :numero,
                periodo = :periodo,
                matricula_colaborador = :matricula,
                nome_colaborador = :colaborador,
                centro_custo = :centro_custo,
                status = :status,
                observacao = :observacao,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :id
        """), {
            "numero": numero,
            "periodo": periodo,
            "matricula": matricula,
            "colaborador": colaborador,
            "centro_custo": centro_custo,
            "status": status,
            "observacao": observacao,
            "id": rd_id
        })

    flash("RD ATUALIZADA COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))
    
@bp.route("/rd/<int:rd_id>/linhas/nova", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def rd_linha_nova(rd_id: int):
    data_lancamento = _nome_preenchido(request.form.get("data_lancamento"))
    descricao = _nome_preenchido(request.form.get("descricao")).upper()
    categoria = _nome_preenchido(request.form.get("categoria")).upper()
    aplicacao = _nome_preenchido(request.form.get("aplicacao")).upper()
    valor_txt = _nome_preenchido(request.form.get("valor")).replace(",", ".")
    anexo_recibo_salvo = _nome_preenchido(request.form.get("anexo_recibo_salvo"))
    forcar_salvamento = request.form.get("forcar_salvamento") == "1"

    if not data_lancamento or not descricao or not categoria or not aplicacao or not valor_txt:
        flash("PREENCHA DATA, DESCRIÇÃO, CATEGORIA, APLICAÇÃO E VALOR.", "warning")
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    try:
        valor = float(valor_txt)
    except ValueError:
        flash("VALOR INVÁLIDO.", "warning")
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    if valor <= 0:
        flash("A RD ACEITA APENAS VALORES POSITIVOS.", "warning")
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    engine = get_engine()
    with engine.begin() as conn:
        rd = conn.execute(text("""
            SELECT id, status
            FROM financeiro2_rd
            WHERE id = :id
        """), {"id": rd_id}).mappings().first()

        if not rd:
            flash("RD NÃO ENCONTRADA.", "danger")
            return redirect(url_for("financeiro_dois.rd"))

        if str(rd["status"]).upper() == "QUITADA":
            flash("ESTA RD ESTÁ QUITADA E BLOQUEADA PARA EDIÇÃO.", "warning")
            return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

        nome_arquivo = anexo_recibo_salvo or None
        arquivo = request.files.get("anexo_recibo")

        if arquivo and arquivo.filename:
            import os
            import uuid
            from werkzeug.utils import secure_filename

            pasta = os.path.join("static", "uploads", "financeiro2", "rd_recibos")
            os.makedirs(pasta, exist_ok=True)

            nome_seguro = secure_filename(arquivo.filename)
            extensao = os.path.splitext(nome_seguro)[1].lower()
            nome_arquivo = f"{uuid.uuid4().hex}{extensao}"
            caminho = os.path.join(pasta, nome_arquivo)
            arquivo.save(caminho)

        if not forcar_salvamento:
            duplicadas = conn.execute(text("""
                SELECT *
                FROM (
                    SELECT
                        'RD' AS origem,
                        rd.numero_rd AS origem_numero,
                        l.id,
                        TO_CHAR(l.data_lancamento, 'DD/MM/YYYY') AS data,
                        NULL AS recibo,
                        UPPER(COALESCE(l.descricao, '')) AS descricao,
                        '' AS detalhes,
                        UPPER(COALESCE(l.categoria, '')) AS categoria,
                        UPPER(COALESCE(l.aplicacao, '')) AS aplicacao,
                        COALESCE(l.valor, 0) AS valor,
                        COALESCE(l.anexo_recibo, '') AS anexo_recibo,
                        'rd_recibos' AS pasta_recibo
                    FROM financeiro2_rd_linhas l
                    JOIN financeiro2_rd rd ON rd.id = l.rd_id
                    WHERE l.data_lancamento = :data_lancamento
                      AND l.valor = :valor
                      AND COALESCE(l.status, 'Ativo') = 'Ativo'

                    UNION ALL

                    SELECT
                        'OM' AS origem,
                        om.numero_om AS origem_numero,
                        l.id,
                        TO_CHAR(l.data_lancamento, 'DD/MM/YYYY') AS data,
                        COALESCE(l.recibo, l.id) AS recibo,
                        UPPER(COALESCE(l.tipo_linha, '')) AS descricao,
                        UPPER(COALESCE(l.detalhes, '')) AS detalhes,
                        UPPER(COALESCE(l.categoria, '')) AS categoria,
                        UPPER(COALESCE(l.aplicacao, '')) AS aplicacao,
                        COALESCE(l.valor, 0) AS valor,
                        COALESCE(l.anexo_recibo, '') AS anexo_recibo,
                        'om_recibos' AS pasta_recibo
                    FROM financeiro2_om_linhas l
                    JOIN financeiro2_om om ON om.id = l.om_id
                    WHERE l.data_lancamento = :data_lancamento
                      AND l.valor = :valor
                      AND COALESCE(l.status, 'Ativo') = 'Ativo'
                ) x
                ORDER BY origem, origem_numero, id
            """), {
                "data_lancamento": data_lancamento,
                "valor": valor
            }).mappings().all()

            if duplicadas:
                return render_template(
                    "financeiro_dois/rd_confirmar_duplicidade.html",
                    subnav_links=build_financeiro_dois_subnav("rd"),
                    rd_id=rd_id,
                    duplicadas=duplicadas,
                    form_data={
                        "data_lancamento": data_lancamento,
                        "descricao": descricao,
                        "categoria": categoria,
                        "aplicacao": aplicacao,
                        "valor": valor_txt,
                        "anexo_recibo_salvo": nome_arquivo or "",
                    }
                )

        conn.execute(text("""
            INSERT INTO financeiro2_rd_linhas (
                rd_id,
                data_lancamento,
                descricao,
                categoria,
                aplicacao,
                valor,
                anexo_recibo,
                status,
                criado_em,
                atualizado_em
            ) VALUES (
                :rd_id,
                :data_lancamento,
                :descricao,
                :categoria,
                :aplicacao,
                :valor,
                :anexo_recibo,
                'Ativo',
                CURRENT_TIMESTAMP,
                CURRENT_TIMESTAMP
            )
        """), {
            "rd_id": rd_id,
            "data_lancamento": data_lancamento,
            "descricao": descricao,
            "categoria": categoria,
            "aplicacao": aplicacao,
            "valor": valor,
            "anexo_recibo": nome_arquivo
        })

    flash("LINHA DA RD SALVA COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

@bp.route("/rd/<int:rd_id>/linhas/<int:linha_id>/editar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def rd_linha_editar(rd_id: int, linha_id: int):
    data_lancamento = _nome_preenchido(request.form.get("data_lancamento"))
    descricao = _nome_preenchido(request.form.get("descricao")).upper()
    categoria = _nome_preenchido(request.form.get("categoria")).upper()
    aplicacao = _nome_preenchido(request.form.get("aplicacao")).upper()
    valor_txt = _nome_preenchido(request.form.get("valor")).replace(",", ".")

    if not data_lancamento or not descricao or not categoria or not aplicacao or not valor_txt:
        flash("PREENCHA DATA, DESCRIÇÃO, CATEGORIA, APLICAÇÃO E VALOR.", "warning")
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    try:
        valor = float(valor_txt)
    except ValueError:
        flash("VALOR INVÁLIDO.", "warning")
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    engine = get_engine()
    with engine.begin() as conn:
        rd = conn.execute(text("""
            SELECT id, status
            FROM financeiro2_rd
            WHERE id = :id
        """), {"id": rd_id}).mappings().first()

        if not rd:
            flash("RD NÃO ENCONTRADA.", "danger")
            return redirect(url_for("financeiro_dois.rd"))

        if str(rd["status"]).upper() == "QUITADA":
            flash("ESTA RD ESTÁ QUITADA E BLOQUEADA PARA EDIÇÃO.", "warning")
            return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

        linha = conn.execute(text("""
            SELECT id, status
            FROM financeiro2_rd_linhas
            WHERE id = :linha_id
              AND rd_id = :rd_id
        """), {"linha_id": linha_id, "rd_id": rd_id}).mappings().first()

        if not linha:
            flash("LINHA NÃO ENCONTRADA.", "danger")
            return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

        if linha["status"] != "Ativo":
            flash("A LINHA ESTÁ INATIVA E NÃO PODE SER EDITADA.", "warning")
            return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

        conn.execute(text("""
            UPDATE financeiro2_rd_linhas
            SET data_lancamento = :data_lancamento,
                descricao = :descricao,
                categoria = :categoria,
                aplicacao = :aplicacao,
                valor = :valor,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :linha_id
              AND rd_id = :rd_id
        """), {
            "data_lancamento": data_lancamento,
            "descricao": descricao,
            "categoria": categoria,
            "aplicacao": aplicacao,
            "valor": valor,
            "linha_id": linha_id,
            "rd_id": rd_id
        })

    flash("LINHA ATUALIZADA COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

@bp.route("/rd/<int:rd_id>/linhas/<int:linha_id>/toggle-status", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def rd_linha_toggle_status(rd_id: int, linha_id: int):
    engine = get_engine()
    with engine.begin() as conn:
        rd = conn.execute(text("""
            SELECT id
            FROM financeiro2_rd
            WHERE id = :id
        """), {"id": rd_id}).mappings().first()

        if not rd:
            flash("RD NÃO ENCONTRADA.", "danger")
            return redirect(url_for("financeiro_dois.rd"))

        linha = conn.execute(text("""
            SELECT id, status
            FROM financeiro2_rd_linhas
            WHERE id = :linha_id
              AND rd_id = :rd_id
        """), {"linha_id": linha_id, "rd_id": rd_id}).mappings().first()

        if not linha:
            flash("LINHA NÃO ENCONTRADA.", "danger")
            return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

        novo_status = "Inativo" if linha["status"] == "Ativo" else "Ativo"

        conn.execute(text("""
            UPDATE financeiro2_rd_linhas
            SET status = :status,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :linha_id
              AND rd_id = :rd_id
        """), {
            "status": novo_status,
            "linha_id": linha_id,
            "rd_id": rd_id
        })

    flash(f"LINHA ALTERADA PARA {novo_status.upper()}.", "success")
    return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))
    
@bp.route("/om/<int:om_id>/linhas/confirmar-duplicidade", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def om_linha_confirmar_duplicidade(om_id: int):
    data_lancamento = _nome_preenchido(request.form.get("data_lancamento"))
    descricao = _nome_preenchido(request.form.get("descricao")).upper()
    detalhes = _nome_preenchido(request.form.get("detalhes")).upper()
    categoria = _nome_preenchido(request.form.get("categoria")).upper()
    aplicacao = _nome_preenchido(request.form.get("aplicacao")).upper()
    valor_txt = _nome_preenchido(request.form.get("valor")).replace(",", ".")
    moeda_codigo = _nome_preenchido(request.form.get("moeda_codigo")).upper() or "BRL"
    anexo_recibo_salvo = _nome_preenchido(request.form.get("anexo_recibo_salvo"))

    try:
        valor = float(valor_txt)
    except ValueError:
        flash("VALOR INVÁLIDO.", "warning")
        return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

    engine = get_engine()
    with engine.connect() as conn:
        duplicadas = conn.execute(text("""
            SELECT *
            FROM (
                SELECT
                    'OM' AS origem,
                    om.numero_om AS origem_numero,
                    l.id,
                    TO_CHAR(l.data_lancamento, 'DD/MM/YYYY') AS data,
                    COALESCE(l.recibo, l.id) AS recibo,
                    UPPER(COALESCE(l.tipo_linha, '')) AS descricao,
                    UPPER(COALESCE(l.detalhes, '')) AS detalhes,
                    UPPER(COALESCE(l.categoria, '')) AS categoria,
                    UPPER(COALESCE(l.aplicacao, '')) AS aplicacao,
                    COALESCE(l.valor, 0) AS valor,
                    COALESCE(l.anexo_recibo, '') AS anexo_recibo,
                    'om_recibos' AS pasta_recibo
                FROM financeiro2_om_linhas l
                JOIN financeiro2_om om ON om.id = l.om_id
                WHERE l.data_lancamento = :data_lancamento
                  AND l.valor = :valor
                  AND COALESCE(l.status, 'Ativo') = 'Ativo'

                UNION ALL

                SELECT
                    'RD' AS origem,
                    rd.numero_rd AS origem_numero,
                    l.id,
                    TO_CHAR(l.data_lancamento, 'DD/MM/YYYY') AS data,
                    NULL AS recibo,
                    UPPER(COALESCE(l.descricao, '')) AS descricao,
                    '' AS detalhes,
                    UPPER(COALESCE(l.categoria, '')) AS categoria,
                    UPPER(COALESCE(l.aplicacao, '')) AS aplicacao,
                    COALESCE(l.valor, 0) AS valor,
                    COALESCE(l.anexo_recibo, '') AS anexo_recibo,
                    'rd_recibos' AS pasta_recibo
                FROM financeiro2_rd_linhas l
                JOIN financeiro2_rd rd ON rd.id = l.rd_id
                WHERE l.data_lancamento = :data_lancamento
                  AND l.valor = :valor
                  AND COALESCE(l.status, 'Ativo') = 'Ativo'
            ) x
            ORDER BY origem, origem_numero, id
        """), {
            "data_lancamento": data_lancamento,
            "valor": valor
        }).mappings().all()

    if not duplicadas:
        return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

    return render_template(
        "financeiro_dois/om_confirmar_duplicidade.html",
        subnav_links=build_financeiro_dois_subnav("om"),
        om_id=om_id,
        duplicadas=duplicadas,
        form_data={
            "data_lancamento": data_lancamento,
            "descricao": descricao,
            "detalhes": detalhes,
            "categoria": categoria,
            "aplicacao": aplicacao,
            "valor": valor_txt,
            "moeda_codigo": moeda_codigo,
            "anexo_recibo_salvo": anexo_recibo_salvo,
        }
    )

@bp.route("/rd/<int:rd_id>/linhas/confirmar-duplicidade", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def rd_linha_confirmar_duplicidade(rd_id: int):
    data_lancamento = _nome_preenchido(request.form.get("data_lancamento"))
    descricao = _nome_preenchido(request.form.get("descricao")).upper()
    categoria = _nome_preenchido(request.form.get("categoria")).upper()
    aplicacao = _nome_preenchido(request.form.get("aplicacao")).upper()
    valor_txt = _nome_preenchido(request.form.get("valor")).replace(",", ".")
    anexo_recibo_salvo = _nome_preenchido(request.form.get("anexo_recibo_salvo"))

    try:
        valor = float(valor_txt)
    except ValueError:
        flash("VALOR INVÁLIDO.", "warning")
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    engine = get_engine()
    with engine.connect() as conn:
        duplicadas = conn.execute(text("""
            SELECT *
            FROM (
                SELECT
                    'RD' AS origem,
                    rd.numero_rd AS origem_numero,
                    l.id,
                    TO_CHAR(l.data_lancamento, 'DD/MM/YYYY') AS data,
                    NULL AS recibo,
                    UPPER(COALESCE(l.descricao, '')) AS descricao,
                    '' AS detalhes,
                    UPPER(COALESCE(l.categoria, '')) AS categoria,
                    UPPER(COALESCE(l.aplicacao, '')) AS aplicacao,
                    COALESCE(l.valor, 0) AS valor,
                    COALESCE(l.anexo_recibo, '') AS anexo_recibo,
                    'rd_recibos' AS pasta_recibo
                FROM financeiro2_rd_linhas l
                JOIN financeiro2_rd rd ON rd.id = l.rd_id
                WHERE l.data_lancamento = :data_lancamento
                  AND l.valor = :valor
                  AND COALESCE(l.status, 'Ativo') = 'Ativo'

                UNION ALL

                SELECT
                    'OM' AS origem,
                    om.numero_om AS origem_numero,
                    l.id,
                    TO_CHAR(l.data_lancamento, 'DD/MM/YYYY') AS data,
                    COALESCE(l.recibo, l.id) AS recibo,
                    UPPER(COALESCE(l.tipo_linha, '')) AS descricao,
                    UPPER(COALESCE(l.detalhes, '')) AS detalhes,
                    UPPER(COALESCE(l.categoria, '')) AS categoria,
                    UPPER(COALESCE(l.aplicacao, '')) AS aplicacao,
                    COALESCE(l.valor, 0) AS valor,
                    COALESCE(l.anexo_recibo, '') AS anexo_recibo,
                    'om_recibos' AS pasta_recibo
                FROM financeiro2_om_linhas l
                JOIN financeiro2_om om ON om.id = l.om_id
                WHERE l.data_lancamento = :data_lancamento
                  AND l.valor = :valor
                  AND COALESCE(l.status, 'Ativo') = 'Ativo'
            ) x
            ORDER BY origem, origem_numero, id
        """), {
            "data_lancamento": data_lancamento,
            "valor": valor
        }).mappings().all()

    if not duplicadas:
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    return render_template(
        "financeiro_dois/rd_confirmar_duplicidade.html",
        subnav_links=build_financeiro_dois_subnav("rd"),
        rd_id=rd_id,
        duplicadas=duplicadas,
        form_data={
            "data_lancamento": data_lancamento,
            "descricao": descricao,
            "categoria": categoria,
            "aplicacao": aplicacao,
            "valor": valor_txt,
            "anexo_recibo_salvo": anexo_recibo_salvo,
        }
    )
# =========================
# DESPESAS
# =========================

@bp.route("/despesas")
@login_required
@permission_required("financeiro", "visualizar")
def despesas():
    busca = _nome_preenchido(request.args.get("busca")).upper()
    status_despesa = _nome_preenchido(request.args.get("status_despesa")).upper()
    status_nd = _nome_preenchido(request.args.get("status_nd")).upper()
    origem = _nome_preenchido(request.args.get("origem")).upper()
    data_inicial = _nome_preenchido(request.args.get("data_inicial"))
    data_final = _nome_preenchido(request.args.get("data_final"))
    venc_inicial = _nome_preenchido(request.args.get("venc_inicial"))
    venc_final = _nome_preenchido(request.args.get("venc_final"))
    nd_numero = _nome_preenchido(request.args.get("nd_numero")).upper()
    somente_vencidas = request.args.get("somente_vencidas") == "1"

    filtros = ["1=1"]
    params = {}

    tem_filtro = any([
        busca,
        status_despesa and status_despesa != "TODOS",
        status_nd and status_nd != "TODOS",
        origem and origem not in ("TODAS", "TODOS"),
        data_inicial,
        data_final,
        venc_inicial,
        venc_final,
        nd_numero,
        somente_vencidas,
    ])

    if busca:
        filtros.append("""
            (
                UPPER(COALESCE(d.numero_despesa, '')) LIKE :busca
                OR UPPER(COALESCE(d.numero_documento, '')) LIKE :busca
                OR UPPER(COALESCE(d.fornecedor, '')) LIKE :busca
                OR UPPER(COALESCE(d.descricao, '')) LIKE :busca
            )
        """)
        params["busca"] = f"%{busca}%"

    if status_despesa and status_despesa != "TODOS":
        filtros.append("UPPER(COALESCE(d.status_despesa, '')) = :status_despesa")
        params["status_despesa"] = status_despesa

    if status_nd and status_nd != "TODOS":
        filtros.append("UPPER(COALESCE(d.status_nd, '')) = :status_nd")
        params["status_nd"] = status_nd

    if origem and origem not in ("TODAS", "TODOS"):
        filtros.append("UPPER(COALESCE(d.origem_tipo, '')) = :origem")
        params["origem"] = origem

    if data_inicial:
        filtros.append("d.data_documento >= :data_inicial")
        params["data_inicial"] = data_inicial

    if data_final:
        filtros.append("d.data_documento <= :data_final")
        params["data_final"] = data_final

    if venc_inicial:
        filtros.append("d.vencimento >= :venc_inicial")
        params["venc_inicial"] = venc_inicial

    if venc_final:
        filtros.append("d.vencimento <= :venc_final")
        params["venc_final"] = venc_final

    if nd_numero:
        filtros.append("UPPER(COALESCE(d.nd_numero, '')) LIKE :nd_numero")
        params["nd_numero"] = f"%{nd_numero}%"

    if somente_vencidas:
        filtros.append("""
            UPPER(COALESCE(d.status_despesa, 'PENDENTE')) <> 'PAGA'
            AND d.vencimento IS NOT NULL
            AND d.vencimento < CURRENT_DATE
        """)

    despesas = []
    total_registros = 0
    total_valor = 0.0

    if tem_filtro:
        engine = get_engine()
        with engine.connect() as conn:
            despesas = conn.execute(text(f"""
                SELECT
                    d.id,
                    TO_CHAR(d.data_documento, 'DD/MM/YYYY') AS data,
                    TO_CHAR(d.vencimento, 'DD/MM/YYYY') AS vencimento,
                    UPPER(COALESCE(d.tipo_documento, '')) AS tipo_documento,
                    UPPER(COALESCE(d.numero_despesa, '')) AS numero_despesa,
                    UPPER(COALESCE(d.numero_documento, '')) AS numero_documento,
                    UPPER(COALESCE(d.fornecedor, '')) AS fornecedor,
                    UPPER(COALESCE(d.descricao, '')) AS descricao,
                    UPPER(COALESCE(d.centro_custo, '')) AS centro_custo,
                    UPPER(COALESCE(d.status_despesa, '')) AS status_despesa,
                    UPPER(COALESCE(d.status_nd, '')) AS status_nd,
                    UPPER(COALESCE(d.origem_tipo, '')) AS origem,
                    UPPER(COALESCE(d.nd_numero, '')) AS nd_numero,

                    CASE
                        WHEN UPPER(COALESCE(d.origem_tipo, '')) = 'OM' THEN COALESCE((
                            SELECT SUM(
                                CASE
                                    WHEN COALESCE(l.valor_brl, 0) > 0 THEN COALESCE(l.valor_brl, 0)
                                    ELSE 0
                                END
                            )
                            FROM financeiro2_om_linhas l
                            WHERE l.om_id = d.origem_id
                              AND COALESCE(l.status, 'Ativo') = 'Ativo'
                        ), 0)

                        WHEN UPPER(COALESCE(d.origem_tipo, '')) = 'RD' THEN COALESCE((
                            SELECT SUM(
                                CASE
                                    WHEN COALESCE(l.valor, 0) > 0 THEN COALESCE(l.valor, 0)
                                    ELSE 0
                                END
                            )
                            FROM financeiro2_rd_linhas l
                            WHERE l.rd_id = d.origem_id
                              AND COALESCE(l.status, 'Ativo') = 'Ativo'
                        ), 0)

                        ELSE COALESCE(d.valor, 0)
                    END AS valor,

                    CASE
                        WHEN UPPER(COALESCE(d.status_despesa, '')) = 'PAGA' THEN 'PAGA'
                        WHEN d.vencimento IS NOT NULL AND d.vencimento < CURRENT_DATE THEN 'VENCIDA'
                        WHEN d.vencimento IS NOT NULL AND d.vencimento <= CURRENT_DATE + INTERVAL '7 day' THEN 'A VENCER'
                        ELSE 'NO PRAZO'
                    END AS situacao_vencimento

                FROM financeiro2_despesas d
                WHERE {' AND '.join(filtros)}
                ORDER BY d.data_documento DESC, d.id DESC
            """), params).mappings().all()

        total_registros = len(despesas)
        total_valor = sum(float(item["valor"] or 0) for item in despesas)

    return render_template(
        "financeiro_dois/despesas.html",
        subnav_links=build_financeiro_dois_subnav("despesas"),
        despesas=despesas,
        mostrar_resultados=tem_filtro,
        total_registros=total_registros,
        total_valor=total_valor,
        filtros={
            "busca": request.args.get("busca", ""),
            "status_despesa": request.args.get("status_despesa", "TODOS"),
            "status_nd": request.args.get("status_nd", "TODOS"),
            "origem": request.args.get("origem", "TODAS"),
            "data_inicial": request.args.get("data_inicial", ""),
            "data_final": request.args.get("data_final", ""),
            "venc_inicial": request.args.get("venc_inicial", ""),
            "venc_final": request.args.get("venc_final", ""),
            "nd_numero": request.args.get("nd_numero", ""),
            "somente_vencidas": somente_vencidas,
        }
    )
    
@bp.route("/despesas/nova")
@login_required
@permission_required("financeiro", "visualizar")
def despesa_nova():
    hoje = date.today().strftime("%Y-%m-%d")
    engine = get_engine()

    with engine.connect() as conn:
        tipos_documento = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_tipos_documento
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        centros_custo = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_centros_custo
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        empresas_nd = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_empresas_nd
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

    despesa = {
        "id": 0,
        "numero_despesa": "NOVA",
        "origem": "OPERACIONAL",
        "origem_tipo": "OPERACIONAL",
        "origem_id": None,
        "data_form": hoje,
        "vencimento_form": "",
        "tipo_documento": "",
        "numero_documento": "",
        "fornecedor": "",
        "cpf_cnpj": "",
        "descricao": "",
        "centro_custo": "",
        "fonte_pagadora": "",
        "valor": "",
        "status_despesa": "PENDENTE",
        "status_nd": "NÃO VINCULADA",
        "nd_numero": "",
        "motivo_status_nd": "",
        "observacao": "",
        "data_pagamento_form": "",
        "valor_pago": "",
        "observacao_pagamento": "",
        "tipo_registro": "OPERACIONAL",
        "eh_nova": True,
        "eh_importada": False,
        "anexos": [],
    }

    return render_template(
        "financeiro_dois/despesa_editar.html",
        subnav_links=build_financeiro_dois_subnav("despesas"),
        despesa=despesa,
        tipos_documento=tipos_documento,
        centros_custo=centros_custo,
        empresas_nd=empresas_nd,
    )
    
@bp.route("/despesas/criar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def despesa_criar():
    data_documento = _nome_preenchido(request.form.get("data_documento"))
    vencimento = _nome_preenchido(request.form.get("vencimento"))
    tipo_documento = _nome_preenchido(request.form.get("tipo_documento")).upper()
    numero_documento = _nome_preenchido(request.form.get("numero_documento")).upper()
    fornecedor = _nome_preenchido(request.form.get("fornecedor")).upper()
    cpf_cnpj = _nome_preenchido(request.form.get("cpf_cnpj")).upper()
    descricao = _nome_preenchido(request.form.get("descricao")).upper()
    centro_custo = _nome_preenchido(request.form.get("centro_custo")).upper()
    fonte_pagadora = _nome_preenchido(request.form.get("fonte_pagadora")).upper()
    valor_txt = _nome_preenchido(request.form.get("valor"))
    observacao = _nome_preenchido(request.form.get("observacao")).upper()
    status_nd = _nome_preenchido(request.form.get("status_nd")).upper() or "NÃO VINCULADA"
    nd_numero = _nome_preenchido(request.form.get("nd_numero")).upper()
    motivo_status_nd = _nome_preenchido(request.form.get("motivo_status_nd")).upper()

    engine = get_engine()
    with engine.connect() as conn:
        tipos_documento = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_tipos_documento
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        centros_custo = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_centros_custo
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        empresas_nd = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_empresas_nd
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

    despesa_form = {
        "id": 0,
        "numero_despesa": "NOVA",
        "origem": "OPERACIONAL",
        "origem_tipo": "OPERACIONAL",
        "origem_id": None,
        "data_form": data_documento,
        "vencimento_form": vencimento,
        "tipo_documento": tipo_documento,
        "numero_documento": numero_documento,
        "fornecedor": fornecedor,
        "cpf_cnpj": cpf_cnpj,
        "descricao": descricao,
        "centro_custo": centro_custo,
        "fonte_pagadora": fonte_pagadora,
        "valor": valor_txt,
        "status_despesa": "PENDENTE",
        "status_nd": status_nd,
        "nd_numero": nd_numero,
        "motivo_status_nd": motivo_status_nd,
        "observacao": observacao,
        "data_pagamento_form": "",
        "valor_pago": 0,
        "observacao_pagamento": "",
        "tipo_registro": "OPERACIONAL",
        "eh_nova": True,
        "eh_importada": False,
        "anexos": [],
        "pagamentos": [],
    }

    if not data_documento or not numero_documento or not descricao or not valor_txt:
        flash("PREENCHA DATA, NÚMERO DO DOCUMENTO, DESCRIÇÃO E VALOR.", "warning")
        return render_template(
            "financeiro_dois/despesa_editar.html",
            subnav_links=build_financeiro_dois_subnav("despesas"),
            despesa=despesa_form,
            tipos_documento=tipos_documento,
            centros_custo=centros_custo,
            empresas_nd=empresas_nd,
        )

    try:
        valor = _valor_decimal(valor_txt)
    except ValueError:
        flash("VALOR INVÁLIDO.", "warning")
        return render_template(
            "financeiro_dois/despesa_editar.html",
            subnav_links=build_financeiro_dois_subnav("despesas"),
            despesa=despesa_form,
            tipos_documento=tipos_documento,
            centros_custo=centros_custo,
            empresas_nd=empresas_nd,
        )

    if valor <= 0:
        flash("O VALOR DA DESPESA DEVE SER MAIOR QUE ZERO.", "warning")
        return render_template(
            "financeiro_dois/despesa_editar.html",
            subnav_links=build_financeiro_dois_subnav("despesas"),
            despesa=despesa_form,
            tipos_documento=tipos_documento,
            centros_custo=centros_custo,
            empresas_nd=empresas_nd,
        )

    with engine.begin() as conn:
        numero_despesa = _proximo_numero_despesa(conn)

        novo_id = conn.execute(text("""
            INSERT INTO financeiro2_despesas (
                numero_despesa,
                tipo_registro,
                origem_tipo,
                origem_id,
                data_documento,
                vencimento,
                tipo_documento,
                numero_documento,
                fornecedor,
                cpf_cnpj,
                descricao,
                centro_custo,
                fonte_pagadora,
                valor,
                status_despesa,
                status_nd,
                nd_numero,
                motivo_status_nd,
                observacao,
                valor_pago,
                observacao_pagamento,
                criado_em,
                atualizado_em
            ) VALUES (
                :numero_despesa,
                'OPERACIONAL',
                'OPERACIONAL',
                NULL,
                :data_documento,
                :vencimento,
                :tipo_documento,
                :numero_documento,
                :fornecedor,
                :cpf_cnpj,
                :descricao,
                :centro_custo,
                :fonte_pagadora,
                :valor,
                'PENDENTE',
                :status_nd,
                :nd_numero,
                :motivo_status_nd,
                :observacao,
                0,
                '',
                CURRENT_TIMESTAMP,
                CURRENT_TIMESTAMP
            )
            RETURNING id
        """), {
            "numero_despesa": numero_despesa,
            "data_documento": data_documento,
            "vencimento": vencimento or None,
            "tipo_documento": tipo_documento,
            "numero_documento": numero_documento,
            "fornecedor": fornecedor,
            "cpf_cnpj": cpf_cnpj,
            "descricao": descricao,
            "centro_custo": centro_custo,
            "fonte_pagadora": fonte_pagadora,
            "valor": valor,
            "status_nd": status_nd,
            "nd_numero": nd_numero,
            "motivo_status_nd": motivo_status_nd,
            "observacao": observacao,
        }).scalar()

        arquivo = request.files.get("anexo")
        nome_salvo, nome_original = _salvar_anexo_despesa(arquivo)
        if nome_salvo:
            conn.execute(text("""
                INSERT INTO financeiro2_despesas_anexos (
                    despesa_id, arquivo, nome_original, criado_em
                ) VALUES (
                    :despesa_id, :arquivo, :nome_original, CURRENT_TIMESTAMP
                )
            """), {
                "despesa_id": novo_id,
                "arquivo": nome_salvo,
                "nome_original": nome_original,
            })

    flash("DESPESA OPERACIONAL CRIADA COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.despesa_editar", despesa_id=novo_id))

@bp.route("/despesas/<int:despesa_id>")
@login_required
@permission_required("financeiro", "visualizar")
def despesa_editar(despesa_id: int):
    engine = get_engine()
    with engine.connect() as conn:
        despesa = conn.execute(text("""
            SELECT
                d.id,
                UPPER(COALESCE(d.numero_despesa, '')) AS numero_despesa,
                UPPER(COALESCE(d.tipo_registro, '')) AS tipo_registro,
                UPPER(COALESCE(d.origem_tipo, '')) AS origem_tipo,
                d.origem_id,
                CASE
                    WHEN d.origem_tipo = 'OM' THEN (
                        SELECT UPPER(COALESCE(om.numero_om, ''))
                        FROM financeiro2_om om
                        WHERE om.id = d.origem_id
                    )
                    WHEN d.origem_tipo = 'RD' THEN (
                        SELECT UPPER(COALESCE(rd.numero_rd, ''))
                        FROM financeiro2_rd rd
                        WHERE rd.id = d.origem_id
                    )
                    ELSE ''
                END AS origem_numero,
                TO_CHAR(d.data_documento, 'YYYY-MM-DD') AS data_form,
                TO_CHAR(d.vencimento, 'YYYY-MM-DD') AS vencimento_form,
                TO_CHAR(d.data_pagamento, 'YYYY-MM-DD') AS data_pagamento_form,
                TO_CHAR(d.data_documento, 'DD/MM/YYYY') AS data,
                TO_CHAR(d.vencimento, 'DD/MM/YYYY') AS vencimento,
                UPPER(COALESCE(d.tipo_documento, '')) AS tipo_documento,
                UPPER(COALESCE(d.numero_documento, '')) AS numero_documento,
                UPPER(COALESCE(d.fornecedor, '')) AS fornecedor,
                UPPER(COALESCE(d.cpf_cnpj, '')) AS cpf_cnpj,
                UPPER(COALESCE(d.descricao, '')) AS descricao,
                UPPER(COALESCE(d.centro_custo, '')) AS centro_custo,
                UPPER(COALESCE(d.fonte_pagadora, '')) AS fonte_pagadora,
                COALESCE(d.valor, 0) AS valor,
                UPPER(COALESCE(d.status_despesa, '')) AS status_despesa,
                UPPER(COALESCE(d.status_nd, '')) AS status_nd,
                UPPER(COALESCE(d.nd_numero, '')) AS nd_numero,
                UPPER(COALESCE(d.motivo_status_nd, '')) AS motivo_status_nd,
                UPPER(COALESCE(d.observacao, '')) AS observacao,
                COALESCE(d.valor_pago, 0) AS valor_pago,
                UPPER(COALESCE(d.observacao_pagamento, '')) AS observacao_pagamento
            FROM financeiro2_despesas d
            WHERE d.id = :id
        """), {"id": despesa_id}).mappings().first()

        if not despesa:
            abort(404)

        anexos = conn.execute(text("""
            SELECT
                id,
                arquivo,
                nome_original
            FROM financeiro2_despesas_anexos
            WHERE despesa_id = :despesa_id
            ORDER BY id DESC
        """), {"despesa_id": despesa_id}).mappings().all()

        pagamentos = conn.execute(text("""
            SELECT
                id,
                TO_CHAR(data_pagamento, 'DD/MM/YYYY') AS data_pagamento,
                COALESCE(valor_pago, 0) AS valor_pago,
                UPPER(COALESCE(observacao, '')) AS observacao
            FROM financeiro2_despesas_pagamentos
            WHERE despesa_id = :despesa_id
            ORDER BY data_pagamento DESC, id DESC
        """), {"despesa_id": despesa_id}).mappings().all()

        tipos_documento = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_tipos_documento
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        centros_custo = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_centros_custo
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        empresas_nd = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_empresas_nd
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        despesa = dict(despesa)
        despesa["origem"] = despesa["origem_tipo"]
        despesa["eh_nova"] = False
        despesa["eh_importada"] = despesa["origem_tipo"] in ("OM", "RD")
        despesa["anexos"] = anexos
        despesa["pagamentos"] = pagamentos

        # Recalcula dinamicamente valores de despesas importadas
        if despesa["eh_importada"] and despesa["origem_tipo"] == "OM":
            totais = conn.execute(text("""
                SELECT
                    COALESCE(SUM(
                        CASE
                            WHEN COALESCE(l.valor_brl, 0) > 0 THEN COALESCE(l.valor_brl, 0)
                            ELSE 0
                        END
                    ), 0) AS valor_total,
                    COALESCE(SUM(
                        CASE
                            WHEN COALESCE(l.valor_brl, 0) < 0 THEN ABS(COALESCE(l.valor_brl, 0))
                            ELSE 0
                        END
                    ), 0) AS valor_pago_total
                FROM financeiro2_om_linhas l
                WHERE l.om_id = :origem_id
                  AND COALESCE(l.status, 'Ativo') = 'Ativo'
            """), {"origem_id": despesa["origem_id"]}).mappings().first()

            despesa["valor"] = float(totais["valor_total"] or 0)
            despesa["valor_pago"] = float(totais["valor_pago_total"] or totais["valor_total"] or 0)

        elif despesa["eh_importada"] and despesa["origem_tipo"] == "RD":
            totais = conn.execute(text("""
                SELECT
                    COALESCE(SUM(
                        CASE
                            WHEN COALESCE(l.valor, 0) > 0 THEN COALESCE(l.valor, 0)
                            ELSE 0
                        END
                    ), 0) AS valor_total,
                    COALESCE(SUM(
                        CASE
                            WHEN COALESCE(l.valor, 0) < 0 THEN ABS(COALESCE(l.valor, 0))
                            ELSE 0
                        END
                    ), 0) AS valor_pago_total
                FROM financeiro2_rd_linhas l
                WHERE l.rd_id = :origem_id
                  AND COALESCE(l.status, 'Ativo') = 'Ativo'
            """), {"origem_id": despesa["origem_id"]}).mappings().first()

            despesa["valor"] = float(totais["valor_total"] or 0)
            despesa["valor_pago"] = float(totais["valor_pago_total"] or totais["valor_total"] or 0)

    return render_template(
        "financeiro_dois/despesa_editar.html",
        subnav_links=build_financeiro_dois_subnav("despesas"),
        despesa=despesa,
        tipos_documento=tipos_documento,
        centros_custo=centros_custo,
        empresas_nd=empresas_nd,
    )
    
@bp.route("/despesas/<int:despesa_id>/salvar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def despesa_salvar(despesa_id: int):
    data_documento = _nome_preenchido(request.form.get("data_documento"))
    vencimento = _nome_preenchido(request.form.get("vencimento"))
    tipo_documento = _nome_preenchido(request.form.get("tipo_documento")).upper()
    numero_documento = _nome_preenchido(request.form.get("numero_documento")).upper()
    fornecedor = _nome_preenchido(request.form.get("fornecedor")).upper()
    cpf_cnpj = _nome_preenchido(request.form.get("cpf_cnpj")).upper()
    descricao = _nome_preenchido(request.form.get("descricao")).upper()
    centro_custo = _nome_preenchido(request.form.get("centro_custo")).upper()
    fonte_pagadora = _nome_preenchido(request.form.get("fonte_pagadora")).upper()
    status_despesa = _nome_preenchido(request.form.get("status_despesa")).upper() or "PENDENTE"
    status_nd = _nome_preenchido(request.form.get("status_nd")).upper() or "NÃO VINCULADA"
    nd_numero = _nome_preenchido(request.form.get("nd_numero")).upper()
    motivo_status_nd = _nome_preenchido(request.form.get("motivo_status_nd")).upper()
    observacao = _nome_preenchido(request.form.get("observacao")).upper()
    valor_txt = _nome_preenchido(request.form.get("valor"))

    if not data_documento or not numero_documento or not descricao or not valor_txt:
        flash("PREENCHA DATA, NÚMERO DO DOCUMENTO, DESCRIÇÃO E VALOR.", "warning")
        return redirect(url_for("financeiro_dois.despesa_editar", despesa_id=despesa_id))

    try:
        valor = _valor_decimal(valor_txt)
    except ValueError:
        flash("VALOR INVÁLIDO.", "warning")
        return redirect(url_for("financeiro_dois.despesa_editar", despesa_id=despesa_id))

    engine = get_engine()
    with engine.begin() as conn:
        atual = conn.execute(text("""
            SELECT origem_tipo
            FROM financeiro2_despesas
            WHERE id = :id
        """), {"id": despesa_id}).mappings().first()

        if not atual:
            abort(404)

        if str(atual["origem_tipo"]).upper() in ("OM", "RD"):
            conn.execute(text("""
                UPDATE financeiro2_despesas
                SET
                    tipo_documento = :tipo_documento,
                    centro_custo = :centro_custo,
                    fonte_pagadora = :fonte_pagadora,
                    status_nd = :status_nd,
                    nd_numero = :nd_numero,
                    motivo_status_nd = :motivo_status_nd,
                    observacao = :observacao,
                    atualizado_em = CURRENT_TIMESTAMP
                WHERE id = :id
            """), {
                "id": despesa_id,
                "tipo_documento": tipo_documento,
                "centro_custo": centro_custo,
                "fonte_pagadora": fonte_pagadora,
                "status_nd": status_nd,
                "nd_numero": nd_numero,
                "motivo_status_nd": motivo_status_nd,
                "observacao": observacao,
            })

            flash("DADOS GERENCIAIS DA DESPESA IMPORTADA SALVOS COM SUCESSO.", "success")
            return redirect(url_for("financeiro_dois.despesa_editar", despesa_id=despesa_id))

        conn.execute(text("""
            UPDATE financeiro2_despesas
            SET
                data_documento = :data_documento,
                vencimento = :vencimento,
                tipo_documento = :tipo_documento,
                numero_documento = :numero_documento,
                fornecedor = :fornecedor,
                cpf_cnpj = :cpf_cnpj,
                descricao = :descricao,
                centro_custo = :centro_custo,
                fonte_pagadora = :fonte_pagadora,
                valor = :valor,
                status_despesa = :status_despesa,
                status_nd = :status_nd,
                nd_numero = :nd_numero,
                motivo_status_nd = :motivo_status_nd,
                observacao = :observacao,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :id
        """), {
            "id": despesa_id,
            "data_documento": data_documento,
            "vencimento": vencimento or None,
            "tipo_documento": tipo_documento,
            "numero_documento": numero_documento,
            "fornecedor": fornecedor,
            "cpf_cnpj": cpf_cnpj,
            "descricao": descricao,
            "centro_custo": centro_custo,
            "fonte_pagadora": fonte_pagadora,
            "valor": valor,
            "status_despesa": status_despesa,
            "status_nd": status_nd,
            "nd_numero": nd_numero,
            "motivo_status_nd": motivo_status_nd,
            "observacao": observacao,
        })

        arquivo = request.files.get("anexo")
        nome_salvo, nome_original = _salvar_anexo_despesa(arquivo)
        if nome_salvo:
            conn.execute(text("""
                INSERT INTO financeiro2_despesas_anexos (
                    despesa_id, arquivo, nome_original, criado_em
                ) VALUES (
                    :despesa_id, :arquivo, :nome_original, CURRENT_TIMESTAMP
                )
            """), {
                "despesa_id": despesa_id,
                "arquivo": nome_salvo,
                "nome_original": nome_original,
            })

    flash("DESPESA SALVA COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.despesa_editar", despesa_id=despesa_id))
    
@bp.route("/despesas/importar")
@login_required
@permission_required("financeiro", "visualizar")
def despesas_importar():
    busca = _nome_preenchido(request.args.get("busca")).upper()
    tipo_origem = _nome_preenchido(request.args.get("tipo_origem")).upper() or "TODAS"
    status_origem = _nome_preenchido(request.args.get("status_origem")).upper() or "TODOS"
    situacao_importacao = _nome_preenchido(request.args.get("situacao_importacao")).upper() or "TODAS"

    engine = get_engine()

    with engine.connect() as conn:
        oms = conn.execute(text("""
            SELECT
                'OM' AS origem_tipo,
                om.id AS origem_id,
                UPPER(COALESCE(om.numero_om, '')) AS numero_origem,
                TO_CHAR(om.criado_em, 'DD/MM/YYYY') AS data_origem,
                UPPER(COALESCE(om.nome_colaborador, '')) AS favorecido,
                CASE
                    WHEN UPPER(COALESCE(om.status, '')) = 'QUITADA' THEN 'PAGA'
                    ELSE UPPER(COALESCE(om.status, ''))
                END AS status_origem,
                COALESCE((
                    SELECT SUM(
                        CASE
                            WHEN COALESCE(l.valor_brl, 0) > 0 THEN COALESCE(l.valor_brl, 0)
                            ELSE 0
                        END
                    )
                    FROM financeiro2_om_linhas l
                    WHERE l.om_id = om.id
                      AND COALESCE(l.status, 'Ativo') = 'Ativo'
                ), 0) AS valor_total,
                CASE
                    WHEN EXISTS (
                        SELECT 1
                        FROM financeiro2_despesas d
                        WHERE d.origem_tipo = 'OM'
                          AND d.origem_id = om.id
                    ) THEN 1 ELSE 0
                END AS ja_importada
            FROM financeiro2_om om
            ORDER BY om.id DESC
        """)).mappings().all()

        rds = conn.execute(text("""
            SELECT
                'RD' AS origem_tipo,
                rd.id AS origem_id,
                UPPER(COALESCE(rd.numero_rd, '')) AS numero_origem,
                TO_CHAR(rd.criado_em, 'DD/MM/YYYY') AS data_origem,
                UPPER(COALESCE(rd.nome_colaborador, '')) AS favorecido,
                CASE
                    WHEN UPPER(COALESCE(rd.status, '')) = 'QUITADA' THEN 'PAGA'
                    ELSE UPPER(COALESCE(rd.status, ''))
                END AS status_origem,
                COALESCE((
                    SELECT SUM(
                        CASE
                            WHEN COALESCE(l.valor, 0) > 0 THEN COALESCE(l.valor, 0)
                            ELSE 0
                        END
                    )
                    FROM financeiro2_rd_linhas l
                    WHERE l.rd_id = rd.id
                      AND COALESCE(l.status, 'Ativo') = 'Ativo'
                ), 0) AS valor_total,
                CASE
                    WHEN EXISTS (
                        SELECT 1
                        FROM financeiro2_despesas d
                        WHERE d.origem_tipo = 'RD'
                          AND d.origem_id = rd.id
                    ) THEN 1 ELSE 0
                END AS ja_importada
            FROM financeiro2_rd rd
            ORDER BY rd.id DESC
        """)).mappings().all()

    itens = []
    for item in list(oms) + list(rds):
        item = dict(item)

        status = item["status_origem"] or ""
        if status in ("ABERTA", "ABERTO"):
            status = "EM ABERTO"
        item["status_origem"] = status

        item["pode_importar"] = item["status_origem"] == "PAGA" and not bool(item["ja_importada"])

        if busca:
            alvo = " ".join([
                str(item["numero_origem"] or ""),
                str(item["favorecido"] or ""),
                str(item["origem_tipo"] or "")
            ]).upper()
            if busca not in alvo:
                continue

        if tipo_origem != "TODAS" and item["origem_tipo"] != tipo_origem:
            continue

        if status_origem != "TODOS" and item["status_origem"] != status_origem:
            continue

        if situacao_importacao == "LIBERADAS" and not item["pode_importar"]:
            continue
        if situacao_importacao == "BLOQUEADAS" and item["pode_importar"]:
            continue
        if situacao_importacao == "JA IMPORTADAS" and not bool(item["ja_importada"]):
            continue
        if situacao_importacao == "NAO IMPORTADAS" and bool(item["ja_importada"]):
            continue

        itens.append(item)

    itens.sort(key=lambda x: (x["origem_tipo"], x["numero_origem"]), reverse=True)

    total_registros = len(itens)
    total_liberadas = sum(1 for i in itens if i["pode_importar"])
    total_bloqueadas = sum(1 for i in itens if not i["pode_importar"] and not i["ja_importada"])
    total_importadas = sum(1 for i in itens if i["ja_importada"])
    total_valor = sum(float(i["valor_total"] or 0) for i in itens)

    return render_template(
        "financeiro_dois/despesas_importar.html",
        subnav_links=build_financeiro_dois_subnav("despesas"),
        itens=itens,
        total_registros=total_registros,
        total_liberadas=total_liberadas,
        total_bloqueadas=total_bloqueadas,
        total_importadas=total_importadas,
        total_valor=total_valor,
        filtros={
            "busca": request.args.get("busca", ""),
            "tipo_origem": tipo_origem,
            "status_origem": status_origem,
            "situacao_importacao": situacao_importacao,
        }
    )
    
@bp.route("/despesas/importar/<string:origem_tipo>/<int:origem_id>", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def despesa_importar_origem(origem_tipo: str, origem_id: int):
    origem_tipo = (origem_tipo or "").upper().strip()
    if origem_tipo not in ("OM", "RD"):
        flash("ORIGEM INVÁLIDA PARA IMPORTAÇÃO.", "warning")
        return redirect(url_for("financeiro_dois.despesas_importar"))

    engine = get_engine()
    with engine.begin() as conn:
        existente = conn.execute(text("""
            SELECT id
            FROM financeiro2_despesas
            WHERE origem_tipo = :origem_tipo
              AND origem_id = :origem_id
            LIMIT 1
        """), {
            "origem_tipo": origem_tipo,
            "origem_id": origem_id
        }).mappings().first()

        if existente:
            flash("ESSA ORIGEM JÁ FOI IMPORTADA COMO DESPESA.", "warning")
            return redirect(url_for("financeiro_dois.despesas_importar"))

        if origem_tipo == "OM":
            origem = conn.execute(text("""
                SELECT
                    id,
                    numero_om AS numero_origem,
                    criado_em::date AS data_documento,
                    UPPER(COALESCE(nome_colaborador, '')) AS fornecedor,
                    UPPER(COALESCE(matricula_colaborador, '')) AS cpf_cnpj,
                    CASE
                        WHEN UPPER(COALESCE(status, '')) = 'QUITADA' THEN 'PAGA'
                        ELSE UPPER(COALESCE(status, ''))
                    END AS status_origem,

                    COALESCE((
                        SELECT SUM(
                            CASE
                                WHEN COALESCE(l.valor_brl, 0) > 0 THEN COALESCE(l.valor_brl, 0)
                                ELSE 0
                            END
                        )
                        FROM financeiro2_om_linhas l
                        WHERE l.om_id = financeiro2_om.id
                          AND COALESCE(l.status, 'Ativo') = 'Ativo'
                    ), 0) AS valor_total,

                    COALESCE((
                        SELECT SUM(
                            CASE
                                WHEN COALESCE(l.valor_brl, 0) < 0 THEN ABS(COALESCE(l.valor_brl, 0))
                                ELSE 0
                            END
                        )
                        FROM financeiro2_om_linhas l
                        WHERE l.om_id = financeiro2_om.id
                          AND COALESCE(l.status, 'Ativo') = 'Ativo'
                    ), 0) AS valor_pago_total
                FROM financeiro2_om
                WHERE id = :id
            """), {"id": origem_id}).mappings().first()

            tipo_documento = "OM"
            numero_documento = origem["numero_origem"] if origem else ""
            descricao = f"IMPORTAÇÃO DA OM {numero_documento}"

        else:
            origem = conn.execute(text("""
                SELECT
                    id,
                    numero_rd AS numero_origem,
                    criado_em::date AS data_documento,
                    UPPER(COALESCE(nome_colaborador, '')) AS fornecedor,
                    UPPER(COALESCE(matricula_colaborador, '')) AS cpf_cnpj,
                    CASE
                        WHEN UPPER(COALESCE(status, '')) = 'QUITADA' THEN 'PAGA'
                        ELSE UPPER(COALESCE(status, ''))
                    END AS status_origem,

                    COALESCE((
                        SELECT SUM(
                            CASE
                                WHEN COALESCE(l.valor, 0) > 0 THEN COALESCE(l.valor, 0)
                                ELSE 0
                            END
                        )
                        FROM financeiro2_rd_linhas l
                        WHERE l.rd_id = financeiro2_rd.id
                          AND COALESCE(l.status, 'Ativo') = 'Ativo'
                    ), 0) AS valor_total,

                    COALESCE((
                        SELECT SUM(
                            CASE
                                WHEN COALESCE(l.valor, 0) < 0 THEN ABS(COALESCE(l.valor, 0))
                                ELSE 0
                            END
                        )
                        FROM financeiro2_rd_linhas l
                        WHERE l.rd_id = financeiro2_rd.id
                          AND COALESCE(l.status, 'Ativo') = 'Ativo'
                    ), 0) AS valor_pago_total
                FROM financeiro2_rd
                WHERE id = :id
            """), {"id": origem_id}).mappings().first()

            tipo_documento = "RD"
            numero_documento = origem["numero_origem"] if origem else ""
            descricao = f"IMPORTAÇÃO DA RD {numero_documento}"

        if not origem:
            flash("ORIGEM NÃO ENCONTRADA.", "danger")
            return redirect(url_for("financeiro_dois.despesas_importar"))

        if origem["status_origem"] != "PAGA":
            flash("A IMPORTAÇÃO SÓ É PERMITIDA QUANDO A ORIGEM ESTIVER PAGA.", "warning")
            return redirect(url_for("financeiro_dois.despesas_importar"))

        numero_despesa = _proximo_numero_despesa(conn, origem["data_documento"])

        conn.execute(text("""
            INSERT INTO financeiro2_despesas (
                numero_despesa,
                tipo_registro,
                origem_tipo,
                origem_id,
                data_documento,
                vencimento,
                tipo_documento,
                numero_documento,
                fornecedor,
                cpf_cnpj,
                descricao,
                centro_custo,
                fonte_pagadora,
                valor,
                status_despesa,
                status_nd,
                nd_numero,
                motivo_status_nd,
                observacao,
                data_pagamento,
                valor_pago,
                observacao_pagamento,
                importada_em,
                criado_em,
                atualizado_em
            ) VALUES (
                :numero_despesa,
                'IMPORTADA',
                :origem_tipo,
                :origem_id,
                :data_documento,
                NULL,
                :tipo_documento,
                :numero_documento,
                :fornecedor,
                :cpf_cnpj,
                :descricao,
                '',
                '',
                :valor,
                'PAGA',
                'NÃO VINCULADA',
                '',
                '',
                '',
                :data_pagamento,
                :valor_pago,
                'IMPORTADA AUTOMATICAMENTE DA ORIGEM',
                CURRENT_TIMESTAMP,
                CURRENT_TIMESTAMP,
                CURRENT_TIMESTAMP
            )
        """), {
            "numero_despesa": numero_despesa,
            "origem_tipo": origem_tipo,
            "origem_id": origem_id,
            "data_documento": origem["data_documento"],
            "tipo_documento": tipo_documento,
            "numero_documento": numero_documento,
            "fornecedor": origem["fornecedor"],
            "cpf_cnpj": origem["cpf_cnpj"],
            "descricao": descricao,
            "valor": float(origem["valor_total"] or 0),
            "data_pagamento": origem["data_documento"],
            "valor_pago": float(origem["valor_pago_total"] or origem["valor_total"] or 0),
        })

    flash(f"{origem_tipo} IMPORTADA COMO DESPESA COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.despesas"))

@bp.route("/rd/<int:rd_id>/adiantar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def rd_adiantar(rd_id: int):
    data_lancamento = _nome_preenchido(request.form.get("data_lancamento"))
    aplicacao = _nome_preenchido(request.form.get("aplicacao")).upper()
    valor_txt = _nome_preenchido(request.form.get("valor")).replace(",", ".")

    if not data_lancamento or not aplicacao or not valor_txt:
        flash("PREENCHA DATA, APLICAÇÃO E VALOR DO ADIANTAMENTO.", "warning")
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    try:
        valor = float(valor_txt)
    except ValueError:
        flash("VALOR INVÁLIDO PARA O ADIANTAMENTO.", "warning")
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    if valor <= 0:
        flash("O ADIANTAMENTO DEVE SER MAIOR QUE ZERO.", "warning")
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    engine = get_engine()
    with engine.begin() as conn:
        rd = conn.execute(text("""
            SELECT id, numero_rd, status
            FROM financeiro2_rd
            WHERE id = :id
        """), {"id": rd_id}).mappings().first()

        if not rd:
            flash("RD NÃO ENCONTRADA.", "danger")
            return redirect(url_for("financeiro_dois.rd"))

        if str(rd["status"]).upper() == "QUITADA":
            flash("ESTA RD ESTÁ QUITADA E BLOQUEADA PARA EDIÇÃO.", "warning")
            return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

        conn.execute(text("""
            INSERT INTO financeiro2_rd_linhas (
                rd_id,
                data_lancamento,
                descricao,
                categoria,
                aplicacao,
                valor,
                status
            )
            VALUES (
                :rd_id,
                :data_lancamento,
                :descricao,
                :categoria,
                :aplicacao,
                :valor,
                'Ativo'
            )
        """), {
            "rd_id": rd_id,
            "data_lancamento": data_lancamento,
            "descricao": f"ADIANTAMENTO DA RD ({rd['numero_rd']})",
            "categoria": "ADIANTAMENTO",
            "aplicacao": aplicacao,
            "valor": -abs(valor),
        })

        conn.execute(text("""
            UPDATE financeiro2_rd
            SET status = 'PARCIAL',
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :id
              AND UPPER(status) = 'ABERTA'
        """), {"id": rd_id})

    flash("ADIANTAMENTO DA RD REGISTRADO COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))
    
@bp.route("/rd/<int:rd_id>/pagar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def rd_pagar(rd_id: int):
    data_lancamento = _nome_preenchido(request.form.get("data_lancamento"))
    aplicacao = (_nome_preenchido(request.form.get("aplicacao")) or "GERAL").upper()

    if not data_lancamento:
        flash("INFORME A DATA DO PAGAMENTO.", "warning")
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    engine = get_engine()
    with engine.begin() as conn:
        rd = conn.execute(text("""
            SELECT id, numero_rd, status
            FROM financeiro2_rd
            WHERE id = :id
        """), {"id": rd_id}).mappings().first()

        if not rd:
            flash("RD NÃO ENCONTRADA.", "danger")
            return redirect(url_for("financeiro_dois.rd"))

        if str(rd["status"]).upper() == "QUITADA":
            flash("ESTA RD JÁ ESTÁ QUITADA.", "warning")
            return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

        saldo_atual = _calcular_saldo_rd(conn, rd_id)

        if saldo_atual <= 0:
            flash("A RD NÃO POSSUI SALDO POSITIVO PARA PAGAMENTO.", "warning")
            return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

        conn.execute(text("""
            INSERT INTO financeiro2_rd_linhas (
                rd_id,
                data_lancamento,
                descricao,
                categoria,
                aplicacao,
                valor,
                status
            )
            VALUES (
                :rd_id,
                :data_lancamento,
                :descricao,
                :categoria,
                :aplicacao,
                :valor,
                'Ativo'
            )
        """), {
            "rd_id": rd_id,
            "data_lancamento": data_lancamento,
            "descricao": f"PAGAMENTO DA RD ({rd['numero_rd']})",
            "categoria": "PAGAMENTO",
            "aplicacao": aplicacao,
            "valor": -abs(saldo_atual),
        })

        conn.execute(text("""
            UPDATE financeiro2_rd
            SET status = 'QUITADA',
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :id
        """), {"id": rd_id})

    flash("PAGAMENTO DA RD REGISTRADO COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))
    
@bp.route("/rd/<int:rd_id>/exportar/excel")
@login_required
@permission_required("financeiro", "visualizar")
def rd_exportar_excel(rd_id: int):
    engine = get_engine()

    with engine.connect() as conn:
        rd = conn.execute(text("""
            SELECT
                id,
                numero_rd,
                periodo,
                matricula_colaborador,
                nome_colaborador,
                centro_custo,
                status,
                COALESCE(observacao, '') AS observacao
            FROM financeiro2_rd
            WHERE id = :id
        """), {"id": rd_id}).mappings().first()

        if not rd:
            abort(404)

        linhas = conn.execute(text("""
            SELECT
                TO_CHAR(data_lancamento, 'DD/MM/YYYY') AS data,
                descricao,
                COALESCE(categoria, '') AS categoria,
                COALESCE(aplicacao, '') AS aplicacao,
                COALESCE(valor, 0) AS valor,
                COALESCE(status, 'Ativo') AS status
            FROM financeiro2_rd_linhas
            WHERE rd_id = :id
              AND COALESCE(status, 'Ativo') = 'Ativo'
            ORDER BY id
        """), {"id": rd_id}).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "RD"

    ws.append(["Número RD", rd["numero_rd"]])
    ws.append(["Período", rd["periodo"]])
    ws.append(["Matrícula", rd["matricula_colaborador"]])
    ws.append(["Colaborador", rd["nome_colaborador"]])
    ws.append(["Centro de custo", rd["centro_custo"]])
    ws.append(["Status", rd["status"]])
    ws.append(["Observação", rd["observacao"]])
    ws.append([])
    ws.append(["Data", "Descrição", "Categoria", "Aplicação", "Valor", "Status"])

    total = 0
    for linha in linhas:
        ws.append([
            linha["data"],
            linha["descricao"],
            linha["categoria"],
            linha["aplicacao"],
            float(linha["valor"]),
            linha["status"],
        ])
        if linha["status"] == "Ativo":
            total += float(linha["valor"])

    ws.append([])
    ws.append(["", "", "", "Saldo", total, ""])

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"{rd['numero_rd']}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    
@bp.route("/rd/<int:rd_id>/exportar/pdf")
@login_required
@permission_required("financeiro", "visualizar")
def rd_exportar_pdf(rd_id: int):
    engine = get_engine()

    with engine.connect() as conn:
        rd = conn.execute(text("""
            SELECT
                id,
                numero_rd,
                periodo,
                matricula_colaborador,
                nome_colaborador,
                centro_custo,
                status,
                COALESCE(observacao, '') AS observacao
            FROM financeiro2_rd
            WHERE id = :id
        """), {"id": rd_id}).mappings().first()

        if not rd:
            abort(404)

        linhas = conn.execute(text("""
            SELECT
                id,
                TO_CHAR(data_lancamento, 'DD/MM/YYYY') AS data,
                descricao,
                COALESCE(categoria, '') AS categoria,
                COALESCE(aplicacao, '') AS aplicacao,
                COALESCE(valor, 0) AS valor,
                COALESCE(status, 'Ativo') AS status,
                COALESCE(anexo_recibo, '') AS anexo_recibo
            FROM financeiro2_rd_linhas
            WHERE rd_id = :id
              AND COALESCE(status, 'Ativo') = 'Ativo'
            ORDER BY id
        """), {"id": rd_id}).mappings().all()

    total = sum(float(l["valor"]) for l in linhas if l["status"] == "Ativo")

    buffer_base = BytesIO()
    pdf = canvas.Canvas(buffer_base, pagesize=A4)
    largura, altura = A4
    y = altura - 40

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(40, y, f"RD {rd['numero_rd']}")
    y -= 20

    pdf.setFont("Helvetica", 10)
    pdf.drawString(40, y, f"Período: {rd['periodo']}")
    y -= 15
    pdf.drawString(40, y, f"Matrícula: {rd['matricula_colaborador']}")
    y -= 15
    pdf.drawString(40, y, f"Colaborador: {rd['nome_colaborador']}")
    y -= 15
    pdf.drawString(40, y, f"Centro de custo: {rd['centro_custo']}")
    y -= 15
    pdf.drawString(40, y, f"Status: {rd['status']}")
    y -= 15
    pdf.drawString(40, y, f"Observação: {rd['observacao']}")
    y -= 25

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(40, y, "Data")
    pdf.drawString(90, y, "Descrição")
    pdf.drawString(260, y, "Categoria")
    pdf.drawString(360, y, "Aplicação")
    pdf.drawString(500, y, "Valor")
    y -= 15

    pdf.setFont("Helvetica", 8)
    for linha in linhas:
        if y < 50:
            pdf.showPage()
            y = altura - 40
            pdf.setFont("Helvetica", 8)

        pdf.drawString(40, y, str(linha["data"]))
        pdf.drawString(90, y, str(linha["descricao"])[:30])
        pdf.drawString(260, y, str(linha["categoria"])[:15])
        pdf.drawString(360, y, str(linha["aplicacao"])[:18])
        pdf.drawRightString(540, y, f"{float(linha['valor']):.2f}")
        y -= 13

    y -= 10
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawRightString(540, y, f"Saldo: {total:.2f}")
    pdf.save()
    buffer_base.seek(0)

    writer = PdfWriter()
    base_reader = PdfReader(buffer_base)
    for page in base_reader.pages:
        writer.add_page(page)

    for linha in linhas:
        nome_anexo = (linha["anexo_recibo"] or "").strip()
        if not nome_anexo:
            continue

        caminho = os.path.join(
            current_app.root_path,
            "static",
            "uploads",
            "financeiro2",
            "rd_recibos",
            nome_anexo
        )

        if not os.path.exists(caminho):
            continue

        extensao = os.path.splitext(caminho)[1].lower()

        try:
            if extensao == ".pdf":
                anexo_reader = PdfReader(caminho)
                for page in anexo_reader.pages:
                    writer.add_page(page)

            elif extensao in [".jpg", ".jpeg", ".png", ".bmp", ".webp"]:
                img = Image.open(caminho)
                if img.mode in ("RGBA", "P"):
                    img = img.convert("RGB")

                img_buffer = BytesIO()
                c = canvas.Canvas(img_buffer, pagesize=A4)
                largura_pg, altura_pg = A4

                iw, ih = img.size
                margem = 30
                area_w = largura_pg - 2 * margem
                area_h = altura_pg - 2 * margem

                escala = min(area_w / iw, area_h / ih)
                novo_w = iw * escala
                novo_h = ih * escala

                x = (largura_pg - novo_w) / 2
                y_img = (altura_pg - novo_h) / 2

                img_temp = BytesIO()
                img.save(img_temp, format="JPEG")
                img_temp.seek(0)

                c.setFont("Helvetica-Bold", 11)
                c.drawString(30, altura_pg - 20, f"RECIBO RD {rd['numero_rd']}")
                c.drawImage(ImageReader(img_temp), x, y_img, width=novo_w, height=novo_h)
                c.showPage()
                c.save()

                img_buffer.seek(0)
                img_reader = PdfReader(img_buffer)
                for page in img_reader.pages:
                    writer.add_page(page)

        except Exception:
            continue

    saida = BytesIO()
    writer.write(saida)
    saida.seek(0)

    return send_file(
        saida,
        as_attachment=True,
        download_name=f"{rd['numero_rd']}.pdf",
        mimetype="application/pdf",
    )

# =========================
# PREVISAO
# =========================

@bp.route("/previsao")
@login_required
@permission_required("financeiro", "visualizar")
def previsao():
    previsoes = [
        {"id": 1, "data": "15/03/2026", "vencimento": "20/03/2026", "tipo_documento": "NF", "numero_documento": "NF-4587", "fornecedor": "Hotel Exemplo", "descricao": "Hospedagem equipe", "centro_custo": "ADM", "valor": 950.00, "status_despesa": "Pendente", "status_nd": "Não vinculada", "motivo_status_nd": ""},
        {"id": 2, "data": "14/03/2026", "vencimento": "18/03/2026", "tipo_documento": "Fatura", "numero_documento": "FAT-9001", "fornecedor": "Posto Modelo", "descricao": "Combustível", "centro_custo": "OPERACAO", "valor": 420.50, "status_despesa": "Paga", "status_nd": "Em espera", "motivo_status_nd": "Aguardando decisão da área"},
        {"id": 3, "data": "13/03/2026", "vencimento": "25/03/2026", "tipo_documento": "NFS", "numero_documento": "NFS-1102", "fornecedor": "Serviço X", "descricao": "Serviço de apoio", "centro_custo": "MANUTENCAO", "valor": 780.30, "status_despesa": "Pendente", "status_nd": "Rejeitada", "motivo_status_nd": "Fora do escopo da ND atual"},
    ]
    return render_template("financeiro_dois/previsao.html", subnav_links=build_financeiro_dois_subnav("previsao"), previsoes=previsoes)


# =========================
# REEMBOLSOS
# =========================

@bp.route("/reembolsos")
@login_required
@permission_required("financeiro", "visualizar")
def reembolsos():
    reembolsos_lista = [
        {"id": 1, "matricula": "LME", "colaborador": "Laercio Melo", "pix": "11999999999", "data_solicitacao": "15/03/2026", "descricao": "Reembolso alimentação viagem", "valor": 180.00, "status": "Solicitado", "fonte_pagadora": "", "aprovacao": "Pendente"},
        {"id": 2, "matricula": "ABC", "colaborador": "Colaborador Exemplo", "pix": "abc@email.com", "data_solicitacao": "14/03/2026", "descricao": "Reembolso táxi", "valor": 75.50, "status": "Aprovado", "fonte_pagadora": "OM-2026-0002", "aprovacao": "Aprovado"},
        {"id": 3, "matricula": "XYZ", "colaborador": "Outro Colaborador", "pix": "123.456.789-00", "data_solicitacao": "13/03/2026", "descricao": "Reembolso combustível", "valor": 220.00, "status": "Pago", "fonte_pagadora": "RD-2026-02-XYZ", "aprovacao": "Aprovado"},
    ]
    return render_template("financeiro_dois/reembolsos.html", subnav_links=build_financeiro_dois_subnav("reembolsos"), reembolsos=reembolsos_lista)


@bp.route("/reembolsos/<int:reembolso_id>")
@login_required
@permission_required("financeiro", "visualizar")
def reembolso_editar(reembolso_id: int):
    reembolsos_map = {
        1: {
            "id": 1, "matricula": "LME", "colaborador": "Laercio Melo", "pix": "11999999999",
            "data_solicitacao": "15/03/2026", "descricao": "Reembolso alimentação viagem", "valor": 180.00,
            "status": "Solicitado", "aprovacao": "Pendente", "fonte_pagadora": "", "comprovante_solicitacao": "anexo_refeicao.pdf",
            "comprovante_pagamento": "", "observacao": "Aguardando aprovação para pagamento.",
        },
        2: {
            "id": 2, "matricula": "ABC", "colaborador": "Colaborador Exemplo", "pix": "abc@email.com",
            "data_solicitacao": "14/03/2026", "descricao": "Reembolso táxi", "valor": 75.50,
            "status": "Aprovado", "aprovacao": "Aprovado", "fonte_pagadora": "OM-2026-0002", "comprovante_solicitacao": "taxi_1403.pdf",
            "comprovante_pagamento": "", "observacao": "Pronto para pagamento.",
        },
        3: {
            "id": 3, "matricula": "XYZ", "colaborador": "Outro Colaborador", "pix": "123.456.789-00",
            "data_solicitacao": "13/03/2026", "descricao": "Reembolso combustível", "valor": 220.00,
            "status": "Pago", "aprovacao": "Aprovado", "fonte_pagadora": "RD-2026-02-XYZ", "comprovante_solicitacao": "combustivel_xyz.pdf",
            "comprovante_pagamento": "pix_220_xyz.pdf", "observacao": "Pagamento realizado com comprovante salvo.",
        },
    }

    reembolso = reembolsos_map.get(reembolso_id)
    if not reembolso:
        abort(404)

    return render_template(
        "financeiro_dois/reembolso_editar.html",
        subnav_links=build_financeiro_dois_subnav("reembolsos"),
        reembolso=reembolso,
    )


# =========================
# APROVACOES
# =========================

@bp.route("/aprovacoes")
@login_required
@permission_required("financeiro", "visualizar")
def aprovacoes():
    solicitacoes = [
        {"id": 1, "tipo": "Aprovação de reembolso", "modulo": "Reembolsos", "referencia": "REB-0001", "motivo": "Solicitação inicial de pagamento", "solicitado_por": "LME", "data_solicitacao": "15/03/2026", "status": "Pendente", "aprovado_por": "", "data_aprovacao": ""},
        {"id": 2, "tipo": "Solicitar alteração", "modulo": "Despesas", "referencia": "FAT-9001", "motivo": "Despesa importada de OM", "solicitado_por": "ABC", "data_solicitacao": "14/03/2026", "status": "Aprovado", "aprovado_por": "ADM", "data_aprovacao": "15/03/2026"},
        {"id": 3, "tipo": "Solicitar exclusão", "modulo": "OM", "referencia": "OM-2026-0002", "motivo": "Registro duplicado", "solicitado_por": "XYZ", "data_solicitacao": "13/03/2026", "status": "Recusado", "aprovado_por": "ADM", "data_aprovacao": "14/03/2026"},
    ]
    return render_template("financeiro_dois/aprovacoes.html", subnav_links=build_financeiro_dois_subnav("aprovacoes"), solicitacoes=solicitacoes)


# =========================
# NOTAS DE DEBITO
# =========================

#@bp.route("/notas-debito")
#@login_required
#@permission_required("financeiro", "visualizar")
#def notas_debito():
#    notas = [
#        {"id": 1, "numero_nd": "ND-2026-0001", "empresa_origem": "MATISA", "data_criacao": "15/03/2026", "status": "Aberta", "total": 1730.30},
#        {"id": 2, "numero_nd": "ND-2026-0002", "empresa_origem": "PRUMAT", "data_criacao": "14/03/2026", "status": "Fechada", "total": 950.00},
#        {"id": 3, "numero_nd": "ND-2026-0003", "empresa_origem": "MATISA", "data_criacao": "13/03/2026", "status": "Exportada", "total": 420.50},
#    ]
#
#    return render_template(
#        "financeiro_dois/notas_debito.html",
#        subnav_links=build_financeiro_dois_subnav("nd"),
#        notas=notas,
#    )


#@bp.route("/notas-debito/<int:nd_id>")
#@login_required
#@permission_required("financeiro", "visualizar")
#def nota_debito_editar(nd_id: int):
#    notas_map = {
#        1: {
#            "id": 1,
#            "numero_nd": "ND-2026-0001",
#            "empresa_origem": "MATISA",
#            "data_criacao": "15/03/2026",
#            "status": "Aberta",
#            "observacao": "ND em montagem com despesas ainda em análise.",
#            "linhas": [
#                {"data": "15/03/2026", "descricao": "Hospedagem equipe", "tipo": "NF", "numero_documento": "NF-4587", "valor": 950.00},
#                {"data": "14/03/2026", "descricao": "Combustível", "tipo": "Fatura", "numero_documento": "FAT-9001", "valor": 420.50},
#                {"data": "13/03/2026", "descricao": "Serviço de apoio", "tipo": "NFS", "numero_documento": "NFS-1102", "valor": 359.80},
#            ],
#        },
#        2: {
#            "id": 2,
#            "numero_nd": "ND-2026-0002",
#            "empresa_origem": "PRUMAT",
#            "data_criacao": "14/03/2026",
#            "status": "Fechada",
#            "observacao": "ND já conferida e fechada.",
#            "linhas": [
#                {"data": "14/03/2026", "descricao": "Hospedagem equipe", "tipo": "NF", "numero_documento": "NF-4587", "valor": 950.00},
#            ],
#        },
#        3: {
#            "id": 3,
#            "numero_nd": "ND-2026-0003",
#            "empresa_origem": "MATISA",
#            "data_criacao": "13/03/2026",
#            "status": "Exportada",
#            "observacao": "ND exportada em PDF e travada para edição direta.",
#            "linhas": [
#                {"data": "14/03/2026", "descricao": "Combustível", "tipo": "Fatura", "numero_documento": "FAT-9001", "valor": 420.50},
#            ],
#        },
#    }
#
#    nd = notas_map.get(nd_id)
#    if not nd:
#        abort(404)
#
#    total_nd = sum(item["valor"] for item in nd["linhas"])
#
#    return render_template(
#        "financeiro_dois/nota_debito_editar.html",
#        subnav_links=build_financeiro_dois_subnav("nd"),
#        nd=nd,
#        total_nd=total_nd,
#    )
    
@bp.route("/despesas/<int:despesa_id>/registrar-pagamento", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def despesa_registrar_pagamento(despesa_id: int):
    data_pagamento = _nome_preenchido(request.form.get("data_pagamento"))
    valor_pago_txt = _nome_preenchido(request.form.get("valor_pago"))
    observacao_pagamento = _nome_preenchido(request.form.get("observacao_pagamento")).upper()

    if not data_pagamento or not valor_pago_txt:
        flash("PREENCHA DATA E VALOR DO PAGAMENTO.", "warning")
        return redirect(url_for("financeiro_dois.despesa_editar", despesa_id=despesa_id))

    try:
        valor_pago = _valor_decimal(valor_pago_txt)
    except ValueError:
        flash("VALOR DE PAGAMENTO INVÁLIDO.", "warning")
        return redirect(url_for("financeiro_dois.despesa_editar", despesa_id=despesa_id))

    if valor_pago <= 0:
        flash("O VALOR DO PAGAMENTO DEVE SER MAIOR QUE ZERO.", "warning")
        return redirect(url_for("financeiro_dois.despesa_editar", despesa_id=despesa_id))

    engine = get_engine()
    with engine.begin() as conn:
        despesa = conn.execute(text("""
            SELECT
                id,
                UPPER(COALESCE(origem_tipo, '')) AS origem_tipo,
                COALESCE(valor, 0) AS valor
            FROM financeiro2_despesas
            WHERE id = :id
        """), {"id": despesa_id}).mappings().first()

        if not despesa:
            abort(404)

        if despesa["origem_tipo"] in ("OM", "RD"):
            flash("PAGAMENTO DE DESPESA IMPORTADA DEVE SER CONTROLADO PELA ORIGEM.", "warning")
            return redirect(url_for("financeiro_dois.despesa_editar", despesa_id=despesa_id))

        conn.execute(text("""
            INSERT INTO financeiro2_despesas_pagamentos (
                despesa_id,
                data_pagamento,
                valor_pago,
                observacao,
                criado_em
            ) VALUES (
                :despesa_id,
                :data_pagamento,
                :valor_pago,
                :observacao,
                CURRENT_TIMESTAMP
            )
        """), {
            "despesa_id": despesa_id,
            "data_pagamento": data_pagamento,
            "valor_pago": valor_pago,
            "observacao": observacao_pagamento,
        })

        totais = conn.execute(text("""
            SELECT
                COALESCE(SUM(COALESCE(valor_pago, 0)), 0) AS total_pago,
                MAX(data_pagamento) AS ultima_data
            FROM financeiro2_despesas_pagamentos
            WHERE despesa_id = :despesa_id
        """), {"despesa_id": despesa_id}).mappings().first()

        total_pago = float(totais["total_pago"] or 0)
        valor_total = float(despesa["valor"] or 0)

        if total_pago <= 0:
            status_despesa = "PENDENTE"
        elif total_pago < valor_total:
            status_despesa = "PARCIAL"
        else:
            status_despesa = "PAGA"

        conn.execute(text("""
            UPDATE financeiro2_despesas
            SET
                data_pagamento = :data_pagamento,
                valor_pago = :valor_pago,
                observacao_pagamento = :observacao_pagamento,
                status_despesa = :status_despesa,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :id
        """), {
            "id": despesa_id,
            "data_pagamento": totais["ultima_data"],
            "valor_pago": total_pago,
            "observacao_pagamento": observacao_pagamento,
            "status_despesa": status_despesa,
        })

    flash("PAGAMENTO REGISTRADO COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.despesa_editar", despesa_id=despesa_id))
    
#====================================================================================
#      EXPORTAR DESPESAS
#====================================================================================

@bp.route("/despesas/exportar/excel")
@login_required
@permission_required("financeiro", "visualizar")
def despesas_exportar_excel():
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    busca = _nome_preenchido(request.args.get("busca")).upper()
    status_despesa = _nome_preenchido(request.args.get("status_despesa")).upper()
    status_nd = _nome_preenchido(request.args.get("status_nd")).upper()
    origem = _nome_preenchido(request.args.get("origem")).upper()
    data_inicial = _nome_preenchido(request.args.get("data_inicial"))
    data_final = _nome_preenchido(request.args.get("data_final"))
    venc_inicial = _nome_preenchido(request.args.get("venc_inicial"))
    venc_final = _nome_preenchido(request.args.get("venc_final"))
    nd_numero = _nome_preenchido(request.args.get("nd_numero")).upper()
    somente_vencidas = request.args.get("somente_vencidas") == "1"

    filtros = ["1=1"]
    params = {}

    tem_filtro = any([
        busca,
        status_despesa and status_despesa != "TODOS",
        status_nd and status_nd != "TODOS",
        origem and origem not in ("TODAS", "TODOS"),
        data_inicial,
        data_final,
        venc_inicial,
        venc_final,
        nd_numero,
        somente_vencidas,
    ])

    if not tem_filtro:
        flash("APLIQUE AO MENOS UM FILTRO PARA EXPORTAR.", "warning")
        return redirect(url_for("financeiro_dois.despesas"))

    if busca:
        filtros.append("""
            (
                UPPER(COALESCE(d.numero_despesa, '')) LIKE :busca
                OR UPPER(COALESCE(d.numero_documento, '')) LIKE :busca
                OR UPPER(COALESCE(d.fornecedor, '')) LIKE :busca
                OR UPPER(COALESCE(d.descricao, '')) LIKE :busca
            )
        """)
        params["busca"] = f"%{busca}%"

    if status_despesa and status_despesa != "TODOS":
        filtros.append("UPPER(COALESCE(d.status_despesa, '')) = :status_despesa")
        params["status_despesa"] = status_despesa

    if status_nd and status_nd != "TODOS":
        filtros.append("UPPER(COALESCE(d.status_nd, '')) = :status_nd")
        params["status_nd"] = status_nd

    if origem and origem not in ("TODAS", "TODOS"):
        filtros.append("UPPER(COALESCE(d.origem_tipo, '')) = :origem")
        params["origem"] = origem

    if data_inicial:
        filtros.append("d.data_documento >= :data_inicial")
        params["data_inicial"] = data_inicial

    if data_final:
        filtros.append("d.data_documento <= :data_final")
        params["data_final"] = data_final

    if venc_inicial:
        filtros.append("d.vencimento >= :venc_inicial")
        params["venc_inicial"] = venc_inicial

    if venc_final:
        filtros.append("d.vencimento <= :venc_final")
        params["venc_final"] = venc_final

    if nd_numero:
        filtros.append("UPPER(COALESCE(d.nd_numero, '')) LIKE :nd_numero")
        params["nd_numero"] = f"%{nd_numero}%"

    if somente_vencidas:
        filtros.append("""
            UPPER(COALESCE(d.status_despesa, 'PENDENTE')) <> 'PAGA'
            AND d.vencimento IS NOT NULL
            AND d.vencimento < CURRENT_DATE
        """)

    engine = get_engine()
    with engine.connect() as conn:
        registros = conn.execute(text(f"""
            SELECT
                TO_CHAR(d.data_documento, 'DD/MM/YYYY') AS data,
                TO_CHAR(d.vencimento, 'DD/MM/YYYY') AS vencimento,
                UPPER(COALESCE(d.tipo_documento, '')) AS tipo_documento,
                UPPER(COALESCE(d.numero_despesa, '')) AS numero_despesa,
                UPPER(COALESCE(d.numero_documento, '')) AS numero_documento,
                UPPER(COALESCE(d.fornecedor, '')) AS fornecedor,
                UPPER(COALESCE(d.descricao, '')) AS descricao,
                UPPER(COALESCE(d.centro_custo, '')) AS centro_custo,
                UPPER(COALESCE(d.nd_numero, '')) AS nd_numero,
                UPPER(COALESCE(d.status_despesa, '')) AS status_despesa,
                UPPER(COALESCE(d.status_nd, '')) AS status_nd,
                UPPER(COALESCE(d.origem_tipo, '')) AS origem,
                CASE
                    WHEN UPPER(COALESCE(d.origem_tipo, '')) = 'OM' THEN COALESCE((
                        SELECT SUM(CASE WHEN COALESCE(l.valor_brl, 0) > 0 THEN COALESCE(l.valor_brl, 0) ELSE 0 END)
                        FROM financeiro2_om_linhas l
                        WHERE l.om_id = d.origem_id
                          AND COALESCE(l.status, 'Ativo') = 'Ativo'
                    ), 0)
                    WHEN UPPER(COALESCE(d.origem_tipo, '')) = 'RD' THEN COALESCE((
                        SELECT SUM(CASE WHEN COALESCE(l.valor, 0) > 0 THEN COALESCE(l.valor, 0) ELSE 0 END)
                        FROM financeiro2_rd_linhas l
                        WHERE l.rd_id = d.origem_id
                          AND COALESCE(l.status, 'Ativo') = 'Ativo'
                    ), 0)
                    ELSE COALESCE(d.valor, 0)
                END AS valor
            FROM financeiro2_despesas d
            WHERE {' AND '.join(filtros)}
            ORDER BY d.data_documento DESC, d.id DESC
        """), params).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "DESPESAS"

    headers = [
        "DATA", "VENCIMENTO", "TIPO", "Nº DESPESA", "Nº DOC.",
        "FORNECEDOR", "DESCRIÇÃO", "CC", "ND", "STATUS DESPESA",
        "STATUS ND", "ORIGEM", "VALOR"
    ]
    ws.append(headers)

    fill = PatternFill("solid", fgColor="16324F")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    total_valor = 0.0
    for r in registros:
        valor = float(r["valor"] or 0)
        total_valor += valor
        ws.append([
            r["data"], r["vencimento"], r["tipo_documento"], r["numero_despesa"],
            r["numero_documento"], r["fornecedor"], r["descricao"], r["centro_custo"],
            r["nd_numero"], r["status_despesa"], r["status_nd"], r["origem"], valor
        ])

    ws.append([])
    ws.append(["", "", "", "", "", "", "", "", "", "", "", "TOTAL", total_valor])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="despesas_filtradas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    
@bp.route("/despesas/exportar/pdf")
@login_required
@permission_required("financeiro", "visualizar")
def despesas_exportar_pdf():
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas

    busca = _nome_preenchido(request.args.get("busca")).upper()
    status_despesa = _nome_preenchido(request.args.get("status_despesa")).upper()
    status_nd = _nome_preenchido(request.args.get("status_nd")).upper()
    origem = _nome_preenchido(request.args.get("origem")).upper()
    data_inicial = _nome_preenchido(request.args.get("data_inicial"))
    data_final = _nome_preenchido(request.args.get("data_final"))
    venc_inicial = _nome_preenchido(request.args.get("venc_inicial"))
    venc_final = _nome_preenchido(request.args.get("venc_final"))
    nd_numero = _nome_preenchido(request.args.get("nd_numero")).upper()
    somente_vencidas = request.args.get("somente_vencidas") == "1"

    filtros = ["1=1"]
    params = {}

    tem_filtro = any([
        busca,
        status_despesa and status_despesa != "TODOS",
        status_nd and status_nd != "TODOS",
        origem and origem not in ("TODAS", "TODOS"),
        data_inicial,
        data_final,
        venc_inicial,
        venc_final,
        nd_numero,
        somente_vencidas,
    ])

    if not tem_filtro:
        flash("APLIQUE AO MENOS UM FILTRO PARA EXPORTAR.", "warning")
        return redirect(url_for("financeiro_dois.despesas"))

    if busca:
        filtros.append("""
            (
                UPPER(COALESCE(d.numero_despesa, '')) LIKE :busca
                OR UPPER(COALESCE(d.numero_documento, '')) LIKE :busca
                OR UPPER(COALESCE(d.fornecedor, '')) LIKE :busca
                OR UPPER(COALESCE(d.descricao, '')) LIKE :busca
            )
        """)
        params["busca"] = f"%{busca}%"

    if status_despesa and status_despesa != "TODOS":
        filtros.append("UPPER(COALESCE(d.status_despesa, '')) = :status_despesa")
        params["status_despesa"] = status_despesa

    if status_nd and status_nd != "TODOS":
        filtros.append("UPPER(COALESCE(d.status_nd, '')) = :status_nd")
        params["status_nd"] = status_nd

    if origem and origem not in ("TODAS", "TODOS"):
        filtros.append("UPPER(COALESCE(d.origem_tipo, '')) = :origem")
        params["origem"] = origem

    if data_inicial:
        filtros.append("d.data_documento >= :data_inicial")
        params["data_inicial"] = data_inicial

    if data_final:
        filtros.append("d.data_documento <= :data_final")
        params["data_final"] = data_final

    if venc_inicial:
        filtros.append("d.vencimento >= :venc_inicial")
        params["venc_inicial"] = venc_inicial

    if venc_final:
        filtros.append("d.vencimento <= :venc_final")
        params["venc_final"] = venc_final

    if nd_numero:
        filtros.append("UPPER(COALESCE(d.nd_numero, '')) LIKE :nd_numero")
        params["nd_numero"] = f"%{nd_numero}%"

    if somente_vencidas:
        filtros.append("""
            UPPER(COALESCE(d.status_despesa, 'PENDENTE')) <> 'PAGA'
            AND d.vencimento IS NOT NULL
            AND d.vencimento < CURRENT_DATE
        """)

    engine = get_engine()
    with engine.connect() as conn:
        registros = conn.execute(text(f"""
            SELECT
                TO_CHAR(d.data_documento, 'DD/MM/YYYY') AS data,
                TO_CHAR(d.vencimento, 'DD/MM/YYYY') AS vencimento,
                UPPER(COALESCE(d.numero_despesa, '')) AS numero_despesa,
                UPPER(COALESCE(d.fornecedor, '')) AS fornecedor,
                UPPER(COALESCE(d.descricao, '')) AS descricao,
                UPPER(COALESCE(d.status_despesa, '')) AS status_despesa,
                UPPER(COALESCE(d.status_nd, '')) AS status_nd,
                UPPER(COALESCE(d.origem_tipo, '')) AS origem,
                CASE
                    WHEN UPPER(COALESCE(d.origem_tipo, '')) = 'OM' THEN COALESCE((
                        SELECT SUM(CASE WHEN COALESCE(l.valor_brl, 0) > 0 THEN COALESCE(l.valor_brl, 0) ELSE 0 END)
                        FROM financeiro2_om_linhas l
                        WHERE l.om_id = d.origem_id
                          AND COALESCE(l.status, 'Ativo') = 'Ativo'
                    ), 0)
                    WHEN UPPER(COALESCE(d.origem_tipo, '')) = 'RD' THEN COALESCE((
                        SELECT SUM(CASE WHEN COALESCE(l.valor, 0) > 0 THEN COALESCE(l.valor, 0) ELSE 0 END)
                        FROM financeiro2_rd_linhas l
                        WHERE l.rd_id = d.origem_id
                          AND COALESCE(l.status, 'Ativo') = 'Ativo'
                    ), 0)
                    ELSE COALESCE(d.valor, 0)
                END AS valor
            FROM financeiro2_despesas d
            WHERE {' AND '.join(filtros)}
            ORDER BY d.data_documento DESC, d.id DESC
        """), params).mappings().all()

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=landscape(A4))
    largura, altura = landscape(A4)

    y = altura - 30
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(30, y, "DESPESAS FILTRADAS")
    y -= 25

    pdf.setFont("Helvetica-Bold", 8)
    pdf.drawString(30, y, "DATA")
    pdf.drawString(80, y, "VENC.")
    pdf.drawString(130, y, "Nº DESPESA")
    pdf.drawString(230, y, "FORNECEDOR")
    pdf.drawString(410, y, "DESCRIÇÃO")
    pdf.drawString(610, y, "STATUS")
    pdf.drawString(680, y, "ND")
    pdf.drawString(750, y, "ORIGEM")
    pdf.drawRightString(820, y, "VALOR")
    y -= 15

    pdf.setFont("Helvetica", 8)
    total_valor = 0.0

    for r in registros:
        if y < 30:
            pdf.showPage()
            y = altura - 30
            pdf.setFont("Helvetica-Bold", 8)
            pdf.drawString(30, y, "DATA")
            pdf.drawString(80, y, "VENC.")
            pdf.drawString(130, y, "Nº DESPESA")
            pdf.drawString(230, y, "FORNECEDOR")
            pdf.drawString(410, y, "DESCRIÇÃO")
            pdf.drawString(610, y, "STATUS")
            pdf.drawString(680, y, "ND")
            pdf.drawString(750, y, "ORIGEM")
            pdf.drawRightString(820, y, "VALOR")
            y -= 15
            pdf.setFont("Helvetica", 8)

        valor = float(r["valor"] or 0)
        total_valor += valor

        pdf.drawString(30, y, str(r["data"]))
        pdf.drawString(80, y, str(r["vencimento"] or "—"))
        pdf.drawString(130, y, str(r["numero_despesa"])[:16])
        pdf.drawString(230, y, str(r["fornecedor"])[:30])
        pdf.drawString(410, y, str(r["descricao"])[:34])
        pdf.drawString(610, y, str(r["status_despesa"])[:10])
        pdf.drawString(680, y, str(r["status_nd"])[:10])
        pdf.drawString(750, y, str(r["origem"])[:8])
        pdf.drawRightString(820, y, f"{valor:.2f}")
        y -= 13

    y -= 10
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawRightString(820, y, f"TOTAL: {total_valor:.2f}")
    pdf.save()

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="despesas_filtradas.pdf",
        mimetype="application/pdf",
    )
    
@bp.route("/notas-debito")
@login_required
@permission_required("financeiro", "visualizar")
def notas_debito():
    numero_nd = _nome_preenchido(request.args.get("numero_nd")).upper()
    empresa_nd = _nome_preenchido(request.args.get("empresa_nd")).upper()
    status = _nome_preenchido(request.args.get("status")).upper()
    data_inicial = _nome_preenchido(request.args.get("data_inicial"))
    data_final = _nome_preenchido(request.args.get("data_final"))

    filtros = ["1=1"]
    params = {}

    if numero_nd:
        filtros.append("UPPER(COALESCE(nd.numero_nd, '')) LIKE :numero_nd")
        params["numero_nd"] = f"%{numero_nd}%"

    if empresa_nd:
        filtros.append("UPPER(COALESCE(nd.empresa_nd, '')) LIKE :empresa_nd")
        params["empresa_nd"] = f"%{empresa_nd}%"

    if status and status != "TODOS":
        filtros.append("UPPER(COALESCE(nd.status, '')) = :status")
        params["status"] = status

    if data_inicial:
        filtros.append("nd.data_nd >= :data_inicial")
        params["data_inicial"] = data_inicial

    if data_final:
        filtros.append("nd.data_nd <= :data_final")
        params["data_final"] = data_final

    engine = get_engine()
    with engine.connect() as conn:
        registros = conn.execute(text(f"""
            SELECT
                nd.id,
                UPPER(COALESCE(nd.numero_nd, '')) AS numero_nd,
                TO_CHAR(nd.data_nd, 'DD/MM/YYYY') AS data_nd,
                UPPER(COALESCE(nd.empresa_nd, '')) AS empresa_nd,
                UPPER(COALESCE(nd.status, '')) AS status,
                UPPER(COALESCE(nd.observacao, '')) AS observacao,
                COALESCE((
                    SELECT COUNT(*)
                    FROM financeiro2_notas_debito_despesas rel
                    WHERE rel.nd_id = nd.id
                ), 0) AS qtd_despesas
            FROM financeiro2_notas_debito nd
            WHERE {' AND '.join(filtros)}
            ORDER BY nd.data_nd DESC, nd.id DESC
        """), params).mappings().all()

    return render_template(
        "financeiro_dois/notas_debito.html",
        subnav_links=build_financeiro_dois_subnav("notas_debito"),
        registros=registros,
        filtros={
            "numero_nd": request.args.get("numero_nd", ""),
            "empresa_nd": request.args.get("empresa_nd", ""),
            "status": request.args.get("status", "TODOS"),
            "data_inicial": request.args.get("data_inicial", ""),
            "data_final": request.args.get("data_final", ""),
        }
    )


@bp.route("/notas-debito/nova")
@login_required
@permission_required("financeiro", "visualizar")
def nota_debito_nova():
    hoje = date.today().strftime("%Y-%m-%d")
    engine = get_engine()

    with engine.connect() as conn:
        empresas_nd = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_empresas_nd
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

    nd = {
        "id": 0,
        "numero_nd": "",
        "data_form": hoje,
        "data_nd": "",
        "empresa_nd": "",
        "status": "ABERTA",
        "observacao": "",
        "eh_nova": True,
        "qtd_despesas": 0,
        "total_linhas_vinculadas": 0,
        "total_linhas_desconsideradas": 0,
        "total_valor_vinculado": 0.0,
        "total_valor_desconsiderado": 0.0,
        "total_valor_pendente": 0.0,
    }

    return render_template(
        "financeiro_dois/nota_debito_editar.html",
        subnav_links=build_financeiro_dois_subnav("notas_debito"),
        nd=nd,
        empresas_nd=empresas_nd,
    )


@bp.route("/notas-debito/criar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def nota_debito_criar():
    numero_nd = _nome_preenchido(request.form.get("numero_nd")).upper()
    data_nd = _nome_preenchido(request.form.get("data_nd"))
    empresa_nd = _nome_preenchido(request.form.get("empresa_nd")).upper()
    status = _nome_preenchido(request.form.get("status")).upper() or "ABERTA"
    observacao = _nome_preenchido(request.form.get("observacao")).upper()

    engine = get_engine()
    with engine.connect() as conn:
        empresas_nd = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_empresas_nd
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

    nd_form = {
        "id": 0,
        "numero_nd": numero_nd,
        "data_form": data_nd,
        "data_nd": "",
        "empresa_nd": empresa_nd,
        "status": status,
        "observacao": observacao,
        "eh_nova": True,
        "qtd_despesas": 0,
        "total_linhas_vinculadas": 0,
        "total_linhas_desconsideradas": 0,
        "total_valor_vinculado": 0.0,
        "total_valor_desconsiderado": 0.0,
        "total_valor_pendente": 0.0,
    }

    if not numero_nd or not data_nd or not empresa_nd:
        flash("PREENCHA NÚMERO ND, DATA E EMPRESA ND.", "warning")
        return render_template(
            "financeiro_dois/nota_debito_editar.html",
            subnav_links=build_financeiro_dois_subnav("notas_debito"),
            nd=nd_form,
            empresas_nd=empresas_nd,
        )

    with engine.begin() as conn:
        existe = conn.execute(text("""
            SELECT id
            FROM financeiro2_notas_debito
            WHERE UPPER(numero_nd) = :numero_nd
            LIMIT 1
        """), {"numero_nd": numero_nd}).mappings().first()

        if existe:
            flash("JÁ EXISTE UMA ND COM ESSE NÚMERO.", "warning")
            return render_template(
                "financeiro_dois/nota_debito_editar.html",
                subnav_links=build_financeiro_dois_subnav("notas_debito"),
                nd=nd_form,
                empresas_nd=empresas_nd,
            )

        novo_id = conn.execute(text("""
            INSERT INTO financeiro2_notas_debito (
                numero_nd,
                data_nd,
                empresa_nd,
                status,
                observacao,
                criado_em,
                atualizado_em
            ) VALUES (
                :numero_nd,
                :data_nd,
                :empresa_nd,
                :status,
                :observacao,
                CURRENT_TIMESTAMP,
                CURRENT_TIMESTAMP
            )
            RETURNING id
        """), {
            "numero_nd": numero_nd,
            "data_nd": data_nd,
            "empresa_nd": empresa_nd,
            "status": status,
            "observacao": observacao,
        }).scalar()

    flash("ND CRIADA COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.nota_debito_editar", nd_id=novo_id))

@bp.route("/notas-debito/<int:nd_id>")
@login_required
@permission_required("financeiro", "visualizar")
def nota_debito_editar(nd_id: int):
    engine = get_engine()

    with engine.connect() as conn:
        nd = conn.execute(text("""
            SELECT
                id,
                UPPER(COALESCE(numero_nd, '')) AS numero_nd,
                TO_CHAR(data_nd, 'YYYY-MM-DD') AS data_form,
                TO_CHAR(data_nd, 'DD/MM/YYYY') AS data_nd,
                UPPER(COALESCE(empresa_nd, '')) AS empresa_nd,
                UPPER(COALESCE(status, '')) AS status,
                UPPER(COALESCE(observacao, '')) AS observacao
            FROM financeiro2_notas_debito
            WHERE id = :id
        """), {"id": nd_id}).mappings().first()

        if not nd:
            abort(404)

        empresas_nd = conn.execute(text("""
            SELECT UPPER(nome) AS nome
            FROM financeiro2_cad_empresas_nd
            WHERE status = 'Ativo'
            ORDER BY nome
        """)).mappings().all()

        despesas_rel = conn.execute(text("""
            SELECT
                d.id,
                UPPER(COALESCE(d.numero_despesa, '')) AS numero_despesa,
                UPPER(COALESCE(d.origem_tipo, '')) AS origem,
                d.origem_id,
                UPPER(COALESCE(d.fornecedor, '')) AS fornecedor,
                UPPER(COALESCE(d.descricao, '')) AS descricao,
                UPPER(COALESCE(d.status_nd, '')) AS status_nd
            FROM financeiro2_notas_debito_despesas rel
            JOIN financeiro2_despesas d ON d.id = rel.despesa_id
            WHERE rel.nd_id = :nd_id
            ORDER BY d.id DESC
        """), {"nd_id": nd_id}).mappings().all()

        despesas_rel = [dict(x) for x in despesas_rel]

        total_linhas_vinculadas = 0
        total_linhas_desconsideradas = 0
        total_valor_vinculado = 0.0
        total_valor_desconsiderado = 0.0
        total_valor_pendente = 0.0

        for item in despesas_rel:
            if item["origem"] == "OM":
                totais = conn.execute(text("""
                    SELECT
                        COUNT(*) FILTER (WHERE COALESCE(numero_nd, '') = :numero_nd) AS qtd_vinculadas,
                        COUNT(*) FILTER (WHERE COALESCE(desconsiderada_nd, FALSE) = TRUE) AS qtd_desconsideradas,
                        COALESCE(SUM(CASE WHEN COALESCE(numero_nd, '') = :numero_nd THEN COALESCE(valor_brl, 0) ELSE 0 END), 0) AS valor_vinculado,
                        COALESCE(SUM(CASE WHEN COALESCE(desconsiderada_nd, FALSE) = TRUE THEN COALESCE(valor_brl, 0) ELSE 0 END), 0) AS valor_desconsiderado,
                        COALESCE(SUM(CASE WHEN COALESCE(numero_nd, '') = '' AND COALESCE(desconsiderada_nd, FALSE) = FALSE THEN COALESCE(valor_brl, 0) ELSE 0 END), 0) AS valor_pendente
                    FROM financeiro2_om_linhas
                    WHERE om_id = :origem_id
                      AND COALESCE(status, 'Ativo') = 'Ativo'
                      AND COALESCE(valor_brl, 0) > 0
                """), {
                    "origem_id": item["origem_id"],
                    "numero_nd": nd["numero_nd"],
                }).mappings().first()
            else:
                totais = conn.execute(text("""
                    SELECT
                        COUNT(*) FILTER (WHERE COALESCE(numero_nd, '') = :numero_nd) AS qtd_vinculadas,
                        COUNT(*) FILTER (WHERE COALESCE(desconsiderada_nd, FALSE) = TRUE) AS qtd_desconsideradas,
                        COALESCE(SUM(CASE WHEN COALESCE(numero_nd, '') = :numero_nd THEN COALESCE(valor, 0) ELSE 0 END), 0) AS valor_vinculado,
                        COALESCE(SUM(CASE WHEN COALESCE(desconsiderada_nd, FALSE) = TRUE THEN COALESCE(valor, 0) ELSE 0 END), 0) AS valor_desconsiderado,
                        COALESCE(SUM(CASE WHEN COALESCE(numero_nd, '') = '' AND COALESCE(desconsiderada_nd, FALSE) = FALSE THEN COALESCE(valor, 0) ELSE 0 END), 0) AS valor_pendente
                    FROM financeiro2_rd_linhas
                    WHERE rd_id = :origem_id
                      AND COALESCE(status, 'Ativo') = 'Ativo'
                      AND COALESCE(valor, 0) > 0
                """), {
                    "origem_id": item["origem_id"],
                    "numero_nd": nd["numero_nd"],
                }).mappings().first()

            item["qtd_vinculadas"] = int(totais["qtd_vinculadas"] or 0)
            item["qtd_desconsideradas"] = int(totais["qtd_desconsideradas"] or 0)
            item["valor_vinculado"] = float(totais["valor_vinculado"] or 0)
            item["valor_desconsiderado"] = float(totais["valor_desconsiderado"] or 0)
            item["valor_pendente"] = float(totais["valor_pendente"] or 0)

            total_linhas_vinculadas += item["qtd_vinculadas"]
            total_linhas_desconsideradas += item["qtd_desconsideradas"]
            total_valor_vinculado += item["valor_vinculado"]
            total_valor_desconsiderado += item["valor_desconsiderado"]
            total_valor_pendente += item["valor_pendente"]

    nd = dict(nd)
    nd["eh_nova"] = False
    nd["qtd_despesas"] = len(despesas_rel)
    nd["total_linhas_vinculadas"] = total_linhas_vinculadas
    nd["total_linhas_desconsideradas"] = total_linhas_desconsideradas
    nd["total_valor_vinculado"] = total_valor_vinculado
    nd["total_valor_desconsiderado"] = total_valor_desconsiderado
    nd["total_valor_pendente"] = total_valor_pendente

    return render_template(
        "financeiro_dois/nota_debito_editar.html",
        subnav_links=build_financeiro_dois_subnav("notas_debito"),
        nd=nd,
        empresas_nd=empresas_nd,
        despesas_rel=despesas_rel,
    )

@bp.route("/notas-debito/<int:nd_id>/salvar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def nota_debito_salvar(nd_id: int):
    numero_nd = _nome_preenchido(request.form.get("numero_nd")).upper()
    data_nd = _nome_preenchido(request.form.get("data_nd"))
    empresa_nd = _nome_preenchido(request.form.get("empresa_nd")).upper()
    status = _nome_preenchido(request.form.get("status")).upper() or "ABERTA"
    observacao = _nome_preenchido(request.form.get("observacao")).upper()

    if not numero_nd or not data_nd or not empresa_nd:
        flash("PREENCHA NÚMERO ND, DATA E EMPRESA ND.", "warning")
        return redirect(url_for("financeiro_dois.nota_debito_editar", nd_id=nd_id))

    engine = get_engine()
    with engine.begin() as conn:
        existe = conn.execute(text("""
            SELECT id
            FROM financeiro2_notas_debito
            WHERE UPPER(numero_nd) = :numero_nd
              AND id <> :id
            LIMIT 1
        """), {
            "numero_nd": numero_nd,
            "id": nd_id
        }).mappings().first()

        if existe:
            flash("JÁ EXISTE OUTRA ND COM ESSE NÚMERO.", "warning")
            return redirect(url_for("financeiro_dois.nota_debito_editar", nd_id=nd_id))

        conn.execute(text("""
            UPDATE financeiro2_notas_debito
            SET
                numero_nd = :numero_nd,
                data_nd = :data_nd,
                empresa_nd = :empresa_nd,
                status = :status,
                observacao = :observacao,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :id
        """), {
            "id": nd_id,
            "numero_nd": numero_nd,
            "data_nd": data_nd,
            "empresa_nd": empresa_nd,
            "status": status,
            "observacao": observacao,
        })

    flash("ND SALVA COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.nota_debito_editar", nd_id=nd_id))
    
@bp.route("/notas-debito/<int:nd_id>/despesas")
@login_required
@permission_required("financeiro", "visualizar")
def nota_debito_despesas(nd_id: int):
    busca = _nome_preenchido(request.args.get("busca")).upper()
    origem = _nome_preenchido(request.args.get("origem")).upper() or "TODAS"
    status_nd = _nome_preenchido(request.args.get("status_nd")).upper() or "TODOS"

    engine = get_engine()
    with engine.connect() as conn:
        nd = conn.execute(text("""
            SELECT
                id,
                UPPER(COALESCE(numero_nd, '')) AS numero_nd,
                TO_CHAR(data_nd, 'DD/MM/YYYY') AS data_nd,
                UPPER(COALESCE(empresa_nd, '')) AS empresa_nd,
                UPPER(COALESCE(status, '')) AS status
            FROM financeiro2_notas_debito
            WHERE id = :id
        """), {"id": nd_id}).mappings().first()

        if not nd:
            abort(404)

        filtros = [
            "UPPER(COALESCE(d.origem_tipo, '')) IN ('OM', 'RD', 'OPERACIONAL')",
            "UPPER(COALESCE(d.status_nd, '')) IN ('NÃO VINCULADA', 'PARCIAL')"
        ]
        params = {}

        if busca:
            filtros.append("""
                (
                    UPPER(COALESCE(d.numero_despesa, '')) LIKE :busca
                    OR UPPER(COALESCE(d.numero_documento, '')) LIKE :busca
                    OR UPPER(COALESCE(d.fornecedor, '')) LIKE :busca
                    OR UPPER(COALESCE(d.descricao, '')) LIKE :busca
                    OR UPPER(COALESCE(d.nd_numero, '')) LIKE :busca
                )
            """)
            params["busca"] = f"%{busca}%"

        if origem != "TODAS":
            filtros.append("UPPER(COALESCE(d.origem_tipo, '')) = :origem")
            params["origem"] = origem

        if status_nd != "TODOS":
            filtros.append("UPPER(COALESCE(d.status_nd, '')) = :status_nd")
            params["status_nd"] = status_nd

        despesas = conn.execute(text(f"""
            SELECT
                d.id,
                UPPER(COALESCE(d.numero_despesa, '')) AS numero_despesa,
                UPPER(COALESCE(d.origem_tipo, '')) AS origem,
                d.origem_id,
                CASE
                    WHEN UPPER(COALESCE(d.origem_tipo, '')) = 'OM' THEN (
                        SELECT UPPER(COALESCE(om.numero_om, ''))
                        FROM financeiro2_om om
                        WHERE om.id = d.origem_id
                    )
                    WHEN UPPER(COALESCE(d.origem_tipo, '')) = 'RD' THEN (
                        SELECT UPPER(COALESCE(rd.numero_rd, ''))
                        FROM financeiro2_rd rd
                        WHERE rd.id = d.origem_id
                    )
                    ELSE UPPER(COALESCE(d.numero_documento, ''))
                END AS numero_origem,
                UPPER(COALESCE(d.fornecedor, '')) AS fornecedor,
                UPPER(COALESCE(d.descricao, '')) AS descricao,
                UPPER(COALESCE(d.status_nd, '')) AS status_nd,
                CASE
                    WHEN UPPER(COALESCE(d.origem_tipo, '')) = 'OM' THEN COALESCE((
                        SELECT SUM(
                            CASE WHEN COALESCE(l.valor_brl, 0) > 0 THEN COALESCE(l.valor_brl, 0) ELSE 0 END
                        )
                        FROM financeiro2_om_linhas l
                        WHERE l.om_id = d.origem_id
                          AND COALESCE(l.status, 'Ativo') = 'Ativo'
                    ), 0)
                    WHEN UPPER(COALESCE(d.origem_tipo, '')) = 'RD' THEN COALESCE((
                        SELECT SUM(
                            CASE WHEN COALESCE(l.valor, 0) > 0 THEN COALESCE(l.valor, 0) ELSE 0 END
                        )
                        FROM financeiro2_rd_linhas l
                        WHERE l.rd_id = d.origem_id
                          AND COALESCE(l.status, 'Ativo') = 'Ativo'
                    ), 0)
                    ELSE COALESCE(d.valor, 0)
                END AS valor
            FROM financeiro2_despesas d
            WHERE {' AND '.join(filtros)}
            ORDER BY d.id DESC
        """), params).mappings().all()

    return render_template(
        "financeiro_dois/nota_debito_despesas.html",
        subnav_links=build_financeiro_dois_subnav("notas_debito"),
        nd=nd,
        despesas=despesas,
        filtros={
            "busca": request.args.get("busca", ""),
            "origem": origem,
            "status_nd": status_nd,
        }
    )
    
@bp.route("/notas-debito/<int:nd_id>/despesas/<int:despesa_id>/linhas")
@login_required
@permission_required("financeiro", "visualizar")
def nota_debito_origem_linhas(nd_id: int, despesa_id: int):
    engine = get_engine()

    with engine.connect() as conn:
        nd = conn.execute(text("""
            SELECT
                id,
                UPPER(COALESCE(numero_nd, '')) AS numero_nd,
                TO_CHAR(data_nd, 'DD/MM/YYYY') AS data_nd,
                UPPER(COALESCE(empresa_nd, '')) AS empresa_nd,
                UPPER(COALESCE(status, '')) AS status
            FROM financeiro2_notas_debito
            WHERE id = :id
        """), {"id": nd_id}).mappings().first()

        if not nd:
            abort(404)

        despesa = conn.execute(text("""
            SELECT
                id,
                UPPER(COALESCE(numero_despesa, '')) AS numero_despesa,
                UPPER(COALESCE(origem_tipo, '')) AS origem,
                origem_id,
                UPPER(COALESCE(status_nd, '')) AS status_nd
            FROM financeiro2_despesas
            WHERE id = :id
        """), {"id": despesa_id}).mappings().first()

        if not despesa:
            abort(404)

        if despesa["origem"] == "OM":
            linhas = conn.execute(text("""
                SELECT
                    id,
                    TO_CHAR(data_lancamento, 'DD/MM/YYYY') AS data,
                    COALESCE(recibo, id) AS referencia,
                    UPPER(COALESCE(tipo_linha, '')) AS descricao,
                    UPPER(COALESCE(categoria, '')) AS categoria,
                    UPPER(COALESCE(aplicacao, '')) AS aplicacao,
                    COALESCE(valor_brl, 0) AS valor,
                    UPPER(COALESCE(numero_nd, '')) AS numero_nd,
                    COALESCE(desconsiderada_nd, FALSE) AS desconsiderada_nd,
                    UPPER(COALESCE(numero_nd_desconsiderada, '')) AS numero_nd_desconsiderada
                FROM financeiro2_om_linhas
                WHERE om_id = :origem_id
                  AND COALESCE(status, 'Ativo') = 'Ativo'
                  AND COALESCE(valor_brl, 0) > 0
                ORDER BY id
            """), {"origem_id": despesa["origem_id"]}).mappings().all()
        else:
            linhas = conn.execute(text("""
                SELECT
                    id,
                    TO_CHAR(data_lancamento, 'DD/MM/YYYY') AS data,
                    id AS referencia,
                    UPPER(COALESCE(descricao, '')) AS descricao,
                    UPPER(COALESCE(categoria, '')) AS categoria,
                    UPPER(COALESCE(aplicacao, '')) AS aplicacao,
                    COALESCE(valor, 0) AS valor,
                    UPPER(COALESCE(numero_nd, '')) AS numero_nd,
                    COALESCE(desconsiderada_nd, FALSE) AS desconsiderada_nd,
                    UPPER(COALESCE(numero_nd_desconsiderada, '')) AS numero_nd_desconsiderada
                FROM financeiro2_rd_linhas
                WHERE rd_id = :origem_id
                  AND COALESCE(status, 'Ativo') = 'Ativo'
                  AND COALESCE(valor, 0) > 0
                ORDER BY id
            """), {"origem_id": despesa["origem_id"]}).mappings().all()

    return render_template(
        "financeiro_dois/nota_debito_linhas.html",
        subnav_links=build_financeiro_dois_subnav("notas_debito"),
        nd=nd,
        despesa=despesa,
        linhas=linhas,
        pode_reverter=_usuario_eh_administrador(),
    )
    
@bp.route("/notas-debito/<int:nd_id>/despesas/<int:despesa_id>/linhas/salvar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def nota_debito_salvar_linhas(nd_id: int, despesa_id: int):
    vincular_ids = set(request.form.getlist("vincular_ids"))
    desconsiderar_ids = set(request.form.getlist("desconsiderar_ids"))

    conflito = vincular_ids.intersection(desconsiderar_ids)
    if conflito:
        flash("UMA MESMA LINHA NÃO PODE SER VINCULADA E DESCONSIDERADA AO MESMO TEMPO.", "warning")
        return redirect(url_for("financeiro_dois.nota_debito_origem_linhas", nd_id=nd_id, despesa_id=despesa_id))

    engine = get_engine()
    with engine.begin() as conn:
        nd = conn.execute(text("""
            SELECT
                id,
                UPPER(COALESCE(numero_nd, '')) AS numero_nd
            FROM financeiro2_notas_debito
            WHERE id = :id
        """), {"id": nd_id}).mappings().first()

        despesa = conn.execute(text("""
            SELECT
                id,
                UPPER(COALESCE(origem_tipo, '')) AS origem,
                origem_id,
                UPPER(COALESCE(nd_numero, '')) AS nd_numero_atual
            FROM financeiro2_despesas
            WHERE id = :id
        """), {"id": despesa_id}).mappings().first()

        if not nd or not despesa:
            abort(404)

        tabela = "financeiro2_om_linhas" if despesa["origem"] == "OM" else "financeiro2_rd_linhas"
        fk = "om_id" if despesa["origem"] == "OM" else "rd_id"
        campo_valor = "valor_brl" if despesa["origem"] == "OM" else "valor"

        # Vincular linhas selecionadas
        for linha_id in vincular_ids:
            conn.execute(text(f"""
                UPDATE {tabela}
                SET
                    numero_nd = :numero_nd,
                    desconsiderada_nd = FALSE
                WHERE id = :linha_id
                  AND {fk} = :origem_id
                  AND COALESCE(status, 'Ativo') = 'Ativo'
                  AND COALESCE(numero_nd, '') = ''
            """), {
                "numero_nd": nd["numero_nd"],
                "linha_id": int(linha_id),
                "origem_id": despesa["origem_id"],
            })

        # Desconsiderar linhas selecionadas
        for linha_id in desconsiderar_ids:
            conn.execute(text(f"""
                UPDATE {tabela}
                SET
                    desconsiderada_nd = TRUE,
                    numero_nd_desconsiderada = :numero_nd
                WHERE id = :linha_id
                  AND {fk} = :origem_id
                  AND COALESCE(status, 'Ativo') = 'Ativo'
                  AND COALESCE(numero_nd, '') = ''
            """), {
                "numero_nd": nd["numero_nd"],
                "linha_id": int(linha_id),
                "origem_id": despesa["origem_id"],
            })

        # Registrar relação ND x despesa
        rel = conn.execute(text("""
            SELECT id
            FROM financeiro2_notas_debito_despesas
            WHERE nd_id = :nd_id
              AND despesa_id = :despesa_id
            LIMIT 1
        """), {"nd_id": nd_id, "despesa_id": despesa_id}).mappings().first()

        if not rel:
            conn.execute(text("""
                INSERT INTO financeiro2_notas_debito_despesas (
                    nd_id, despesa_id, criado_em
                ) VALUES (
                    :nd_id, :despesa_id, CURRENT_TIMESTAMP
                )
            """), {
                "nd_id": nd_id,
                "despesa_id": despesa_id,
            })

        # Recalcular status da despesa com base nas linhas da origem
        totais = conn.execute(text(f"""
            SELECT
                COUNT(*) AS total_linhas,
                SUM(
                    CASE
                        WHEN COALESCE(numero_nd, '') <> ''
                             OR COALESCE(desconsiderada_nd, FALSE) = TRUE
                        THEN 1 ELSE 0
                    END
                ) AS total_resolvidas
            FROM {tabela}
            WHERE {fk} = :origem_id
              AND COALESCE(status, 'Ativo') = 'Ativo'
              AND COALESCE({campo_valor}, 0) > 0
        """), {"origem_id": despesa["origem_id"]}).mappings().first()

        total_linhas = int(totais["total_linhas"] or 0)
        total_resolvidas = int(totais["total_resolvidas"] or 0)

        if total_resolvidas <= 0:
            novo_status = "NÃO VINCULADA"
        elif total_resolvidas < total_linhas:
            novo_status = "PARCIAL"
        else:
            novo_status = "VINCULADA"

        nd_numero_principal = despesa["nd_numero_atual"] or nd["numero_nd"]

        conn.execute(text("""
            UPDATE financeiro2_despesas
            SET
                status_nd = :status_nd,
                nd_numero = :nd_numero,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :id
        """), {
            "id": despesa_id,
            "status_nd": novo_status,
            "nd_numero": nd_numero_principal,
        })

        # Recalcular status da própria ND
        _recalcular_status_nd(conn, nd_id)

    flash("LINHAS DA ORIGEM PROCESSADAS COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.nota_debito_editar", nd_id=nd_id))
    
@bp.route("/notas-debito/<int:nd_id>/despesas/<int:despesa_id>/linhas/<int:linha_id>/reverter", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def nota_debito_reverter_linha(nd_id: int, despesa_id: int, linha_id: int):
    if not _usuario_eh_administrador():
        flash("APENAS O PERFIL ADMINISTRADOR PODE REVERTER LINHAS.", "danger")
        return redirect(url_for("financeiro_dois.nota_debito_origem_linhas", nd_id=nd_id, despesa_id=despesa_id))

    engine = get_engine()
    with engine.begin() as conn:
        nd = conn.execute(text("""
            SELECT
                id,
                UPPER(COALESCE(numero_nd, '')) AS numero_nd
            FROM financeiro2_notas_debito
            WHERE id = :id
        """), {"id": nd_id}).mappings().first()

        despesa = conn.execute(text("""
            SELECT
                id,
                UPPER(COALESCE(origem_tipo, '')) AS origem,
                origem_id,
                UPPER(COALESCE(nd_numero, '')) AS nd_numero_atual
            FROM financeiro2_despesas
            WHERE id = :id
        """), {"id": despesa_id}).mappings().first()

        if not nd or not despesa:
            abort(404)

        tabela = "financeiro2_om_linhas" if despesa["origem"] == "OM" else "financeiro2_rd_linhas"
        fk = "om_id" if despesa["origem"] == "OM" else "rd_id"
        campo_valor = "valor_brl" if despesa["origem"] == "OM" else "valor"

        linha = conn.execute(text(f"""
            SELECT
                id,
                UPPER(COALESCE(numero_nd, '')) AS numero_nd,
                COALESCE(desconsiderada_nd, FALSE) AS desconsiderada_nd,
                UPPER(COALESCE(numero_nd_desconsiderada, '')) AS numero_nd_desconsiderada
            FROM {tabela}
            WHERE id = :linha_id
              AND {fk} = :origem_id
              AND COALESCE(status, 'Ativo') = 'Ativo'
            LIMIT 1
        """), {
            "linha_id": linha_id,
            "origem_id": despesa["origem_id"],
        }).mappings().first()

        if not linha:
            flash("LINHA NÃO ENCONTRADA.", "warning")
            return redirect(url_for("financeiro_dois.nota_debito_origem_linhas", nd_id=nd_id, despesa_id=despesa_id))

        # Reverte vínculo desta ND
        if linha["numero_nd"] == nd["numero_nd"]:
            conn.execute(text(f"""
                UPDATE {tabela}
                SET numero_nd = NULL
                WHERE id = :linha_id
            """), {"linha_id": linha_id})

        # Reverte desconsideração desta ND
        elif linha["desconsiderada_nd"] and linha["numero_nd_desconsiderada"] == nd["numero_nd"]:
            conn.execute(text(f"""
                UPDATE {tabela}
                SET
                    desconsiderada_nd = FALSE,
                    numero_nd_desconsiderada = NULL
                WHERE id = :linha_id
            """), {"linha_id": linha_id})

        else:
            flash("ESSA LINHA NÃO ESTÁ VINCULADA OU DESCONSIDERADA POR ESTA ND.", "warning")
            return redirect(url_for("financeiro_dois.nota_debito_origem_linhas", nd_id=nd_id, despesa_id=despesa_id))

        totais = conn.execute(text(f"""
            SELECT
                COUNT(*) AS total_linhas,
                SUM(
                    CASE
                        WHEN COALESCE(numero_nd, '') <> ''
                          OR COALESCE(desconsiderada_nd, FALSE) = TRUE
                        THEN 1 ELSE 0
                    END
                ) AS total_resolvidas
            FROM {tabela}
            WHERE {fk} = :origem_id
              AND COALESCE(status, 'Ativo') = 'Ativo'
              AND COALESCE({campo_valor}, 0) > 0
        """), {"origem_id": despesa["origem_id"]}).mappings().first()

        total_linhas = int(totais["total_linhas"] or 0)
        total_resolvidas = int(totais["total_resolvidas"] or 0)

        if total_resolvidas <= 0:
            novo_status = "NÃO VINCULADA"
        elif total_resolvidas < total_linhas:
            novo_status = "PARCIAL"
        else:
            novo_status = "VINCULADA"

        conn.execute(text("""
            UPDATE financeiro2_despesas
            SET
                status_nd = :status_nd,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :id
        """), {
            "id": despesa_id,
            "status_nd": novo_status,
        })

        uso_nd = conn.execute(text(f"""
            SELECT COUNT(*) AS total
            FROM {tabela}
            WHERE {fk} = :origem_id
              AND (
                    COALESCE(numero_nd, '') = :numero_nd
                 OR (
                        COALESCE(desconsiderada_nd, FALSE) = TRUE
                    AND COALESCE(numero_nd_desconsiderada, '') = :numero_nd
                 )
              )
        """), {
            "origem_id": despesa["origem_id"],
            "numero_nd": nd["numero_nd"],
        }).mappings().first()

        if int(uso_nd["total"] or 0) <= 0:
            conn.execute(text("""
                DELETE FROM financeiro2_notas_debito_despesas
                WHERE nd_id = :nd_id
                  AND despesa_id = :despesa_id
            """), {
                "nd_id": nd_id,
                "despesa_id": despesa_id,
            })

        _recalcular_status_nd(conn, nd_id)

    flash("LINHA REVERTIDA COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.nota_debito_origem_linhas", nd_id=nd_id, despesa_id=despesa_id))
    
@bp.route("/notas-debito/<int:nd_id>/despesas/<int:despesa_id>/operacional", methods=["GET", "POST"])
@login_required
@permission_required("financeiro", "visualizar")
def nota_debito_operacional(nd_id: int, despesa_id: int):
    engine = get_engine()

    if request.method == "POST":
        if not _usuario_eh_administrador() and request.form.get("acao") == "desconsiderar":
            flash("APENAS O PERFIL ADMINISTRADOR PODE DESCONSIDERAR DESPESA OPERACIONAL.", "danger")
            return redirect(url_for("financeiro_dois.nota_debito_operacional", nd_id=nd_id, despesa_id=despesa_id))

        acao = _nome_preenchido(request.form.get("acao")).lower()

        with engine.begin() as conn:
            nd = conn.execute(text("""
                SELECT id, UPPER(COALESCE(numero_nd, '')) AS numero_nd
                FROM financeiro2_notas_debito
                WHERE id = :id
            """), {"id": nd_id}).mappings().first()

            despesa = conn.execute(text("""
                SELECT
                    id,
                    UPPER(COALESCE(origem_tipo, '')) AS origem,
                    UPPER(COALESCE(nd_numero, '')) AS nd_numero,
                    UPPER(COALESCE(numero_nd_desconsiderada, '')) AS numero_nd_desconsiderada
                FROM financeiro2_despesas
                WHERE id = :id
            """), {"id": despesa_id}).mappings().first()

            if not nd or not despesa:
                abort(404)

            if despesa["origem"] != "OPERACIONAL":
                flash("ESSA TELA É APENAS PARA DESPESA OPERACIONAL.", "warning")
                return redirect(url_for("financeiro_dois.nota_debito_despesas", nd_id=nd_id))

            relacao_outra_nd = conn.execute(text("""
                SELECT
                    r.nd_id,
                    UPPER(COALESCE(n.numero_nd, '')) AS numero_nd
                FROM financeiro2_notas_debito_despesas r
                JOIN financeiro2_notas_debito n
                  ON n.id = r.nd_id
                WHERE r.despesa_id = :despesa_id
                  AND r.nd_id <> :nd_id
                LIMIT 1
            """), {
                "despesa_id": despesa_id,
                "nd_id": nd_id,
            }).mappings().first()

            if relacao_outra_nd:
                flash(
                    f"DESPESA OPERACIONAL JÁ ESTÁ RELACIONADA À ND {relacao_outra_nd['numero_nd']} E SÓ PODE PARTICIPAR DE UMA ÚNICA ND.",
                    "warning"
                )
                return redirect(url_for("financeiro_dois.nota_debito_despesas", nd_id=nd_id))

            if despesa["nd_numero"] and despesa["nd_numero"] != nd["numero_nd"]:
                flash("DESPESA OPERACIONAL JÁ ESTÁ VINCULADA A OUTRA ND E SÓ PODE PARTICIPAR DE UMA ÚNICA ND.", "warning")
                return redirect(url_for("financeiro_dois.nota_debito_despesas", nd_id=nd_id))

            if despesa["numero_nd_desconsiderada"] and despesa["numero_nd_desconsiderada"] != nd["numero_nd"]:
                flash("DESPESA OPERACIONAL JÁ FOI DESCONSIDERADA EM OUTRA ND E SÓ PODE PARTICIPAR DE UMA ÚNICA ND.", "warning")
                return redirect(url_for("financeiro_dois.nota_debito_despesas", nd_id=nd_id))

            if acao == "vincular":
                rel = conn.execute(text("""
                    SELECT
                        r.id,
                        r.nd_id,
                        UPPER(COALESCE(n.numero_nd, '')) AS numero_nd
                    FROM financeiro2_notas_debito_despesas r
                    JOIN financeiro2_notas_debito n
                      ON n.id = r.nd_id
                    WHERE r.despesa_id = :despesa_id
                    LIMIT 1
                """), {"despesa_id": despesa_id}).mappings().first()

                if rel and int(rel["nd_id"]) != int(nd_id):
                    flash(
                        f"DESPESA OPERACIONAL JÁ ESTÁ RELACIONADA À ND {rel['numero_nd']} E SÓ PODE PARTICIPAR DE UMA ÚNICA ND.",
                        "warning"
                    )
                    return redirect(url_for("financeiro_dois.nota_debito_despesas", nd_id=nd_id))

                if not rel:
                    conn.execute(text("""
                        INSERT INTO financeiro2_notas_debito_despesas (
                            nd_id, despesa_id, criado_em
                        ) VALUES (
                            :nd_id, :despesa_id, CURRENT_TIMESTAMP
                        )
                    """), {"nd_id": nd_id, "despesa_id": despesa_id})

                if not rel:
                    conn.execute(text("""
                        INSERT INTO financeiro2_notas_debito_despesas (
                            nd_id, despesa_id, criado_em
                        ) VALUES (
                            :nd_id, :despesa_id, CURRENT_TIMESTAMP
                        )
                    """), {"nd_id": nd_id, "despesa_id": despesa_id})

                conn.execute(text("""
                    UPDATE financeiro2_despesas
                    SET
                        status_nd = 'VINCULADA',
                        nd_numero = CASE
                            WHEN COALESCE(nd_numero, '') = '' THEN :numero_nd
                            ELSE nd_numero
                        END,
                        numero_nd_desconsiderada = NULL,
                        atualizado_em = CURRENT_TIMESTAMP
                    WHERE id = :id
                """), {
                    "id": despesa_id,
                    "numero_nd": nd["numero_nd"],
                })

                _recalcular_status_nd(conn, nd_id)
                flash("DESPESA OPERACIONAL VINCULADA À ND.", "success")

            elif acao == "desconsiderar":
                rel = conn.execute(text("""
                    SELECT
                        r.id,
                        r.nd_id,
                        UPPER(COALESCE(n.numero_nd, '')) AS numero_nd
                    FROM financeiro2_notas_debito_despesas r
                    JOIN financeiro2_notas_debito n
                      ON n.id = r.nd_id
                    WHERE r.despesa_id = :despesa_id
                    LIMIT 1
                """), {"despesa_id": despesa_id}).mappings().first()

                if rel and int(rel["nd_id"]) != int(nd_id):
                    flash(
                        f"DESPESA OPERACIONAL JÁ ESTÁ RELACIONADA À ND {rel['numero_nd']} E SÓ PODE PARTICIPAR DE UMA ÚNICA ND.",
                        "warning"
                    )
                    return redirect(url_for("financeiro_dois.nota_debito_despesas", nd_id=nd_id))

                if not rel:
                    conn.execute(text("""
                        INSERT INTO financeiro2_notas_debito_despesas (
                            nd_id, despesa_id, criado_em
                        ) VALUES (
                            :nd_id, :despesa_id, CURRENT_TIMESTAMP
                        )
                    """), {"nd_id": nd_id, "despesa_id": despesa_id})

                if not rel:
                    conn.execute(text("""
                        INSERT INTO financeiro2_notas_debito_despesas (
                            nd_id, despesa_id, criado_em
                        ) VALUES (
                            :nd_id, :despesa_id, CURRENT_TIMESTAMP
                        )
                    """), {"nd_id": nd_id, "despesa_id": despesa_id})

                conn.execute(text("""
                    UPDATE financeiro2_despesas
                    SET
                        status_nd = 'VINCULADA',
                        nd_numero = CASE
                            WHEN COALESCE(nd_numero, '') = '' THEN :numero_nd
                            ELSE nd_numero
                        END,
                        numero_nd_desconsiderada = :numero_nd,
                        atualizado_em = CURRENT_TIMESTAMP
                    WHERE id = :id
                """), {
                    "id": despesa_id,
                    "numero_nd": nd["numero_nd"],
                })

                _recalcular_status_nd(conn, nd_id)
                flash("DESPESA OPERACIONAL DESCONSIDERADA PARA ESTA ND.", "success")

            else:
                flash("AÇÃO INVÁLIDA.", "warning")

        return redirect(url_for("financeiro_dois.nota_debito_editar", nd_id=nd_id))

    with engine.connect() as conn:
        nd = conn.execute(text("""
            SELECT
                id,
                UPPER(COALESCE(numero_nd, '')) AS numero_nd,
                TO_CHAR(data_nd, 'DD/MM/YYYY') AS data_nd,
                UPPER(COALESCE(empresa_nd, '')) AS empresa_nd,
                UPPER(COALESCE(status, '')) AS status
            FROM financeiro2_notas_debito
            WHERE id = :id
        """), {"id": nd_id}).mappings().first()

        despesa = conn.execute(text("""
            SELECT
                id,
                UPPER(COALESCE(numero_despesa, '')) AS numero_despesa,
                UPPER(COALESCE(origem_tipo, '')) AS origem,
                UPPER(COALESCE(numero_documento, '')) AS numero_documento,
                UPPER(COALESCE(fornecedor, '')) AS fornecedor,
                UPPER(COALESCE(descricao, '')) AS descricao,
                UPPER(COALESCE(status_nd, '')) AS status_nd,
                UPPER(COALESCE(nd_numero, '')) AS nd_numero,
                UPPER(COALESCE(numero_nd_desconsiderada, '')) AS numero_nd_desconsiderada,
                COALESCE(valor, 0) AS valor
            FROM financeiro2_despesas
            WHERE id = :id
        """), {"id": despesa_id}).mappings().first()

        if not nd or not despesa:
            abort(404)

        if despesa["origem"] != "OPERACIONAL":
            flash("ESSA TELA É APENAS PARA DESPESA OPERACIONAL.", "warning")
            return redirect(url_for("financeiro_dois.nota_debito_despesas", nd_id=nd_id))

    return render_template(
        "financeiro_dois/nota_debito_operacional.html",
        subnav_links=build_financeiro_dois_subnav("notas_debito"),
        nd=nd,
        despesa=despesa,
        pode_reverter=_usuario_eh_administrador(),
    )
    
@bp.route("/notas-debito/<int:nd_id>/despesas/<int:despesa_id>/operacional/reverter", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def nota_debito_operacional_reverter(nd_id: int, despesa_id: int):
    if not _usuario_eh_administrador():
        flash("APENAS O PERFIL ADMINISTRADOR PODE REVERTER DESPESA OPERACIONAL.", "danger")
        return redirect(url_for("financeiro_dois.nota_debito_operacional", nd_id=nd_id, despesa_id=despesa_id))

    engine = get_engine()
    with engine.begin() as conn:
        nd = conn.execute(text("""
            SELECT id, UPPER(COALESCE(numero_nd, '')) AS numero_nd
            FROM financeiro2_notas_debito
            WHERE id = :id
        """), {"id": nd_id}).mappings().first()

        despesa = conn.execute(text("""
            SELECT
                id,
                UPPER(COALESCE(origem_tipo, '')) AS origem,
                UPPER(COALESCE(nd_numero, '')) AS nd_numero,
                UPPER(COALESCE(numero_nd_desconsiderada, '')) AS numero_nd_desconsiderada
            FROM financeiro2_despesas
            WHERE id = :id
        """), {"id": despesa_id}).mappings().first()

        if not nd or not despesa:
            abort(404)

        if despesa["origem"] != "OPERACIONAL":
            flash("ESSA TELA É APENAS PARA DESPESA OPERACIONAL.", "warning")
            return redirect(url_for("financeiro_dois.nota_debito_despesas", nd_id=nd_id))

        if despesa["nd_numero"] != nd["numero_nd"] and despesa["numero_nd_desconsiderada"] != nd["numero_nd"]:
            flash("ESSA DESPESA NÃO ESTÁ RELACIONADA A ESTA ND.", "warning")
            return redirect(url_for("financeiro_dois.nota_debito_operacional", nd_id=nd_id, despesa_id=despesa_id))

        conn.execute(text("""
            UPDATE financeiro2_despesas
            SET
                status_nd = 'NÃO VINCULADA',
                nd_numero = CASE WHEN UPPER(COALESCE(nd_numero, '')) = :numero_nd THEN NULL ELSE nd_numero END,
                numero_nd_desconsiderada = CASE WHEN UPPER(COALESCE(numero_nd_desconsiderada, '')) = :numero_nd THEN NULL ELSE numero_nd_desconsiderada END,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = :id
        """), {"id": despesa_id, "numero_nd": nd["numero_nd"]})

        conn.execute(text("""
            DELETE FROM financeiro2_notas_debito_despesas
            WHERE nd_id = :nd_id
              AND despesa_id = :despesa_id
        """), {"nd_id": nd_id, "despesa_id": despesa_id})

        _recalcular_status_nd(conn, nd_id)

    flash("DESPESA OPERACIONAL REVERTIDA COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.nota_debito_operacional", nd_id=nd_id, despesa_id=despesa_id))
    
@bp.route("/om/<int:om_id>/desvincular")
@login_required
@permission_required("financeiro", "visualizar")
def om_desvincular(om_id: int):
    if not _usuario_eh_administrador():
        flash("APENAS O PERFIL ADMINISTRADOR PODE DESVINCULAR LINHAS DA OM.", "danger")
        return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

    engine = get_engine()
    with engine.connect() as conn:
        om = conn.execute(text("""
            SELECT
                id,
                numero_om AS numero,
                matricula_colaborador AS matricula,
                nome_colaborador AS colaborador,
                status
            FROM financeiro2_om
            WHERE id = :id
        """), {"id": om_id}).mappings().first()

        if not om:
            abort(404)

        if (om["status"] or "").upper() != "PAGA":
            flash("SÓ É POSSÍVEL DESVINCULAR QUANDO A OM ESTIVER PAGA.", "warning")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        linhas = conn.execute(text("""
            SELECT
                id,
                TO_CHAR(data_lancamento, 'DD/MM/YYYY') AS data,
                COALESCE(recibo, id) AS recibo,
                UPPER(COALESCE(tipo_linha, '')) AS descricao,
                UPPER(COALESCE(detalhes, '')) AS detalhes,
                UPPER(COALESCE(categoria, '')) AS categoria,
                UPPER(COALESCE(aplicacao, '')) AS aplicacao,
                COALESCE(valor_brl, 0) AS valor_brl,
                UPPER(COALESCE(numero_nd, '')) AS numero_nd,
                COALESCE(desconsiderada_nd, FALSE) AS desconsiderada_nd,
                UPPER(COALESCE(numero_nd_desconsiderada, '')) AS numero_nd_desconsiderada
            FROM financeiro2_om_linhas
            WHERE om_id = :om_id
              AND COALESCE(status, 'Ativo') = 'Ativo'
              AND COALESCE(valor_brl, 0) > 0
            ORDER BY id
        """), {"om_id": om_id}).mappings().all()

    return render_template(
        "financeiro_dois/om_desvincular.html",
        subnav_links=build_financeiro_dois_subnav("om"),
        om=om,
        linhas=linhas,
    )


@bp.route("/om/<int:om_id>/desvincular/salvar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def om_desvincular_salvar(om_id: int):
    if not _usuario_eh_administrador():
        flash("APENAS O PERFIL ADMINISTRADOR PODE DESVINCULAR LINHAS DA OM.", "danger")
        return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

    senha = _nome_preenchido(request.form.get("senha_confirmacao"))
    linha_ids = request.form.getlist("linha_ids")

    if not linha_ids:
        flash("SELECIONE AO MENOS UMA LINHA PARA DESVINCULAR.", "warning")
        return redirect(url_for("financeiro_dois.om_desvincular", om_id=om_id))

    engine = get_engine()
    with engine.begin() as conn:
        if not _validar_senha_usuario_atual(conn, senha):
            flash("SENHA INVÁLIDA.", "danger")
            return redirect(url_for("financeiro_dois.om_desvincular", om_id=om_id))

        om = conn.execute(text("""
            SELECT id, status
            FROM financeiro2_om
            WHERE id = :id
        """), {"id": om_id}).mappings().first()

        if not om:
            abort(404)

        if (om["status"] or "").upper() != "PAGA":
            flash("SÓ É POSSÍVEL DESVINCULAR QUANDO A OM ESTIVER PAGA.", "warning")
            return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))

        nds_afetadas = set()

        for linha_id in linha_ids:
            linha = conn.execute(text("""
                SELECT
                    id,
                    UPPER(COALESCE(numero_nd, '')) AS numero_nd,
                    COALESCE(desconsiderada_nd, FALSE) AS desconsiderada_nd,
                    UPPER(COALESCE(numero_nd_desconsiderada, '')) AS numero_nd_desconsiderada
                FROM financeiro2_om_linhas
                WHERE id = :linha_id
                  AND om_id = :om_id
                  AND COALESCE(status, 'Ativo') = 'Ativo'
                  AND COALESCE(valor_brl, 0) > 0
            """), {
                "linha_id": int(linha_id),
                "om_id": om_id,
            }).mappings().first()

            if not linha:
                continue

            if linha["numero_nd"]:
                nds_afetadas.add(linha["numero_nd"])
                conn.execute(text("""
                    UPDATE financeiro2_om_linhas
                    SET
                        numero_nd = NULL,
                        atualizado_em = CURRENT_TIMESTAMP
                    WHERE id = :linha_id
                      AND om_id = :om_id
                """), {
                    "linha_id": int(linha_id),
                    "om_id": om_id,
                })

            if linha["desconsiderada_nd"]:
                if linha["numero_nd_desconsiderada"]:
                    nds_afetadas.add(linha["numero_nd_desconsiderada"])
                conn.execute(text("""
                    UPDATE financeiro2_om_linhas
                    SET
                        desconsiderada_nd = FALSE,
                        numero_nd_desconsiderada = NULL,
                        atualizado_em = CURRENT_TIMESTAMP
                    WHERE id = :linha_id
                      AND om_id = :om_id
                """), {
                    "linha_id": int(linha_id),
                    "om_id": om_id,
                })

        _recalcular_status_despesa_origem(conn, "OM", om_id)

        for numero_nd in nds_afetadas:
            _recalcular_relacao_nd_despesa(conn, numero_nd, "OM", om_id)

    flash("LINHAS DA OM DESVINCULADAS COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.om_editar", om_id=om_id))
    
@bp.route("/rd/<int:rd_id>/desvincular")
@login_required
@permission_required("financeiro", "visualizar")
def rd_desvincular(rd_id: int):
    if not _usuario_eh_administrador():
        flash("APENAS O PERFIL ADMINISTRADOR PODE DESVINCULAR LINHAS DA RD.", "danger")
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    engine = get_engine()
    with engine.connect() as conn:
        rd = conn.execute(text("""
            SELECT
                id,
                numero_rd AS numero,
                periodo,
                matricula_colaborador AS matricula,
                nome_colaborador AS colaborador,
                status
            FROM financeiro2_rd
            WHERE id = :id
        """), {"id": rd_id}).mappings().first()

        if not rd:
            abort(404)

        if (rd["status"] or "").upper() != "QUITADA":
            flash("SÓ É POSSÍVEL DESVINCULAR QUANDO A RD ESTIVER QUITADA.", "warning")
            return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

        linhas = conn.execute(text("""
            SELECT
                id,
                TO_CHAR(data_lancamento, 'DD/MM/YYYY') AS data,
                id AS referencia,
                UPPER(COALESCE(descricao, '')) AS descricao,
                UPPER(COALESCE(categoria, '')) AS categoria,
                UPPER(COALESCE(aplicacao, '')) AS aplicacao,
                COALESCE(valor, 0) AS valor,
                UPPER(COALESCE(numero_nd, '')) AS numero_nd,
                COALESCE(desconsiderada_nd, FALSE) AS desconsiderada_nd,
                UPPER(COALESCE(numero_nd_desconsiderada, '')) AS numero_nd_desconsiderada
            FROM financeiro2_rd_linhas
            WHERE rd_id = :rd_id
              AND COALESCE(status, 'Ativo') = 'Ativo'
              AND COALESCE(valor, 0) > 0
            ORDER BY id
        """), {"rd_id": rd_id}).mappings().all()

    return render_template(
        "financeiro_dois/rd_desvincular.html",
        subnav_links=build_financeiro_dois_subnav("rd"),
        rd=rd,
        linhas=linhas,
    )


@bp.route("/rd/<int:rd_id>/desvincular/salvar", methods=["POST"])
@login_required
@permission_required("financeiro", "visualizar")
def rd_desvincular_salvar(rd_id: int):
    if not _usuario_eh_administrador():
        flash("APENAS O PERFIL ADMINISTRADOR PODE DESVINCULAR LINHAS DA RD.", "danger")
        return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

    senha = _nome_preenchido(request.form.get("senha_confirmacao"))
    linha_ids = request.form.getlist("linha_ids")

    if not linha_ids:
        flash("SELECIONE AO MENOS UMA LINHA PARA DESVINCULAR.", "warning")
        return redirect(url_for("financeiro_dois.rd_desvincular", rd_id=rd_id))

    engine = get_engine()
    with engine.begin() as conn:
        if not _validar_senha_usuario_atual(conn, senha):
            flash("SENHA INVÁLIDA.", "danger")
            return redirect(url_for("financeiro_dois.rd_desvincular", rd_id=rd_id))

        rd = conn.execute(text("""
            SELECT id, status
            FROM financeiro2_rd
            WHERE id = :id
        """), {"id": rd_id}).mappings().first()

        if not rd:
            abort(404)

        if (rd["status"] or "").upper() != "QUITADA":
            flash("SÓ É POSSÍVEL DESVINCULAR QUANDO A RD ESTIVER QUITADA.", "warning")
            return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))

        nds_afetadas = set()

        for linha_id in linha_ids:
            linha = conn.execute(text("""
                SELECT
                    id,
                    UPPER(COALESCE(numero_nd, '')) AS numero_nd,
                    COALESCE(desconsiderada_nd, FALSE) AS desconsiderada_nd,
                    UPPER(COALESCE(numero_nd_desconsiderada, '')) AS numero_nd_desconsiderada
                FROM financeiro2_rd_linhas
                WHERE id = :linha_id
                  AND rd_id = :rd_id
                  AND COALESCE(status, 'Ativo') = 'Ativo'
                  AND COALESCE(valor, 0) > 0
            """), {
                "linha_id": int(linha_id),
                "rd_id": rd_id,
            }).mappings().first()

            if not linha:
                continue

            if linha["numero_nd"]:
                nds_afetadas.add(linha["numero_nd"])
                conn.execute(text("""
                    UPDATE financeiro2_rd_linhas
                    SET
                        numero_nd = NULL,
                        atualizado_em = CURRENT_TIMESTAMP
                    WHERE id = :linha_id
                      AND rd_id = :rd_id
                """), {
                    "linha_id": int(linha_id),
                    "rd_id": rd_id,
                })

            if linha["desconsiderada_nd"]:
                if linha["numero_nd_desconsiderada"]:
                    nds_afetadas.add(linha["numero_nd_desconsiderada"])
                conn.execute(text("""
                    UPDATE financeiro2_rd_linhas
                    SET
                        desconsiderada_nd = FALSE,
                        numero_nd_desconsiderada = NULL,
                        atualizado_em = CURRENT_TIMESTAMP
                    WHERE id = :linha_id
                      AND rd_id = :rd_id
                """), {
                    "linha_id": int(linha_id),
                    "rd_id": rd_id,
                })

        _recalcular_status_despesa_origem(conn, "RD", rd_id)

        for numero_nd in nds_afetadas:
            _recalcular_relacao_nd_despesa(conn, numero_nd, "RD", rd_id)

    flash("LINHAS DA RD DESVINCULADAS COM SUCESSO.", "success")
    return redirect(url_for("financeiro_dois.rd_editar", rd_id=rd_id))