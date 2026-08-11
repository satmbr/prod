from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader, PdfWriter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


AZUL = "173B5F"
AZUL_CLARO = "EAF1F7"
VERDE = "177E64"
CINZA = "667085"
BORDA = "D8E1EA"


def _valor(numero) -> Decimal:
    return Decimal(str(numero or 0))


def _texto(valor) -> str:
    return "" if valor is None else str(valor)


def _data(valor) -> str:
    if isinstance(valor, (date, datetime)):
        return valor.strftime("%d/%m/%Y")
    return _texto(valor)


def _moeda(valor, codigo: str) -> str:
    numero = _valor(valor)
    formatado = f"{numero:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    return f"{codigo} {formatado}"


def nome_base_om(numero_om: str) -> str:
    seguro = "".join(c if c.isalnum() or c in "-_" else "-" for c in _texto(numero_om).strip())
    seguro = "-".join(parte for parte in seguro.split("-") if parte)
    if seguro.upper().startswith("OM-"):
        return seguro
    return f"OM-{seguro or 'sem-numero'}"


def gerar_excel_om(om: Mapping, itens: Iterable[Mapping], pagamentos: Iterable[Mapping]) -> BytesIO:
    itens = list(itens)
    pagamentos = list(pagamentos)
    moeda = _texto(om.get("moeda")) or "BRL"
    total = sum((_valor(item.get("valor")) for item in itens), Decimal("0"))
    total_reembolsos = _valor(om.get("valor_reembolsos"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo OM"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A11"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"ORDEM DE MISSÃO · {om.get('numero_om', '')}"
    ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34

    metadados = [
        ("Favorecido", om.get("solicitante")),
        ("Matrícula", om.get("matricula_favorecido")),
        ("Centro padrão", f"{om.get('centro_codigo', '')} · {om.get('centro_nome', '')}"),
        ("Status", _texto(om.get("status")).replace("_", " ")),
        ("Moeda", moeda),
        ("Observações", om.get("observacoes") or "—"),
    ]
    for indice, (rotulo, valor) in enumerate(metadados, start=3):
        coluna = 1 if indice <= 5 else 4
        linha = indice if indice <= 5 else indice - 3
        ws.cell(linha, coluna, rotulo).font = Font(bold=True, color=CINZA)
        ws.merge_cells(start_row=linha, start_column=coluna + 1, end_row=linha, end_column=coluna + 2)
        ws.cell(linha, coluna + 1, _texto(valor)).alignment = Alignment(wrap_text=True)

    ws["A7"] = "Total das linhas"
    ws["A8"] = "Reembolsos"
    ws["D7"] = "Total geral"
    for celula in ("A7", "A8", "D7"):
        ws[celula].font = Font(bold=True, color=CINZA)
    ws["B7"] = float(total)
    ws["B8"] = float(total_reembolsos)
    ws["E7"] = "=B7+B8"
    for celula in ("B7", "B8", "E7"):
        ws[celula].number_format = f'"{moeda}" #,##0.00;[Red]-"{moeda}" #,##0.00;–'
        ws[celula].font = Font(size=12, bold=True, color=VERDE if celula == "E7" else AZUL)

    cabecalho_linha = 10
    cabecalhos = ["Linha", "Data", "Centro de custo", "Categoria", "Descrição", "Recibo", "Valor"]
    for coluna, titulo in enumerate(cabecalhos, start=1):
        celula = ws.cell(cabecalho_linha, coluna, titulo)
        celula.fill = PatternFill("solid", fgColor=AZUL)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.alignment = Alignment(vertical="center")
    ws.row_dimensions[cabecalho_linha].height = 24

    primeira = cabecalho_linha + 1
    for linha, item in enumerate(itens, start=primeira):
        recibo = item.get("nome_original") if item.get("arquivo_id") else (
            f"Sem recibo · {item.get('justificativa_sem_comprovante')}"
            if item.get("justificativa_sem_comprovante") else "Sem recibo"
        )
        ws.append([
            int(item.get("numero_linha") or linha - cabecalho_linha),
            item.get("data_despesa"),
            f"{item.get('centro_codigo', '')} · {item.get('centro_nome', '')}",
            item.get("categoria"),
            item.get("descricao"),
            recibo,
            float(_valor(item.get("valor"))),
        ])
        ws.cell(linha, 2).number_format = "dd/mm/yyyy"
        ws.cell(linha, 7).number_format = f'"{moeda}" #,##0.00;[Red]-"{moeda}" #,##0.00;–'
        ws.cell(linha, 7).alignment = Alignment(horizontal="right")
        for coluna in range(1, 8):
            ws.cell(linha, coluna).alignment = Alignment(
                vertical="top", wrap_text=coluna in {3, 4, 5, 6},
                horizontal="center" if coluna in {1, 2} else None,
            )
        if linha % 2 == 0:
            for coluna in range(1, 8):
                ws.cell(linha, coluna).fill = PatternFill("solid", fgColor="F7F9FB")

    ultima = max(primeira, primeira + len(itens) - 1)
    total_linha = primeira + len(itens)
    ws.cell(total_linha, 6, "Total das linhas").font = Font(bold=True, color=AZUL)
    ws.cell(total_linha, 7, f"=SUM(G{primeira}:G{ultima})" if itens else "=0")
    ws.cell(total_linha, 7).number_format = f'"{moeda}" #,##0.00;[Red]-"{moeda}" #,##0.00;–'
    ws.cell(total_linha, 7).font = Font(bold=True, color=VERDE)
    ws.auto_filter.ref = f"A{cabecalho_linha}:G{ultima}" if itens else f"A{cabecalho_linha}:G{cabecalho_linha}"

    larguras = [14, 13, 28, 24, 42, 30, 18]
    for indice, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = largura
    ws.print_title_rows = f"1:{cabecalho_linha}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    ws.page_margins.left = ws.page_margins.right = 0.25

    pg = wb.create_sheet("Pagamentos")
    pg.sheet_view.showGridLines = False
    pg.append(["Tipo", "Data do pagamento", "Status", "Comprovante", "Observações", "Valor"])
    for celula in pg[1]:
        celula.fill = PatternFill("solid", fgColor=AZUL)
        celula.font = Font(bold=True, color="FFFFFF")
    for pagamento in pagamentos:
        pg.append([
            _texto(pagamento.get("tipo")).replace("_", " "),
            pagamento.get("data_pagamento") or pagamento.get("data_prevista_pagamento"),
            "REGISTRADO" if pagamento.get("status") == "PAGO" else pagamento.get("status"),
            pagamento.get("nome_original") or "—",
            pagamento.get("observacoes") or "—",
            float(_valor(pagamento.get("valor"))),
        ])
    for linha in range(2, pg.max_row + 1):
        pg.cell(linha, 2).number_format = "dd/mm/yyyy"
        pg.cell(linha, 6).number_format = f'"{moeda}" #,##0.00;[Red]-"{moeda}" #,##0.00;–'
    for indice, largura in enumerate([18, 20, 18, 30, 42, 18], start=1):
        pg.column_dimensions[get_column_letter(indice)].width = largura
    pg.freeze_panes = "A2"
    pg.auto_filter.ref = f"A1:F{max(pg.max_row, 1)}"

    borda = Side(style="thin", color=BORDA)
    for aba in (ws, pg):
        for row in aba.iter_rows():
            for celula in row:
                if celula.value is not None:
                    celula.border = Border(bottom=borda)
                    if celula.font.name == "Calibri":
                        celula.font = Font(name="Aptos", size=10, bold=celula.font.bold, color=celula.font.color)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo


def _rodape_pdf(canvas_pdf, documento):
    canvas_pdf.saveState()
    canvas_pdf.setStrokeColor(colors.HexColor("#D8E1EA"))
    canvas_pdf.line(14 * mm, 10 * mm, landscape(A4)[0] - 14 * mm, 10 * mm)
    canvas_pdf.setFont("Helvetica", 7.5)
    canvas_pdf.setFillColor(colors.HexColor("#667085"))
    canvas_pdf.drawString(14 * mm, 6 * mm, "Resumo da Ordem de Missão")
    canvas_pdf.drawRightString(landscape(A4)[0] - 14 * mm, 6 * mm, f"Página {documento.page}")
    canvas_pdf.restoreState()


def _resumo_pdf(om: Mapping, itens: list[Mapping]) -> BytesIO:
    buffer = BytesIO()
    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle("TituloOM", parent=estilos["Title"], fontName="Helvetica-Bold", fontSize=18,
                            leading=22, textColor=colors.HexColor("#173B5F"), alignment=TA_LEFT)
    rotulo = ParagraphStyle("RotuloOM", parent=estilos["Normal"], fontName="Helvetica-Bold", fontSize=7.5,
                            leading=9, textColor=colors.HexColor("#667085"))
    corpo = ParagraphStyle("CorpoOM", parent=estilos["Normal"], fontName="Helvetica", fontSize=8.5,
                           leading=11, textColor=colors.HexColor("#172B3A"))
    tabela_texto = ParagraphStyle("TabelaOM", parent=corpo, fontSize=7.2, leading=9)
    tabela_direita = ParagraphStyle("TabelaDireitaOM", parent=tabela_texto, alignment=TA_RIGHT)
    tabela_centro = ParagraphStyle("TabelaCentroOM", parent=tabela_texto, alignment=TA_CENTER)

    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=12 * mm, bottomMargin=14 * mm, title=f"OM {om.get('numero_om', '')}")
    moeda = _texto(om.get("moeda")) or "BRL"
    total = sum((_valor(item.get("valor")) for item in itens), Decimal("0"))
    reembolsos = _valor(om.get("valor_reembolsos"))
    recibos = sum(1 for item in itens if item.get("arquivo_id"))
    disponiveis = sum(1 for item in itens if item.get("caminho_recibo"))

    historia = [
        Paragraph(f"Ordem de Missão · {om.get('numero_om', '')}", titulo),
        Spacer(1, 3 * mm),
    ]
    dados = [
        [Paragraph("FAVORECIDO", rotulo), Paragraph("MATRÍCULA", rotulo), Paragraph("CENTRO PADRÃO", rotulo), Paragraph("STATUS", rotulo)],
        [Paragraph(_texto(om.get("solicitante")), corpo), Paragraph(_texto(om.get("matricula_favorecido")), corpo),
         Paragraph(f"{om.get('centro_codigo', '')} · {om.get('centro_nome', '')}", corpo),
         Paragraph(_texto(om.get("status")).replace("_", " "), corpo)],
    ]
    quadro = Table(dados, colWidths=[78 * mm, 42 * mm, 75 * mm, 52 * mm], rowHeights=[6 * mm, 10 * mm])
    quadro.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F7F9FB")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8E1EA")),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#E7EDF3")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
    ]))
    historia.extend([quadro, Spacer(1, 4 * mm)])

    kpis = [
        [Paragraph("TOTAL DAS LINHAS", rotulo), Paragraph("REEMBOLSOS VINCULADOS", rotulo), Paragraph("TOTAL GERAL", rotulo), Paragraph("RECIBOS", rotulo)],
        [Paragraph(_moeda(total, moeda), corpo), Paragraph(_moeda(reembolsos, moeda), corpo),
         Paragraph(_moeda(total + reembolsos, moeda), corpo), Paragraph(f"{disponiveis} disponíveis de {recibos} vinculados", corpo)],
    ]
    cards = Table(kpis, colWidths=[62 * mm] * 4, rowHeights=[6 * mm, 11 * mm])
    cards.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EAF1F7")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFD0DF")),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#BFD0DF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
    ]))
    historia.extend([cards, Spacer(1, 5 * mm), Paragraph("Resumo das despesas", ParagraphStyle(
        "SecaoOM", parent=titulo, fontSize=11, leading=14)), Spacer(1, 2 * mm)])

    cabecalho = ["#", "Data", "Centro de custo", "Categoria", "Descrição", "Recibo", "Valor"]
    linhas = [[Paragraph(c, ParagraphStyle("CabecalhoOM", parent=tabela_texto, textColor=colors.white,
                                            fontName="Helvetica-Bold", alignment=TA_CENTER)) for c in cabecalho]]
    for item in itens:
        if item.get("arquivo_id"):
            recibo = item.get("nome_original") or "Recibo"
            if not item.get("caminho_recibo"):
                recibo = f"{recibo} · indisponível"
        else:
            recibo = "Sem recibo"
            if item.get("justificativa_sem_comprovante"):
                recibo += f" · {item.get('justificativa_sem_comprovante')}"
        linhas.append([
            Paragraph(_texto(item.get("numero_linha")), tabela_centro),
            Paragraph(_data(item.get("data_despesa")), tabela_centro),
            Paragraph(f"{item.get('centro_codigo', '')} · {item.get('centro_nome', '')}", tabela_texto),
            Paragraph(_texto(item.get("categoria")), tabela_texto),
            Paragraph(_texto(item.get("descricao")), tabela_texto),
            Paragraph(_texto(recibo), tabela_texto),
            Paragraph(_moeda(item.get("valor"), moeda), tabela_direita),
        ])
    if not itens:
        linhas.append([Paragraph("Nenhuma linha de despesa registrada.", tabela_texto), "", "", "", "", "", ""])
    tabela = Table(linhas, colWidths=[10 * mm, 21 * mm, 43 * mm, 34 * mm, 62 * mm, 51 * mm, 28 * mm], repeatRows=1)
    estilo_tabela = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#173B5F")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D8E1EA")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for indice in range(1, len(linhas)):
        if indice % 2 == 0:
            estilo_tabela.append(("BACKGROUND", (0, indice), (-1, indice), colors.HexColor("#F7F9FB")))
    if not itens:
        estilo_tabela.append(("SPAN", (0, 1), (-1, 1)))
    tabela.setStyle(TableStyle(estilo_tabela))
    historia.append(tabela)
    if om.get("observacoes"):
        historia.extend([Spacer(1, 4 * mm), KeepTogether([
            Paragraph("Observações", rotulo), Spacer(1, 1 * mm), Paragraph(_texto(om.get("observacoes")), corpo)
        ])])
    historia.extend([Spacer(1, 4 * mm), Paragraph(
        "Os recibos vinculados são apresentados após este resumo, seguindo exatamente a ordem numérica das linhas.",
        ParagraphStyle("NotaOM", parent=corpo, textColor=colors.HexColor("#667085")))])
    doc.build(historia, onFirstPage=_rodape_pdf, onLaterPages=_rodape_pdf)
    buffer.seek(0)
    return buffer


def _pagina_indisponivel(item: Mapping) -> BytesIO:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    largura, altura = A4
    pdf.setFillColor(colors.HexColor("#173B5F"))
    pdf.rect(0, altura - 32 * mm, largura, 32 * mm, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(20 * mm, altura - 20 * mm, f"Recibo da linha {item.get('numero_linha', '')}")
    pdf.setFillColor(colors.HexColor("#172B3A"))
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(20 * mm, altura - 58 * mm, "Arquivo indisponível")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(20 * mm, altura - 68 * mm, "O vínculo do recibo existe, mas o arquivo não pôde ser lido no armazenamento.")
    pdf.drawString(20 * mm, altura - 76 * mm, f"Arquivo: {_texto(item.get('nome_original')) or 'não informado'}")
    pdf.drawString(20 * mm, altura - 84 * mm, f"Despesa: {_texto(item.get('descricao'))[:90]}")
    pdf.save()
    buffer.seek(0)
    return buffer


def gerar_pdf_om(om: Mapping, itens: Iterable[Mapping]) -> BytesIO:
    itens = list(itens)
    resumo = _resumo_pdf(om, itens)
    writer = PdfWriter()
    for pagina in PdfReader(resumo).pages:
        writer.add_page(pagina)
    writer.add_outline_item("Resumo da OM", 0)

    for item in itens:
        if not item.get("arquivo_id"):
            continue
        inicio = len(writer.pages)
        origem = None
        caminho = item.get("caminho_recibo")
        try:
            if caminho:
                origem = BytesIO(Path(caminho).read_bytes())
                leitor = PdfReader(origem)
                if not leitor.pages:
                    raise ValueError("PDF sem páginas")
            else:
                leitor = PdfReader(_pagina_indisponivel(item))
            for pagina in leitor.pages:
                writer.add_page(pagina)
        except Exception:
            for pagina in PdfReader(_pagina_indisponivel(item)).pages:
                writer.add_page(pagina)
        writer.add_outline_item(f"Linha {item.get('numero_linha')} · {_texto(item.get('descricao'))[:55]}", inicio)

    writer.add_metadata({
        "/Title": f"OM {om.get('numero_om', '')}",
        "/Subject": "Resumo da Ordem de Missão e recibos ordenados por linha",
        "/Creator": "Sistema PROD · Financeiro Novo",
    })
    arquivo = BytesIO()
    writer.write(arquivo)
    arquivo.seek(0)
    return arquivo
