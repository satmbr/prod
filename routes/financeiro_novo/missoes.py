import os
import re
import unicodedata
import uuid
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile

from flask import abort, current_app, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError

from db import get_engine
from routes.auth import login_required, permission_required
from routes.financeiro_novo import bp
from routes.financeiro_novo.despesas import FORMAS_PAGAMENTO, _opcoes
from routes.financeiro_novo.services.anexos import AnexoInvalido, nome_objeto_pdf, normalizar_anexo
from routes.financeiro_novo.services.auditoria import registrar_evento
from routes.financeiro_novo.services.exportacao_om import gerar_excel_om, gerar_pdf_om, nome_base_om
from routes.financeiro_novo.services.valores import ValorInvalido, data_iso, decimal_br
from routes.financeiro_novo.views import build_subnav


EDITAVEIS = {"RASCUNHO", "REJEITADA"}
DOCUMENTOS = {
    "om": ("financeiro3_oms", "OM"),
    "rd": ("financeiro3_rds", "RD"),
    "nd": ("financeiro3_notas_debito", "ND"),
}


def _registro(conn, tabela, registro_id, bloquear=False):
    ativo = " AND removido_em IS NULL" if tabela == "financeiro3_oms" else ""
    return conn.execute(
        text(f"SELECT * FROM {tabela} WHERE id=:id{ativo}" + (" FOR UPDATE" if bloquear else "")),
        {"id": registro_id},
    ).mappings().first()


def _proibir_autoaprovacao(registro):
    return (
        registro["criado_por"] == session.get("usuario_id")
        and "auth:administrar" not in session.get("permissoes", [])
    )


def _tem_anexo(conn, entidade, registro_id):
    return bool(conn.execute(text("""
        SELECT COUNT(*) FROM financeiro3_anexos
        WHERE entidade=:entidade AND entidade_id=:id AND status='ATIVO'
    """), {"entidade": entidade, "id": registro_id}).scalar())


@bp.get("/missoes")
@login_required
@permission_required("financeiro_novo", "visualizar")
def missoes():
    with get_engine().connect() as conn:
        oms = conn.execute(text("""
            SELECT o.*, p.nome_razao AS solicitante, cc.codigo AS centro, m.codigo AS moeda,
              COALESCE((SELECT SUM(pg.valor) FROM financeiro3_om_pagamentos pg
                WHERE pg.om_id=o.id AND pg.status='PAGO'),0) AS valor_pago,
              COALESCE((SELECT SUM(pg.valor) FROM financeiro3_om_pagamentos pg
                WHERE pg.om_id=o.id AND pg.status='PREVISTO'),0) AS valor_previsto,
              COALESCE((SELECT SUM(i.valor) FROM financeiro3_reembolso_itens i
                JOIN financeiro3_reembolsos rb ON rb.id=i.reembolso_id
                WHERE rb.om_pagadora_id=o.id AND rb.forma_liquidacao='OM'
                  AND rb.status='APROVADO' AND i.status='ATIVO'),0) AS valor_reembolsos
            FROM financeiro3_oms o JOIN financeiro3_pessoas p ON p.id=o.solicitante_id
            JOIN financeiro3_centros_custo cc ON cc.id=o.centro_custo_id
            JOIN financeiro3_moedas m ON m.id=o.moeda_id
            WHERE o.removido_em IS NULL ORDER BY o.id DESC LIMIT 500
        """)).mappings().all()
        rds = conn.execute(text("""
            SELECT r.*,p.nome_razao AS responsavel,cc.codigo AS centro,m.codigo AS moeda,
              COALESCE((SELECT SUM(pg.valor) FROM financeiro3_rd_pagamentos pg
                WHERE pg.rd_id=r.id AND pg.status='PAGO'),0) AS valor_pago,
              COALESCE((SELECT SUM(i.valor) FROM financeiro3_reembolso_itens i
                JOIN financeiro3_reembolsos rb ON rb.id=i.reembolso_id
                WHERE rb.rd_pagadora_id=r.id AND rb.forma_liquidacao='RD'
                  AND rb.status='APROVADO' AND i.status='ATIVO'),0) AS valor_reembolsos
            FROM financeiro3_rds r
            JOIN financeiro3_pessoas p ON p.id=r.responsavel_id
            JOIN financeiro3_centros_custo cc ON cc.id=r.centro_custo_id
            JOIN financeiro3_moedas m ON m.id=r.moeda_id
            ORDER BY r.id DESC LIMIT 500
        """)).mappings().all()
        acertos = conn.execute(text("""
            SELECT a.*, p.nome_razao AS responsavel, m.codigo AS moeda
            FROM financeiro3_rd_acertos a JOIN financeiro3_rds r ON r.id=a.rd_id
            JOIN financeiro3_pessoas p ON p.id=r.responsavel_id
            JOIN financeiro3_moedas m ON m.id=r.moeda_id
            ORDER BY a.id DESC LIMIT 200
        """)).mappings().all()
    return render_template(
        "financeiro_novo/missoes.html", oms=oms, rds=rds, acertos=acertos,
        subnav_links=build_subnav("missoes"),
    )


def _dados_om(form):
    try:
        dados = {
            "solicitante_id": int(form.get("solicitante_id") or 0),
            "centro_custo_id": int(form.get("centro_custo_id") or 0),
            "moeda_id": int(form.get("moeda_id") or 0),
        }
    except ValueError as exc:
        raise ValorInvalido("Selecione cadastros válidos.") from exc
    dados.update({
        "numero_om": (form.get("numero_om") or "").strip().upper(),
        "matricula_favorecido": (form.get("matricula_favorecido") or "").strip().upper(),
        "observacoes": (form.get("observacoes") or "").strip() or None,
    })
    if not dados["numero_om"] or not dados["matricula_favorecido"]:
        raise ValorInvalido("Informe o número da OM e a matrícula do favorecido.")
    if len(dados["numero_om"]) > 80 or len(dados["matricula_favorecido"]) > 40:
        raise ValorInvalido("Número da OM ou matrícula excede o tamanho permitido.")
    return dados


def _validar_om_refs(conn, dados):
    ok = conn.execute(text("""
        SELECT EXISTS(SELECT 1 FROM financeiro3_pessoas WHERE id=:solicitante_id AND ativo AND favorecido)
           AND EXISTS(SELECT 1 FROM financeiro3_centros_custo WHERE id=:centro_custo_id AND ativo)
           AND EXISTS(SELECT 1 FROM financeiro3_moedas WHERE id=:moeda_id AND ativo)
    """), dados).scalar()
    if not ok:
        raise ValorInvalido("Solicitante, centro de custo ou moeda está inativo ou inválido.")


@bp.route("/oms/nova", methods=["GET", "POST"])
@login_required
@permission_required("financeiro_novo", "criar")
def om_nova():
    with get_engine().connect() as conn:
        opcoes = _opcoes(conn)
    if request.method == "POST":
        try:
            dados = _dados_om(request.form)
            dados["usuario"] = session.get("usuario_id")
            with get_engine().begin() as conn:
                _validar_om_refs(conn, dados)
                om = conn.execute(text("""
                    INSERT INTO financeiro3_oms(numero_om,matricula_favorecido,solicitante_id,
                        centro_custo_id,moeda_id,observacoes,criado_por)
                    VALUES (:numero_om,:matricula_favorecido,:solicitante_id,
                        :centro_custo_id,:moeda_id,:observacoes,:usuario) RETURNING *
                """), dados).mappings().one()
                registrar_evento(conn, entidade="OM", entidade_id=om["id"], evento="CRIADA", dados_novos=dict(om))
            flash("OM criada em rascunho.", "sucesso")
            return redirect(url_for("financeiro_novo.om_detalhe", om_id=om["id"]))
        except (ValorInvalido, IntegrityError) as exc:
            if isinstance(exc, IntegrityError):
                exc = ValorInvalido("Já existe uma OM ativa com este número.")
            flash(str(exc), "erro")
    return render_template("financeiro_novo/om_form.html", opcoes=opcoes, subnav_links=build_subnav("missoes"))


@bp.get("/oms/<int:om_id>")
@login_required
@permission_required("financeiro_novo", "visualizar")
def om_detalhe(om_id):
    from routes.financeiro_novo.homologacao import diagnosticar_armazenamento
    with get_engine().connect() as conn:
        om = conn.execute(text("""
            SELECT o.*, p.nome_razao AS solicitante, cc.codigo AS centro_codigo,
                   cc.nome AS centro_nome, m.codigo AS moeda,
              COALESCE((SELECT SUM(i.valor) FROM financeiro3_reembolso_itens i
                JOIN financeiro3_reembolsos rb ON rb.id=i.reembolso_id
                WHERE rb.om_pagadora_id=o.id AND rb.forma_liquidacao='OM'
                  AND rb.status='APROVADO' AND i.status='ATIVO'),0) AS valor_reembolsos
            FROM financeiro3_oms o JOIN financeiro3_pessoas p ON p.id=o.solicitante_id
            JOIN financeiro3_centros_custo cc ON cc.id=o.centro_custo_id
            JOIN financeiro3_moedas m ON m.id=o.moeda_id
            WHERE o.id=:id AND o.removido_em IS NULL
        """), {"id": om_id}).mappings().first()
        if not om:
            abort(404)
        decisoes = conn.execute(text(
            "SELECT * FROM financeiro3_om_decisoes WHERE om_id=:id ORDER BY id DESC"
        ), {"id": om_id}).mappings().all()
        anexos = _listar_anexos(conn, "OM", om_id)
        itens = conn.execute(text("""
            SELECT i.*,ROW_NUMBER() OVER (ORDER BY i.id) AS numero_linha,
              c.nome AS categoria,cc.codigo AS centro_codigo,cc.nome AS centro_nome,
              a.id AS anexo_id,ar.id AS arquivo_id,ar.nome_original,ar.paginas
            FROM financeiro3_om_itens i JOIN financeiro3_categorias c ON c.id=i.categoria_id
            JOIN financeiro3_centros_custo cc ON cc.id=i.centro_custo_id
            LEFT JOIN financeiro3_anexos a ON a.entidade='OM_ITEM' AND a.entidade_id=i.id AND a.status='ATIVO'
            LEFT JOIN financeiro3_arquivos ar ON ar.id=a.arquivo_id
            WHERE i.om_id=:id AND i.status='ATIVO' ORDER BY i.id
        """), {"id": om_id}).mappings().all()
        pagamentos = conn.execute(text("""
            SELECT pg.*,a.id AS anexo_id,ar.id AS arquivo_id,ar.nome_original
            FROM financeiro3_om_pagamentos pg
            LEFT JOIN financeiro3_anexos a ON a.entidade='OM_PAGAMENTO' AND a.entidade_id=pg.id AND a.status='ATIVO'
            LEFT JOIN financeiro3_arquivos ar ON ar.id=a.arquivo_id AND ar.status='ATIVO'
            WHERE pg.om_id=:id ORDER BY pg.data_prevista_pagamento,pg.id
        """), {"id": om_id}).mappings().all()
        despesa_importada = conn.execute(text(
            "SELECT id FROM financeiro3_despesas WHERE origem_om_id=:id"
        ), {"id": om_id}).scalar()
        opcoes = _opcoes(conn)
    valor_pago = sum((pg["valor"] for pg in pagamentos if pg["status"] == "PAGO"), start=0)
    valor_previsto = sum((pg["valor"] for pg in pagamentos if pg["status"] == "PREVISTO"), start=0)
    return render_template(
        "financeiro_novo/om_detalhe.html", om=om, itens=itens, decisoes=decisoes, anexos=anexos,
        pagamentos=pagamentos, valor_pago=valor_pago, valor_previsto=valor_previsto,
        despesa_importada=despesa_importada,
        armazenamento=diagnosticar_armazenamento(),
        opcoes=opcoes, editavel=om["status"] in EDITAVEIS,
        subnav_links=build_subnav("missoes"),
    )


def _dados_exportacao_om(om_id):
    with get_engine().connect() as conn:
        om = conn.execute(text("""
            SELECT o.*, p.nome_razao AS solicitante, cc.codigo AS centro_codigo,
                   cc.nome AS centro_nome, m.codigo AS moeda,
              COALESCE((SELECT SUM(i.valor) FROM financeiro3_reembolso_itens i
                JOIN financeiro3_reembolsos rb ON rb.id=i.reembolso_id
                WHERE rb.om_pagadora_id=o.id AND rb.forma_liquidacao='OM'
                  AND rb.status='APROVADO' AND i.status='ATIVO'),0) AS valor_reembolsos
            FROM financeiro3_oms o
            JOIN financeiro3_pessoas p ON p.id=o.solicitante_id
            JOIN financeiro3_centros_custo cc ON cc.id=o.centro_custo_id
            JOIN financeiro3_moedas m ON m.id=o.moeda_id
            WHERE o.id=:id AND o.removido_em IS NULL
        """), {"id": om_id}).mappings().first()
        if not om:
            abort(404)
        itens = conn.execute(text("""
            SELECT i.*, ROW_NUMBER() OVER (ORDER BY i.id) AS numero_linha,
              c.nome AS categoria, cc.codigo AS centro_codigo, cc.nome AS centro_nome,
              ar.id AS arquivo_id, ar.nome_original, ar.object_key
            FROM financeiro3_om_itens i
            JOIN financeiro3_categorias c ON c.id=i.categoria_id
            JOIN financeiro3_centros_custo cc ON cc.id=i.centro_custo_id
            LEFT JOIN LATERAL (
              SELECT arq.id, arq.nome_original, arq.object_key
              FROM financeiro3_anexos a
              JOIN financeiro3_arquivos arq ON arq.id=a.arquivo_id AND arq.status='ATIVO'
              WHERE a.entidade='OM_ITEM' AND a.entidade_id=i.id AND a.status='ATIVO'
              ORDER BY a.id DESC LIMIT 1
            ) ar ON TRUE
            WHERE i.om_id=:id AND i.status='ATIVO'
            ORDER BY i.id
        """), {"id": om_id}).mappings().all()
        pagamentos = conn.execute(text("""
            SELECT pg.*, ar.id AS arquivo_id, ar.nome_original
            FROM financeiro3_om_pagamentos pg
            LEFT JOIN LATERAL (
              SELECT arq.id, arq.nome_original
              FROM financeiro3_anexos a
              JOIN financeiro3_arquivos arq ON arq.id=a.arquivo_id AND arq.status='ATIVO'
              WHERE a.entidade='OM_PAGAMENTO' AND a.entidade_id=pg.id AND a.status='ATIVO'
              ORDER BY a.id DESC LIMIT 1
            ) ar ON TRUE
            WHERE pg.om_id=:id
            ORDER BY COALESCE(pg.data_pagamento,pg.data_prevista_pagamento),pg.id
        """), {"id": om_id}).mappings().all()

    itens_exportacao = []
    for item in itens:
        dados = dict(item)
        caminho = _caminho(item["object_key"]) if item.get("object_key") else None
        dados["caminho_recibo"] = caminho if caminho and caminho.is_file() else None
        itens_exportacao.append(dados)
    return dict(om), itens_exportacao, [dict(pg) for pg in pagamentos]


@bp.get("/oms/<int:om_id>/exportar/excel")
@login_required
@permission_required("financeiro_novo", "visualizar")
def om_exportar_excel(om_id):
    om, itens, pagamentos = _dados_exportacao_om(om_id)
    return send_file(
        gerar_excel_om(om, itens, pagamentos),
        as_attachment=True,
        download_name=f"{nome_base_om(om['numero_om'])}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/oms/<int:om_id>/exportar/om")
@login_required
@permission_required("financeiro_novo", "visualizar")
def om_exportar_pdf(om_id):
    om, itens, _ = _dados_exportacao_om(om_id)
    return send_file(
        gerar_pdf_om(om, itens),
        as_attachment=True,
        download_name=f"{nome_base_om(om['numero_om'])}.pdf",
        mimetype="application/pdf",
    )


@bp.post("/oms/<int:om_id>/editar")
@login_required
@permission_required("financeiro_novo", "editar")
def om_editar(om_id):
    try:
        dados = _dados_om(request.form)
        dados.update({"id": om_id, "usuario": session.get("usuario_id")})
        with get_engine().begin() as conn:
            anterior = _registro(conn, "financeiro3_oms", om_id, True)
            if not anterior:
                abort(404)
            if anterior["status"] not in EDITAVEIS:
                abort(409)
            _validar_om_refs(conn, dados)
            novo = conn.execute(text("""
                UPDATE financeiro3_oms SET numero_om=:numero_om,matricula_favorecido=:matricula_favorecido,
                    solicitante_id=:solicitante_id,centro_custo_id=:centro_custo_id,moeda_id=:moeda_id,
                    observacoes=:observacoes,atualizado_por=:usuario,atualizado_em=NOW()
                WHERE id=:id RETURNING *
            """), dados).mappings().one()
            registrar_evento(conn, entidade="OM", entidade_id=om_id, evento="EDITADA", dados_anteriores=dict(anterior), dados_novos=dict(novo))
        flash("OM atualizada.", "sucesso")
    except (ValorInvalido, IntegrityError) as exc:
        if isinstance(exc, IntegrityError):
            exc = ValorInvalido("Já existe uma OM ativa com este número.")
        flash(str(exc), "erro")
    return redirect(url_for("financeiro_novo.om_detalhe", om_id=om_id))


def _origens_duplicadas(conn, data, valor):
    registros = conn.execute(text("""
        SELECT * FROM (
          SELECT 'OM'::text AS tipo,o.id AS origem_id,i.id AS item_id,o.numero_om AS numero,
            i.data_despesa AS data,i.descricao,c.nome AS categoria,
            cc.codigo||' · '||cc.nome AS centro,i.valor,m.codigo AS moeda,
            ar.id AS arquivo_id,ar.nome_original AS nome_arquivo
          FROM financeiro3_om_itens i JOIN financeiro3_oms o ON o.id=i.om_id
          JOIN financeiro3_categorias c ON c.id=i.categoria_id
          JOIN financeiro3_centros_custo cc ON cc.id=i.centro_custo_id
          JOIN financeiro3_moedas m ON m.id=o.moeda_id
          LEFT JOIN financeiro3_anexos a ON a.entidade='OM_ITEM' AND a.entidade_id=i.id AND a.status='ATIVO'
          LEFT JOIN financeiro3_arquivos ar ON ar.id=a.arquivo_id AND ar.status='ATIVO'
          WHERE i.data_despesa=:data AND i.valor=:valor AND i.status='ATIVO'
            AND o.removido_em IS NULL
          UNION ALL
          SELECT 'RD',r.id,i.id,r.numero_rd,i.data_despesa,i.descricao,c.nome,
            cc.codigo||' · '||cc.nome,i.valor,m.codigo,NULL::uuid,NULL::varchar
          FROM financeiro3_rd_itens i
          JOIN financeiro3_rds r ON r.id=i.rd_id
          JOIN financeiro3_categorias c ON c.id=i.categoria_id
          JOIN financeiro3_centros_custo cc ON cc.id=r.centro_custo_id
          JOIN financeiro3_moedas m ON m.id=r.moeda_id
          WHERE i.data_despesa=:data AND i.valor=:valor AND i.status='ATIVO'
          UNION ALL
          SELECT 'REEMBOLSO',x.id,i.id,'REEMBOLSO-'||LPAD(x.id::text,6,'0'),
            i.data_despesa,i.descricao,c.nome,cc.codigo||' · '||cc.nome,i.valor,m.codigo,
            ar.id,ar.nome_original
          FROM financeiro3_reembolso_itens i
          JOIN financeiro3_reembolsos x ON x.id=i.reembolso_id
          JOIN financeiro3_categorias c ON c.id=i.categoria_id
          JOIN financeiro3_centros_custo cc ON cc.id=x.centro_custo_id
          JOIN financeiro3_moedas m ON m.id=x.moeda_id
          LEFT JOIN financeiro3_anexos a ON a.entidade='REEMBOLSO_ITEM' AND a.entidade_id=i.id AND a.status='ATIVO'
          LEFT JOIN financeiro3_arquivos ar ON ar.id=a.arquivo_id AND ar.status='ATIVO'
          WHERE i.data_despesa=:data AND i.valor=:valor AND i.status='ATIVO'
        ) duplicadas ORDER BY tipo,numero,item_id
    """), {"data": data, "valor": valor}).mappings().all()
    resultado = []
    for registro in registros:
        item = {
            "tipo": registro["tipo"], "numero": registro["numero"],
            "data": registro["data"].isoformat(), "descricao": registro["descricao"],
            "categoria": registro["categoria"], "centro": registro["centro"],
            "valor": str(registro["valor"]), "moeda": registro["moeda"],
            "nome_arquivo": registro["nome_arquivo"],
        }
        if registro["tipo"] == "OM":
            item["documento_url"] = url_for("financeiro_novo.om_detalhe", om_id=registro["origem_id"])
            if registro["arquivo_id"]:
                item["comprovante_url"] = url_for("financeiro_novo.om_item_anexo_baixar",
                    om_id=registro["origem_id"], item_id=registro["item_id"], arquivo_id=registro["arquivo_id"])
        elif registro["tipo"] == "RD":
            item["documento_url"] = url_for("financeiro_novo.rd_detalhe", rd_id=registro["origem_id"])
        else:
            item["documento_url"] = url_for("financeiro_novo.reembolso_detalhe", reembolso_id=registro["origem_id"])
            if registro["arquivo_id"]:
                item["comprovante_url"] = url_for("financeiro_novo.reembolso_anexo_baixar",
                    reembolso_id=registro["origem_id"], arquivo_id=registro["arquivo_id"])
        resultado.append(item)
    return resultado


def _linhas_om_formulario():
    datas = request.form.getlist("data_despesa")
    centros = request.form.getlist("centro_custo_id")
    categorias = request.form.getlist("categoria_id")
    descricoes = request.form.getlist("descricao")
    valores = request.form.getlist("valor")
    justificativas = request.form.getlist("justificativa_sem_comprovante")
    arquivos = request.files.getlist("arquivo")
    quantidade = len(datas)
    if not quantidade or any(len(lista) != quantidade for lista in (
        centros, categorias, descricoes, valores, justificativas, arquivos,
    )):
        raise ValorInvalido("As linhas da OM estão incompletas. Revise e tente novamente.")
    linhas = []
    for indice in range(quantidade):
        try:
            centro = int(centros[indice] or 0)
            categoria = int(categorias[indice] or 0)
        except ValueError as exc:
            raise ValorInvalido(f"Linha {indice + 1}: centro ou categoria inválido.") from exc
        descricao = descricoes[indice].strip()
        justificativa = justificativas[indice].strip() or None
        arquivo = arquivos[indice]
        if not descricao or len(descricao) > 220:
            raise ValorInvalido(f"Linha {indice + 1}: informe uma descrição com até 220 caracteres.")
        if (not arquivo or not arquivo.filename) and not justificativa:
            raise ValorInvalido(f"Linha {indice + 1}: anexe o recibo ou justifique sua ausência.")
        linhas.append({
            "numero": indice + 1,
            "data": data_iso(datas[indice], f"Data da linha {indice + 1}"),
            "centro": centro,
            "categoria": categoria,
            "descricao": descricao,
            "valor": decimal_br(valores[indice], positivo=True),
            "justificativa": justificativa,
            "arquivo": arquivo,
        })
    return linhas


COLUNAS_EXCEL_OM = ("data", "centro de custo", "categoria", "descricao", "valor")
LIMITE_LINHAS_EXCEL_OM = 1000
LIMITE_BYTES_EXCEL_OM = 10 * 1024 * 1024


def _chave_excel(valor):
    texto = unicodedata.normalize("NFKD", str(valor or "").strip())
    texto = "".join(caractere for caractere in texto if not unicodedata.combining(caractere))
    return re.sub(r"[^a-z0-9]+", " ", texto.casefold()).strip()


def _indice_cadastro_excel(registros):
    indice, ambiguos = {}, set()
    for registro in registros:
        rotulos = {
            _chave_excel(registro["codigo"]),
            _chave_excel(registro["nome"]),
            _chave_excel(f"{registro['codigo']} {registro['nome']}"),
        }
        codigo = str(registro["codigo"] or "").strip()
        if codigo.isdigit():
            rotulos.add(str(int(codigo)))
        for rotulo in rotulos - {""}:
            if rotulo in indice and indice[rotulo] != registro["id"]:
                ambiguos.add(rotulo)
            else:
                indice[rotulo] = registro["id"]
    for rotulo in ambiguos:
        indice.pop(rotulo, None)
    return indice


def _data_excel(valor, linha):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto_data = str(valor or "").strip()
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto_data, formato).date()
        except ValueError:
            continue
    raise ValorInvalido(f"Linha {linha}: Data inválida. Use DD/MM/AAAA.")


def _ler_linhas_om_excel(conteudo, centros, categorias):
    try:
        planilha = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError, ValueError, KeyError) as exc:
        raise ValorInvalido("O arquivo não é um Excel .xlsx válido.") from exc
    try:
        aba = planilha.active
        cabecalho = next(aba.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not cabecalho:
            raise ValorInvalido("O Excel está vazio.")
        posicoes = {}
        for indice, valor in enumerate(cabecalho):
            chave = _chave_excel(valor)
            if chave and chave not in posicoes:
                posicoes[chave] = indice
        ausentes = [coluna for coluna in COLUNAS_EXCEL_OM if coluna not in posicoes]
        if ausentes:
            nomes = ", ".join(coluna.title() for coluna in ausentes)
            raise ValorInvalido(f"Colunas obrigatórias não encontradas: {nomes}.")

        centros_por_nome = _indice_cadastro_excel(centros)
        categorias_por_nome = _indice_cadastro_excel(categorias)
        linhas, erros, linhas_lidas = [], [], 0
        for numero_excel, valores in enumerate(aba.iter_rows(min_row=2, values_only=True), 2):
            dados = {coluna: valores[posicoes[coluna]] if posicoes[coluna] < len(valores) else None
                     for coluna in COLUNAS_EXCEL_OM}
            if all(valor is None or str(valor).strip() == "" for valor in dados.values()):
                continue
            linhas_lidas += 1
            if linhas_lidas > LIMITE_LINHAS_EXCEL_OM:
                raise ValorInvalido(f"O Excel excede o limite de {LIMITE_LINHAS_EXCEL_OM} linhas.")
            try:
                centro = centros_por_nome.get(_chave_excel(dados["centro de custo"]))
                categoria = categorias_por_nome.get(_chave_excel(dados["categoria"]))
                if not centro:
                    raise ValorInvalido(f"Centro de custo '{dados['centro de custo']}' não encontrado ou ambíguo.")
                if not categoria:
                    raise ValorInvalido(f"Categoria '{dados['categoria']}' não encontrada ou ambígua.")
                descricao = str(dados["descricao"] or "").strip()
                if not descricao or len(descricao) > 220:
                    raise ValorInvalido("Descrição obrigatória com no máximo 220 caracteres.")
                valor = decimal_br(dados["valor"], positivo=True)
                linhas.append({
                    "data": _data_excel(dados["data"], numero_excel).isoformat(),
                    "centro_custo_id": centro,
                    "categoria_id": categoria,
                    "descricao": descricao,
                    "valor": f"{valor:.2f}".replace(".", ","),
                })
            except ValorInvalido as exc:
                erros.append(f"Linha {numero_excel}: {str(exc).removeprefix(f'Linha {numero_excel}: ')}")
        if erros:
            resumo = " ".join(erros[:20])
            if len(erros) > 20:
                resumo += f" Mais {len(erros) - 20} erro(s)."
            raise ValorInvalido(resumo)
        if not linhas:
            raise ValorInvalido("O Excel não possui linhas de despesas preenchidas.")
        return linhas
    finally:
        planilha.close()


@bp.post("/oms/<int:om_id>/itens/carregar-excel")
@login_required
@permission_required("financeiro_novo", "editar")
def om_itens_carregar_excel(om_id):
    arquivo = request.files.get("arquivo_excel")
    if not arquivo or not arquivo.filename:
        return jsonify({"erro": "Selecione um arquivo Excel .xlsx."}), 400
    if not arquivo.filename.lower().endswith(".xlsx"):
        return jsonify({"erro": "Formato inválido. Envie um arquivo .xlsx."}), 400
    conteudo = arquivo.read(LIMITE_BYTES_EXCEL_OM + 1)
    if len(conteudo) > LIMITE_BYTES_EXCEL_OM:
        return jsonify({"erro": "O Excel excede o limite de 10 MB."}), 413
    try:
        with get_engine().connect() as conn:
            om = _registro(conn, "financeiro3_oms", om_id)
            if not om:
                abort(404)
            if om["status"] not in EDITAVEIS:
                return jsonify({"erro": "Esta OM não aceita novas linhas."}), 409
            centros = conn.execute(text("""
                SELECT id,codigo,nome FROM financeiro3_centros_custo WHERE ativo ORDER BY codigo,nome
            """)).mappings().all()
            categorias = conn.execute(text("""
                SELECT id,codigo,nome FROM financeiro3_categorias
                WHERE ativo AND natureza='DESPESA' ORDER BY codigo,nome
            """)).mappings().all()
        linhas = _ler_linhas_om_excel(conteudo, centros, categorias)
        return jsonify({"linhas": linhas, "quantidade": len(linhas)})
    except ValorInvalido as exc:
        return jsonify({"erro": str(exc)}), 422


@bp.get("/oms/<int:om_id>/verificar-duplicidades")
@login_required
@permission_required("financeiro_novo", "editar")
def om_verificar_duplicidades(om_id):
    datas = request.args.getlist("data")
    valores = request.args.getlist("valor")
    linhas = [{"data": data, "valor": valor} for data, valor in zip(datas, valores)]
    resultado, vistos = [], {}
    with get_engine().connect() as conn:
        if not _registro(conn, "financeiro3_oms", om_id):
            abort(404)
        for indice, linha in enumerate(linhas, 1):
            try:
                data = data_iso(linha.get("data"), f"Data da linha {indice}")
                valor = decimal_br(linha.get("valor"), positivo=True)
            except (ValorInvalido, AttributeError):
                continue
            chave = (data, valor)
            registros = _origens_duplicadas(conn, data, valor)
            if chave in vistos:
                registros.append({
                    "tipo": "LOTE", "numero": f"Linha {vistos[chave]}",
                    "data": data.isoformat(), "descricao": "Outra linha deste lote",
                    "categoria": "—", "centro": "—", "valor": str(valor), "moeda": None,
                    "nome_arquivo": None,
                })
            else:
                vistos[chave] = indice
            if registros:
                resultado.append({"linha": indice, "registros": registros})
    return jsonify({"duplicidades": resultado})


@bp.post("/oms/<int:om_id>/itens")
@login_required
@permission_required("financeiro_novo", "editar")
def om_item_novo(om_id):
    from routes.financeiro_novo.reembolsos import _preparar_anexo, _vincular_anexo
    preparados = []
    try:
        linhas = _linhas_om_formulario()
        preparados = [_preparar_anexo(linha["arquivo"]) for linha in linhas]
        with get_engine().begin() as conn:
            om = _registro(conn, "financeiro3_oms", om_id, True)
            if not om: abort(404)
            if om["status"] not in EDITAVEIS: abort(409)
            duplicadas, vistos = [], {}
            for linha in linhas:
                refs = conn.execute(text("""
                    SELECT EXISTS(SELECT 1 FROM financeiro3_categorias
                        WHERE id=:categoria AND ativo AND natureza='DESPESA')
                      AND EXISTS(SELECT 1 FROM financeiro3_centros_custo
                        WHERE id=:centro AND ativo)
                """), linha).scalar()
                if not refs:
                    raise ValorInvalido(f"Linha {linha['numero']}: centro ou categoria inválido.")
                chave = (linha["data"], linha["valor"])
                registros = _origens_duplicadas(conn, *chave)
                if chave in vistos:
                    registros.append({"tipo": "LOTE", "numero": f"linha {vistos[chave]} deste lote"})
                else:
                    vistos[chave] = linha["numero"]
                if registros:
                    rotulos = [f"{item['tipo']} {item['numero']}" for item in registros]
                    duplicadas.append(f"linha {linha['numero']} ({', '.join(rotulos)})")
            if duplicadas and request.form.get("forcar_salvamento") != "1":
                raise ValorInvalido("Possíveis duplicidades: " + "; ".join(duplicadas) + ". Confirme para salvar.")
            for linha, preparado in zip(linhas, preparados):
                item = conn.execute(text("""
                    INSERT INTO financeiro3_om_itens(om_id,data_despesa,centro_custo_id,
                      categoria_id,descricao,valor,justificativa_sem_comprovante,criado_por)
                    VALUES (:om,:data,:centro,:categoria,:descricao,:valor,:justificativa,:usuario)
                    RETURNING *
                """), {**linha, "om": om_id, "usuario": session.get("usuario_id")}).mappings().one()
                vinculo = _vincular_anexo(conn, preparado, "OM_ITEM", item["id"], "COMPROVANTE")
                registrar_evento(conn, entidade="OM_ITEM", entidade_id=item["id"], evento="CRIADO",
                    dados_novos={**dict(item), "anexo_id": vinculo})
        flash(f"{len(linhas)} linha(s) de despesa incluída(s) na OM.", "sucesso")
    except (ValorInvalido, ValueError, AnexoInvalido) as exc:
        for preparado in preparados:
            if preparado: preparado[3].unlink(missing_ok=True)
        flash(str(exc), "erro")
    except Exception:
        for preparado in preparados:
            if preparado: preparado[3].unlink(missing_ok=True)
        raise
    return redirect(url_for("financeiro_novo.om_detalhe", om_id=om_id))


@bp.post("/oms/<int:om_id>/itens/<int:item_id>/remover")
@login_required
@permission_required("financeiro_novo", "editar")
def om_item_remover(om_id, item_id):
    with get_engine().begin() as conn:
        om = _registro(conn, "financeiro3_oms", om_id, True)
        if not om or om["status"] not in EDITAVEIS: abort(409 if om else 404)
        anterior = conn.execute(text("""
            SELECT * FROM financeiro3_om_itens
            WHERE id=:item AND om_id=:om AND status='ATIVO' FOR UPDATE
        """), {"item": item_id, "om": om_id}).mappings().first()
        if not anterior: abort(404)
        novo = conn.execute(text("""
            UPDATE financeiro3_om_itens SET status='REMOVIDO',removido_por=:u,removido_em=NOW()
            WHERE id=:item RETURNING *
        """), {"u": session.get("usuario_id"), "item": item_id}).mappings().one()
        registrar_evento(conn, entidade="OM_ITEM", entidade_id=item_id, evento="REMOVIDO",
            dados_anteriores=dict(anterior), dados_novos=dict(novo))
    flash("Linha removida e preservada na auditoria.", "sucesso")
    return redirect(url_for("financeiro_novo.om_detalhe", om_id=om_id))


@bp.post("/oms/<int:om_id>/itens/<int:item_id>/recibo")
@login_required
@permission_required("financeiro_novo", "editar")
def om_item_recibo_substituir(om_id, item_id):
    from routes.financeiro_novo.reembolsos import _preparar_anexo, _vincular_anexo
    preparado = None
    try:
        preparado = _preparar_anexo(request.files.get("arquivo"))
        if not preparado:
            raise ValorInvalido("Selecione uma foto ou PDF do recibo.")
        with get_engine().begin() as conn:
            om = _registro(conn, "financeiro3_oms", om_id, True)
            if not om:
                abort(404)
            if om["status"] not in EDITAVEIS:
                abort(409)
            item = conn.execute(text("""
                SELECT * FROM financeiro3_om_itens
                WHERE id=:item AND om_id=:om AND status='ATIVO' FOR UPDATE
            """), {"item": item_id, "om": om_id}).mappings().first()
            if not item:
                abort(404)
            anteriores = conn.execute(text("""
                UPDATE financeiro3_anexos SET status='REMOVIDO',removido_por=:usuario,
                  removido_em=NOW(),motivo_remocao='Recibo substituído pelo usuário'
                WHERE entidade='OM_ITEM' AND entidade_id=:item AND status='ATIVO'
                RETURNING id,arquivo_id
            """), {"usuario": session.get("usuario_id"), "item": item_id}).mappings().all()
            vinculo = _vincular_anexo(conn, preparado, "OM_ITEM", item_id, "COMPROVANTE")
            registrar_evento(conn, entidade="OM_ITEM", entidade_id=item_id, evento="RECIBO_SUBSTITUIDO",
                dados_anteriores={"anexos": [dict(anexo) for anexo in anteriores]},
                dados_novos={"anexo_id": vinculo})
        flash("Recibo convertido para PDF e vinculado novamente.", "sucesso")
    except (ValorInvalido, AnexoInvalido) as exc:
        if preparado:
            preparado[3].unlink(missing_ok=True)
        flash(str(exc), "erro")
    except Exception:
        if preparado:
            preparado[3].unlink(missing_ok=True)
        raise
    return redirect(url_for("financeiro_novo.om_detalhe", om_id=om_id))


def _pagamento_config(tipo):
    if tipo == "om":
        return "financeiro3_oms", "financeiro3_om_pagamentos", "om_id", "OM", "om_detalhe", "om_id"
    return "financeiro3_rds", "financeiro3_rd_pagamentos", "rd_id", "RD", "rd_detalhe", "rd_id"


def _redirecionar_documento(tipo, registro_id):
    _, _, _, _, endpoint, parametro = _pagamento_config(tipo)
    return redirect(url_for(f"financeiro_novo.{endpoint}", **{parametro: registro_id}))


def _programar_pagamento(tipo, registro_id):
    tabela_documento, tabela, fk, entidade, _, _ = _pagamento_config(tipo)
    try:
        tipo_pagamento = (request.form.get("tipo_pagamento") or "").upper()
        if tipo_pagamento not in {"ADIANTAMENTO", "QUITACAO"}:
            raise ValorInvalido("Tipo de pagamento inválido.")
        dados = {
            "documento": registro_id,
            "tipo": tipo_pagamento,
            "data": data_iso(request.form.get("data_prevista_pagamento"), "Data prevista"),
            "valor": decimal_br(request.form.get("valor"), positivo=True),
            "observacoes": (request.form.get("observacoes") or "").strip() or None,
            "usuario": session.get("usuario_id"),
        }
        with get_engine().begin() as conn:
            documento = _registro(conn, tabela_documento, registro_id, True)
            if not documento:
                abort(404)
            if documento["status"] in {"CANCELADA", "ENCERRADA", "LIQUIDADA"}:
                abort(409)
            pagamento = conn.execute(text(f"""
                INSERT INTO {tabela}({fk},tipo,data_prevista_pagamento,valor,observacoes,criado_por)
                VALUES (:documento,:tipo,:data,:valor,:observacoes,:usuario) RETURNING *
            """), dados).mappings().one()
            registrar_evento(conn, entidade=f"{entidade}_PAGAMENTO", entidade_id=pagamento["id"],
                evento="PROGRAMADO", dados_novos=dict(pagamento))
        flash("Pagamento incluído na previsão.", "sucesso")
    except ValorInvalido as exc:
        flash(str(exc), "erro")
    return _redirecionar_documento(tipo, registro_id)


def _registrar_pagamento_om_direto(om_id):
    from routes.financeiro_novo.reembolsos import _preparar_anexo, _vincular_anexo
    preparado = None
    try:
        tipo_pagamento = (request.form.get("tipo_pagamento") or "").upper()
        if tipo_pagamento not in {"ADIANTAMENTO", "QUITACAO"}:
            raise ValorInvalido("Tipo de pagamento inválido.")
        data_pagamento = data_iso(request.form.get("data_pagamento"), "Data do pagamento")
        valor = decimal_br(request.form.get("valor"), positivo=True)
        observacoes = (request.form.get("observacoes") or "").strip() or None
        preparado = _preparar_anexo(request.files.get("arquivo"))
        with get_engine().begin() as conn:
            om = _registro(conn, "financeiro3_oms", om_id, True)
            if not om:
                abort(404)
            if om["status"] in {"CANCELADA", "ENCERRADA"}:
                abort(409)
            pagamento = conn.execute(text("""
                INSERT INTO financeiro3_om_pagamentos(om_id,tipo,data_prevista_pagamento,
                  data_pagamento,valor,status,observacoes,criado_por,pago_por,pago_em)
                VALUES (:om,:tipo,:data,:data,:valor,'PAGO',:observacoes,:usuario,:usuario,NOW())
                RETURNING *
            """), {"om": om_id, "tipo": tipo_pagamento, "data": data_pagamento,
                    "valor": valor, "observacoes": observacoes,
                    "usuario": session.get("usuario_id")}).mappings().one()
            vinculo = _vincular_anexo(conn, preparado, "OM_PAGAMENTO", pagamento["id"], "COMPROVANTE")
            registrar_evento(conn, entidade="OM_PAGAMENTO", entidade_id=pagamento["id"],
                evento="REGISTRADO", dados_novos={**dict(pagamento), "anexo_id": vinculo})
        flash("Adiantamento registrado." if tipo_pagamento == "ADIANTAMENTO" else "Quitação registrada.", "sucesso")
    except (ValorInvalido, AnexoInvalido) as exc:
        if preparado:
            preparado[3].unlink(missing_ok=True)
        flash(str(exc), "erro")
    except Exception:
        if preparado:
            preparado[3].unlink(missing_ok=True)
        raise
    return redirect(url_for("financeiro_novo.om_detalhe", om_id=om_id))


def _registrar_pagamento(tipo, registro_id, pagamento_id):
    from routes.financeiro_novo.reembolsos import _preparar_anexo, _vincular_anexo
    tabela_documento, tabela, fk, entidade, _, _ = _pagamento_config(tipo)
    preparado = None
    try:
        data_pagamento = data_iso(request.form.get("data_pagamento"), "Data do pagamento")
        preparado = _preparar_anexo(request.files.get("arquivo"))
        with get_engine().begin() as conn:
            if not _registro(conn, tabela_documento, registro_id, True):
                abort(404)
            anterior = conn.execute(text(f"""
                SELECT * FROM {tabela} WHERE id=:pagamento AND {fk}=:documento FOR UPDATE
            """), {"pagamento": pagamento_id, "documento": registro_id}).mappings().first()
            if not anterior:
                abort(404)
            if anterior["status"] != "PREVISTO":
                abort(409)
            novo = conn.execute(text(f"""
                UPDATE {tabela} SET status='PAGO',data_pagamento=:data,pago_por=:usuario,pago_em=NOW()
                WHERE id=:id RETURNING *
            """), {"data": data_pagamento, "usuario": session.get("usuario_id"), "id": pagamento_id}).mappings().one()
            vinculo = _vincular_anexo(conn, preparado, f"{entidade}_PAGAMENTO", pagamento_id, "COMPROVANTE")
            registrar_evento(conn, entidade=f"{entidade}_PAGAMENTO", entidade_id=pagamento_id,
                evento="PAGO", dados_anteriores=dict(anterior), dados_novos={**dict(novo), "anexo_id": vinculo})
        flash("Pagamento realizado e movido da previsão para o realizado.", "sucesso")
    except (ValorInvalido, AnexoInvalido) as exc:
        if preparado:
            preparado[3].unlink(missing_ok=True)
        flash(str(exc), "erro")
    except Exception:
        if preparado:
            preparado[3].unlink(missing_ok=True)
        raise
    return _redirecionar_documento(tipo, registro_id)


def _cancelar_pagamento(tipo, registro_id, pagamento_id):
    _, tabela, fk, entidade, _, _ = _pagamento_config(tipo)
    motivo = (request.form.get("motivo") or "").strip()
    if not motivo:
        flash("Informe o motivo do cancelamento.", "erro")
        return _redirecionar_documento(tipo, registro_id)
    with get_engine().begin() as conn:
        anterior = conn.execute(text(f"SELECT * FROM {tabela} WHERE id=:id AND {fk}=:doc FOR UPDATE"),
            {"id": pagamento_id, "doc": registro_id}).mappings().first()
        if not anterior:
            abort(404)
        if anterior["status"] != "PREVISTO":
            abort(409)
        novo = conn.execute(text(f"""
            UPDATE {tabela} SET status='CANCELADO',cancelado_por=:usuario,cancelado_em=NOW(),motivo_cancelamento=:motivo
            WHERE id=:id RETURNING *
        """), {"usuario": session.get("usuario_id"), "motivo": motivo, "id": pagamento_id}).mappings().one()
        registrar_evento(conn, entidade=f"{entidade}_PAGAMENTO", entidade_id=pagamento_id,
            evento="CANCELADO", dados_anteriores=dict(anterior), dados_novos=dict(novo), justificativa=motivo)
    flash("Previsão de pagamento cancelada.", "sucesso")
    return _redirecionar_documento(tipo, registro_id)


@bp.post("/oms/<int:om_id>/pagamentos")
@login_required
@permission_required("financeiro_novo", "pagar")
def om_pagamento_programar(om_id):
    return _registrar_pagamento_om_direto(om_id)


@bp.post("/oms/<int:om_id>/pagamentos/<int:pagamento_id>/pagar")
@login_required
@permission_required("financeiro_novo", "pagar")
def om_pagamento_pagar(om_id, pagamento_id):
    return _registrar_pagamento("om", om_id, pagamento_id)


@bp.post("/oms/<int:om_id>/pagamentos/<int:pagamento_id>/cancelar")
@login_required
@permission_required("financeiro_novo", "cancelar")
def om_pagamento_cancelar(om_id, pagamento_id):
    return _cancelar_pagamento("om", om_id, pagamento_id)


@bp.post("/rds/<int:rd_id>/pagamentos")
@login_required
@permission_required("financeiro_novo", "editar")
def rd_pagamento_programar(rd_id):
    return _programar_pagamento("rd", rd_id)


@bp.post("/rds/<int:rd_id>/pagamentos/<int:pagamento_id>/pagar")
@login_required
@permission_required("financeiro_novo", "pagar")
def rd_pagamento_pagar(rd_id, pagamento_id):
    return _registrar_pagamento("rd", rd_id, pagamento_id)


@bp.post("/rds/<int:rd_id>/pagamentos/<int:pagamento_id>/cancelar")
@login_required
@permission_required("financeiro_novo", "cancelar")
def rd_pagamento_cancelar(rd_id, pagamento_id):
    return _cancelar_pagamento("rd", rd_id, pagamento_id)


@bp.post("/oms/<int:om_id>/enviar")
@login_required
@permission_required("financeiro_novo", "editar")
def om_enviar(om_id):
    with get_engine().begin() as conn:
        anterior = _registro(conn, "financeiro3_oms", om_id, True)
        if not anterior:
            abort(404)
        if anterior["status"] not in EDITAVEIS:
            abort(409)
        if anterior["valor_total"] <= 0:
            flash("Inclua ao menos uma linha de despesa antes de enviar a OM.", "erro")
            return redirect(url_for("financeiro_novo.om_detalhe", om_id=om_id))
        novo = conn.execute(text("UPDATE financeiro3_oms SET status='EM_APROVACAO',atualizado_em=NOW() WHERE id=:id RETURNING *"), {"id": om_id}).mappings().one()
        _decisao(conn, "om", om_id, "ENVIO", anterior["status"], "EM_APROVACAO")
        registrar_evento(conn, entidade="OM", entidade_id=om_id, evento="ENVIADA_APROVACAO", dados_anteriores=dict(anterior), dados_novos=dict(novo))
    flash("OM enviada para aprovação.", "sucesso")
    return redirect(url_for("financeiro_novo.om_detalhe", om_id=om_id))


def _decisao(conn, tipo, registro_id, acao, anterior, novo, justificativa=None):
    fk = "om_id" if tipo == "om" else "rd_id"
    tabela = "financeiro3_om_decisoes" if tipo == "om" else "financeiro3_rd_decisoes"
    conn.execute(text(f"""
        INSERT INTO {tabela}({fk},acao,status_anterior,status_novo,justificativa,usuario_id)
        VALUES (:id,:acao,:anterior,:novo,:justificativa,:usuario)
    """), {"id": registro_id, "acao": acao, "anterior": anterior, "novo": novo,
            "justificativa": justificativa, "usuario": session.get("usuario_id")})


def _decidir_om(om_id, aprovar):
    justificativa = (request.form.get("justificativa") or "").strip()
    if not aprovar and not justificativa:
        flash("Informe a justificativa da rejeição.", "erro")
        return
    with get_engine().begin() as conn:
        anterior = _registro(conn, "financeiro3_oms", om_id, True)
        if not anterior:
            abort(404)
        if anterior["status"] != "EM_APROVACAO":
            abort(409)
        if _proibir_autoaprovacao(anterior):
            flash("O responsável pelo lançamento não pode decidir a própria OM.", "erro")
            return
        status = "APROVADA" if aprovar else "REJEITADA"
        novo = conn.execute(text("UPDATE financeiro3_oms SET status=:status,atualizado_em=NOW() WHERE id=:id RETURNING *"), {"status": status, "id": om_id}).mappings().one()
        acao = "APROVACAO" if aprovar else "REJEICAO"
        _decisao(conn, "om", om_id, acao, "EM_APROVACAO", status, justificativa or None)
        registrar_evento(conn, entidade="OM", entidade_id=om_id, evento=acao, dados_anteriores=dict(anterior), dados_novos=dict(novo), justificativa=justificativa)
    flash("OM aprovada." if aprovar else "OM rejeitada para correção.", "sucesso")


@bp.post("/oms/<int:om_id>/aprovar")
@login_required
@permission_required("financeiro_novo", "aprovar")
def om_aprovar(om_id):
    _decidir_om(om_id, True)
    return redirect(url_for("financeiro_novo.om_detalhe", om_id=om_id))


@bp.post("/oms/<int:om_id>/rejeitar")
@login_required
@permission_required("financeiro_novo", "aprovar")
def om_rejeitar(om_id):
    _decidir_om(om_id, False)
    return redirect(url_for("financeiro_novo.om_detalhe", om_id=om_id))


@bp.post("/oms/<int:om_id>/cancelar")
@login_required
@permission_required("financeiro_novo", "cancelar")
def om_cancelar(om_id):
    motivo = (request.form.get("motivo") or "").strip()
    if not motivo:
        flash("Informe o motivo do cancelamento.", "erro")
        return redirect(url_for("financeiro_novo.om_detalhe", om_id=om_id))
    with get_engine().begin() as conn:
        anterior = _registro(conn, "financeiro3_oms", om_id, True)
        if not anterior:
            abort(404)
        if anterior["status"] in {"ENCERRADA", "CANCELADA"}:
            abort(409)
        if conn.execute(text("""
            SELECT EXISTS(SELECT 1 FROM financeiro3_om_pagamentos
              WHERE om_id=:id AND status='PAGO')
        """), {"id": om_id}).scalar():
            flash("A OM possui pagamento realizado e não pode ser cancelada.", "erro")
            return redirect(url_for("financeiro_novo.om_detalhe", om_id=om_id))
        conn.execute(text("""
            UPDATE financeiro3_om_pagamentos SET status='CANCELADO',cancelado_por=:usuario,
              cancelado_em=NOW(),motivo_cancelamento=:motivo
            WHERE om_id=:id AND status='PREVISTO'
        """), {"usuario": session.get("usuario_id"), "motivo": motivo, "id": om_id})
        novo = conn.execute(text("UPDATE financeiro3_oms SET status='CANCELADA',atualizado_em=NOW() WHERE id=:id RETURNING *"), {"id": om_id}).mappings().one()
        _decisao(conn, "om", om_id, "CANCELAMENTO", anterior["status"], "CANCELADA", motivo)
        registrar_evento(conn, entidade="OM", entidade_id=om_id, evento="CANCELADA", dados_anteriores=dict(anterior), dados_novos=dict(novo), justificativa=motivo)
    flash("OM cancelada.", "sucesso")
    return redirect(url_for("financeiro_novo.om_detalhe", om_id=om_id))


def _dados_rd(form):
    try:
        dados = {
            "responsavel": int(form.get("responsavel_id") or 0),
            "centro": int(form.get("centro_custo_id") or 0),
            "moeda": int(form.get("moeda_id") or 0),
        }
    except ValueError as exc:
        raise ValorInvalido("Selecione responsável, centro e moeda válidos.") from exc
    dados.update({
        "numero": (form.get("numero_rd") or "").strip().upper(),
        "matricula": (form.get("matricula_responsavel") or "").strip().upper(),
        "inicio": data_iso(form.get("periodo_inicio"), "Início do período"),
        "fim": data_iso(form.get("periodo_fim"), "Fim do período"),
        "observacoes": (form.get("observacoes") or "").strip() or None,
    })
    if not dados["numero"] or not dados["matricula"]:
        raise ValorInvalido("Informe o número da RD e a matrícula do responsável.")
    if len(dados["numero"]) > 80 or len(dados["matricula"]) > 40:
        raise ValorInvalido("Número da RD ou matrícula excede o tamanho permitido.")
    if dados["fim"] < dados["inicio"]:
        raise ValorInvalido("O fim do período não pode ser anterior ao início.")
    return dados


@bp.route("/rds/nova", methods=["GET", "POST"])
@login_required
@permission_required("financeiro_novo", "criar")
def rd_nova():
    with get_engine().connect() as conn:
        opcoes = _opcoes(conn)
    if request.method == "POST":
        try:
            dados = _dados_rd(request.form)
            dados["usuario"] = session.get("usuario_id")
            with get_engine().begin() as conn:
                refs = conn.execute(text("""
                    SELECT EXISTS(SELECT 1 FROM financeiro3_pessoas
                        WHERE id=:responsavel AND ativo AND favorecido)
                      AND EXISTS(SELECT 1 FROM financeiro3_centros_custo WHERE id=:centro AND ativo)
                      AND EXISTS(SELECT 1 FROM financeiro3_moedas WHERE id=:moeda AND ativo)
                """), dados).scalar()
                if not refs:
                    raise ValorInvalido("Responsável, centro ou moeda está inativo ou inválido.")
                rd = conn.execute(text("""
                    INSERT INTO financeiro3_rds(numero_rd,matricula_responsavel,responsavel_id,
                      centro_custo_id,moeda_id,periodo_inicio,periodo_fim,valor_adiantamento,
                      observacoes,criado_por)
                    VALUES (:numero,:matricula,:responsavel,:centro,:moeda,:inicio,:fim,
                      0,:observacoes,:usuario) RETURNING *
                """), dados).mappings().one()
                registrar_evento(conn, entidade="RD", entidade_id=rd["id"], evento="CRIADA",
                    dados_novos=dict(rd))
            flash("RD criada em rascunho.", "sucesso")
            return redirect(url_for("financeiro_novo.rd_detalhe", rd_id=rd["id"]))
        except (ValorInvalido, IntegrityError) as exc:
            if isinstance(exc, IntegrityError):
                exc = ValorInvalido("Já existe uma RD com este número.")
            flash(str(exc), "erro")
    return render_template("financeiro_novo/rd_form.html", opcoes=opcoes,
        subnav_links=build_subnav("missoes"))


@bp.get("/rds/<int:rd_id>")
@login_required
@permission_required("financeiro_novo", "visualizar")
def rd_detalhe(rd_id):
    with get_engine().connect() as conn:
        rd = conn.execute(text("""
            SELECT r.*, p.nome_razao AS responsavel, cc.codigo AS centro_codigo,
                   cc.nome AS centro_nome, m.codigo AS moeda,
              COALESCE((SELECT SUM(i.valor) FROM financeiro3_reembolso_itens i
                JOIN financeiro3_reembolsos rb ON rb.id=i.reembolso_id
                WHERE rb.rd_pagadora_id=r.id AND rb.forma_liquidacao='RD'
                  AND rb.status='APROVADO' AND i.status='ATIVO'),0) AS valor_reembolsos
            FROM financeiro3_rds r JOIN financeiro3_pessoas p ON p.id=r.responsavel_id
            JOIN financeiro3_centros_custo cc ON cc.id=r.centro_custo_id
            JOIN financeiro3_moedas m ON m.id=r.moeda_id
            WHERE r.id=:id
        """), {"id": rd_id}).mappings().first()
        if not rd:
            abort(404)
        itens = conn.execute(text("""
            SELECT i.*, c.nome AS categoria, p.nome_razao AS fornecedor
            FROM financeiro3_rd_itens i JOIN financeiro3_categorias c ON c.id=i.categoria_id
            LEFT JOIN financeiro3_pessoas p ON p.id=i.fornecedor_id
            WHERE i.rd_id=:id AND i.status='ATIVO' ORDER BY i.data_despesa,i.id
        """), {"id": rd_id}).mappings().all()
        anexos = _listar_anexos(conn, "RD", rd_id)
        decisoes = conn.execute(text("SELECT * FROM financeiro3_rd_decisoes WHERE rd_id=:id ORDER BY id DESC"), {"id": rd_id}).mappings().all()
        acerto = conn.execute(text("SELECT * FROM financeiro3_rd_acertos WHERE rd_id=:id"), {"id": rd_id}).mappings().first()
        pagamentos = conn.execute(text("""
            SELECT pg.*,a.id AS anexo_id,ar.id AS arquivo_id,ar.nome_original
            FROM financeiro3_rd_pagamentos pg
            LEFT JOIN financeiro3_anexos a ON a.entidade='RD_PAGAMENTO' AND a.entidade_id=pg.id AND a.status='ATIVO'
            LEFT JOIN financeiro3_arquivos ar ON ar.id=a.arquivo_id AND ar.status='ATIVO'
            WHERE pg.rd_id=:id ORDER BY pg.data_prevista_pagamento,pg.id
        """), {"id": rd_id}).mappings().all()
        despesa_importada = conn.execute(text(
            "SELECT id FROM financeiro3_despesas WHERE origem_rd_id=:id"
        ), {"id": rd_id}).scalar()
        opcoes = _opcoes(conn)
    valor_pago = sum((pg["valor"] for pg in pagamentos if pg["status"] == "PAGO"), start=0)
    valor_previsto = sum((pg["valor"] for pg in pagamentos if pg["status"] == "PREVISTO"), start=0)
    return render_template(
        "financeiro_novo/rd_detalhe.html", rd=rd, itens=itens, anexos=anexos,
        decisoes=decisoes, acerto=acerto, pagamentos=pagamentos,
        valor_pago=valor_pago, valor_previsto=valor_previsto,
        despesa_importada=despesa_importada, opcoes=opcoes, formas=FORMAS_PAGAMENTO,
        editavel=rd["status"] in EDITAVEIS, subnav_links=build_subnav("missoes"),
    )


@bp.get("/rds/<int:rd_id>/verificar-duplicidades")
@login_required
@permission_required("financeiro_novo", "editar")
def rd_verificar_duplicidades(rd_id):
    resultado = []
    with get_engine().connect() as conn:
        if not _registro(conn, "financeiro3_rds", rd_id):
            abort(404)
        try:
            data = data_iso(request.args.get("data"), "Data da despesa")
            valor = decimal_br(request.args.get("valor"), positivo=True)
        except ValorInvalido:
            return jsonify({"duplicidades": []})
        registros = _origens_duplicadas(conn, data, valor)
        if registros:
            resultado.append({"linha": 1, "registros": registros})
    return jsonify({"duplicidades": resultado})


@bp.post("/rds/<int:rd_id>/itens")
@login_required
@permission_required("financeiro_novo", "editar")
def rd_item_novo(rd_id):
    try:
        dados = {"data": data_iso(request.form.get("data_despesa"), "Data da despesa"),
                 "categoria": int(request.form.get("categoria_id") or 0),
                 "fornecedor": int(request.form.get("fornecedor_id") or 0) or None,
                 "descricao": (request.form.get("descricao") or "").strip(),
                 "documento": (request.form.get("numero_documento") or "").strip() or None,
                 "valor": decimal_br(request.form.get("valor"), positivo=True)}
        if not dados["descricao"]:
            raise ValorInvalido("Informe a descrição do gasto.")
        with get_engine().begin() as conn:
            rd = _registro(conn, "financeiro3_rds", rd_id, True)
            if not rd:
                abort(404)
            if rd["status"] not in EDITAVEIS:
                abort(409)
            if not (rd["periodo_inicio"] <= dados["data"] <= rd["periodo_fim"]):
                raise ValorInvalido("A data do gasto deve estar dentro do período da RD.")
            referencia_ok = conn.execute(text("""
                SELECT EXISTS(SELECT 1 FROM financeiro3_categorias
                    WHERE id=:categoria AND ativo AND natureza='DESPESA')
                  AND (:fornecedor IS NULL OR EXISTS(SELECT 1 FROM financeiro3_pessoas
                    WHERE id=:fornecedor AND ativo AND fornecedor))
            """), dados).scalar()
            if not referencia_ok:
                raise ValorInvalido("Categoria ou fornecedor está inativo ou inválido para a RD.")
            duplicidades = _origens_duplicadas(conn, dados["data"], dados["valor"])
            if duplicidades and request.form.get("forcar_salvamento") != "1":
                rotulos = ", ".join(f"{item['tipo']} {item['numero']}" for item in duplicidades)
                raise ValorInvalido(
                    f"Possível lançamento duplicado ({rotulos}). Confirme para salvar."
                )
            item = conn.execute(text("""
                INSERT INTO financeiro3_rd_itens(rd_id,data_despesa,categoria_id,fornecedor_id,
                    descricao,numero_documento,valor,criado_por)
                VALUES (:rd,:data,:categoria,:fornecedor,:descricao,:documento,:valor,:usuario) RETURNING *
            """), {**dados, "rd": rd_id, "usuario": session.get("usuario_id")}).mappings().one()
            registrar_evento(conn, entidade="RD_ITEM", entidade_id=item["id"], evento="CRIADO", dados_novos=dict(item))
        flash("Gasto incluído na RD.", "sucesso")
    except (ValorInvalido, ValueError) as exc:
        flash(str(exc), "erro")
    return redirect(url_for("financeiro_novo.rd_detalhe", rd_id=rd_id))


@bp.post("/rds/<int:rd_id>/itens/<int:item_id>/remover")
@login_required
@permission_required("financeiro_novo", "editar")
def rd_item_remover(rd_id, item_id):
    with get_engine().begin() as conn:
        rd = _registro(conn, "financeiro3_rds", rd_id, True)
        if not rd or rd["status"] not in EDITAVEIS:
            abort(409 if rd else 404)
        anterior = conn.execute(text("SELECT * FROM financeiro3_rd_itens WHERE id=:item AND rd_id=:rd AND status='ATIVO' FOR UPDATE"), {"item": item_id, "rd": rd_id}).mappings().first()
        if not anterior:
            abort(404)
        novo = conn.execute(text("UPDATE financeiro3_rd_itens SET status='REMOVIDO',removido_por=:u,removido_em=NOW() WHERE id=:id RETURNING *"), {"u": session.get("usuario_id"), "id": item_id}).mappings().one()
        registrar_evento(conn, entidade="RD_ITEM", entidade_id=item_id, evento="REMOVIDO", dados_anteriores=dict(anterior), dados_novos=dict(novo))
    return redirect(url_for("financeiro_novo.rd_detalhe", rd_id=rd_id))


@bp.post("/rds/<int:rd_id>/enviar")
@login_required
@permission_required("financeiro_novo", "editar")
def rd_enviar(rd_id):
    with get_engine().begin() as conn:
        anterior = _registro(conn, "financeiro3_rds", rd_id, True)
        if not anterior or anterior["status"] not in EDITAVEIS:
            abort(409 if anterior else 404)
        if anterior["valor_total"] <= 0 or not _tem_anexo(conn, "RD", rd_id):
            flash("Inclua gastos e ao menos um comprovante antes de enviar.", "erro")
            return redirect(url_for("financeiro_novo.rd_detalhe", rd_id=rd_id))
        novo = conn.execute(text("UPDATE financeiro3_rds SET status='EM_APROVACAO',atualizado_em=NOW() WHERE id=:id RETURNING *"), {"id": rd_id}).mappings().one()
        _decisao(conn, "rd", rd_id, "ENVIO", anterior["status"], "EM_APROVACAO")
        registrar_evento(conn, entidade="RD", entidade_id=rd_id, evento="ENVIADA_APROVACAO", dados_anteriores=dict(anterior), dados_novos=dict(novo))
    flash("RD enviada para aprovação.", "sucesso")
    return redirect(url_for("financeiro_novo.rd_detalhe", rd_id=rd_id))


def _decidir_rd(rd_id, aprovar):
    justificativa = (request.form.get("justificativa") or "").strip()
    if not aprovar and not justificativa:
        flash("Informe a justificativa da rejeição.", "erro")
        return
    with get_engine().begin() as conn:
        anterior = _registro(conn, "financeiro3_rds", rd_id, True)
        if not anterior or anterior["status"] != "EM_APROVACAO":
            abort(409 if anterior else 404)
        if _proibir_autoaprovacao(anterior):
            flash("O responsável pelo lançamento não pode decidir a própria RD.", "erro")
            return
        status = "APROVADA" if aprovar else "REJEITADA"
        total_pago = conn.execute(text("""
            SELECT COALESCE(SUM(valor),0) FROM financeiro3_rd_pagamentos
            WHERE rd_id=:id AND status='PAGO'
        """), {"id": rd_id}).scalar()
        reembolsos_vinculados = conn.execute(text("""
            SELECT COALESCE(SUM(i.valor),0) FROM financeiro3_reembolso_itens i
            JOIN financeiro3_reembolsos r ON r.id=i.reembolso_id
            WHERE r.forma_liquidacao='RD' AND r.rd_pagadora_id=:id
              AND r.status='APROVADO' AND i.status='ATIVO'
        """), {"id": rd_id}).scalar()
        diferenca = anterior["valor_total"] + reembolsos_vinculados - total_pago
        data_prevista = None
        if aprovar and diferenca != 0:
            try:
                data_prevista = data_iso(request.form.get("data_prevista_liquidacao"), "Data prevista do acerto")
            except ValorInvalido as exc:
                flash(str(exc), "erro")
                return
        if aprovar and diferenca == 0:
            status = "LIQUIDADA"
        novo = conn.execute(text("UPDATE financeiro3_rds SET status=:status,valor_adiantamento=:pago,atualizado_em=NOW() WHERE id=:id RETURNING *"), {"status": status, "pago": total_pago, "id": rd_id}).mappings().one()
        acao = "APROVACAO" if aprovar else "REJEICAO"
        _decisao(conn, "rd", rd_id, acao, "EM_APROVACAO", status, justificativa or None)
        if aprovar and diferenca != 0:
            conn.execute(text("""
                INSERT INTO financeiro3_rd_acertos(rd_id,tipo,valor,data_prevista_liquidacao,criado_por)
                VALUES (:rd,:tipo,:valor,:data,:usuario)
            """), {"rd": rd_id, "tipo": "REEMBOLSO" if diferenca > 0 else "DEVOLUCAO",
                    "valor": abs(diferenca), "data": data_prevista,
                    "usuario": session.get("usuario_id")})
        registrar_evento(conn, entidade="RD", entidade_id=rd_id, evento=acao, dados_anteriores=dict(anterior), dados_novos=dict(novo), justificativa=justificativa)
    flash("RD aprovada e acerto calculado." if aprovar else "RD rejeitada para correção.", "sucesso")


@bp.post("/rds/<int:rd_id>/aprovar")
@login_required
@permission_required("financeiro_novo", "aprovar")
def rd_aprovar(rd_id):
    _decidir_rd(rd_id, True)
    return redirect(url_for("financeiro_novo.rd_detalhe", rd_id=rd_id))


@bp.post("/rds/<int:rd_id>/rejeitar")
@login_required
@permission_required("financeiro_novo", "aprovar")
def rd_rejeitar(rd_id):
    _decidir_rd(rd_id, False)
    return redirect(url_for("financeiro_novo.rd_detalhe", rd_id=rd_id))


@bp.post("/rds/<int:rd_id>/liquidar")
@login_required
@permission_required("financeiro_novo", "pagar")
def rd_liquidar(rd_id):
    try:
        conta = int(request.form.get("conta_id") or 0)
        data = data_iso(request.form.get("data_liquidacao"), "Data de liquidação")
        forma = (request.form.get("forma") or "").upper()
        if forma not in dict(FORMAS_PAGAMENTO):
            raise ValorInvalido("Forma inválida.")
        with get_engine().begin() as conn:
            rd = _registro(conn, "financeiro3_rds", rd_id, True)
            acerto = conn.execute(text("SELECT * FROM financeiro3_rd_acertos WHERE rd_id=:id FOR UPDATE"), {"id": rd_id}).mappings().first()
            if not rd or not acerto or rd["status"] != "APROVADA" or acerto["status"] != "PENDENTE":
                abort(409)
            conta_ok = conn.execute(text("SELECT EXISTS(SELECT 1 FROM financeiro3_contas WHERE id=:id AND ativo AND moeda_id=:moeda)"), {"id": conta, "moeda": rd["moeda_id"]}).scalar()
            if not conta_ok:
                raise ValorInvalido("Selecione uma conta ativa na moeda da RD.")
            conn.execute(text("""
                UPDATE financeiro3_rd_acertos SET status='LIQUIDADO',conta_id=:conta,
                    data_liquidacao=:data,forma=:forma,referencia=:referencia,
                    liquidado_por=:usuario,liquidado_em=NOW() WHERE id=:id
            """), {"conta": conta, "data": data, "forma": forma,
                    "referencia": (request.form.get("referencia") or "").strip() or None,
                    "usuario": session.get("usuario_id"), "id": acerto["id"]})
            novo = conn.execute(text("UPDATE financeiro3_rds SET status='LIQUIDADA',atualizado_em=NOW() WHERE id=:id RETURNING *"), {"id": rd_id}).mappings().one()
            _decisao(conn, "rd", rd_id, "LIQUIDACAO", "APROVADA", "LIQUIDADA")
            registrar_evento(conn, entidade="RD", entidade_id=rd_id, evento="ACERTO_LIQUIDADO", dados_anteriores=dict(rd), dados_novos=dict(novo))
        flash("Acerto da RD liquidado.", "sucesso")
    except (ValorInvalido, ValueError) as exc:
        flash(str(exc), "erro")
    return redirect(url_for("financeiro_novo.rd_detalhe", rd_id=rd_id))


def _listar_anexos(conn, entidade, registro_id):
    return conn.execute(text("""
        SELECT a.id, a.categoria, ar.id AS arquivo_id, ar.nome_original, ar.tamanho_canonico, ar.paginas
        FROM financeiro3_anexos a JOIN financeiro3_arquivos ar ON ar.id=a.arquivo_id
        WHERE a.entidade=:entidade AND a.entidade_id=:id AND a.status='ATIVO' ORDER BY a.id DESC
    """), {"entidade": entidade, "id": registro_id}).mappings().all()


def _caminho(object_key):
    raiz = Path(current_app.config["UPLOAD_ROOT"]).resolve()
    destino = (raiz / object_key).resolve()
    if not destino.is_relative_to(raiz):
        abort(404)
    return destino


@bp.get("/oms/<int:om_id>/itens/<int:item_id>/anexos/<uuid:arquivo_id>")
@login_required
@permission_required("financeiro_novo", "visualizar")
def om_item_anexo_baixar(om_id, item_id, arquivo_id):
    with get_engine().connect() as conn:
        arquivo = conn.execute(text("""
            SELECT ar.object_key,ar.nome_original FROM financeiro3_om_itens i
            JOIN financeiro3_anexos a ON a.entidade='OM_ITEM' AND a.entidade_id=i.id AND a.status='ATIVO'
            JOIN financeiro3_arquivos ar ON ar.id=a.arquivo_id AND ar.status='ATIVO'
            JOIN financeiro3_oms o ON o.id=i.om_id AND o.removido_em IS NULL
            WHERE i.id=:item AND i.om_id=:om AND a.arquivo_id=:arquivo AND i.status='ATIVO'
        """), {"item": item_id, "om": om_id, "arquivo": arquivo_id}).mappings().first()
    if not arquivo: abort(404)
    caminho = _caminho(arquivo["object_key"])
    if not caminho.is_file():
        flash("O recibo não está mais disponível no armazenamento. Configure o volume persistente e use Substituir para reenviá-lo.", "erro")
        return redirect(url_for("financeiro_novo.om_detalhe", om_id=om_id))
    return send_file(caminho, mimetype="application/pdf", as_attachment=True,
        download_name=f"{Path(arquivo['nome_original']).stem}.pdf")


@bp.post("/documentos/<tipo>/<int:registro_id>/anexos")
@login_required
@permission_required("financeiro_novo", "editar")
def documento_anexo_novo(tipo, registro_id):
    if tipo not in DOCUMENTOS:
        abort(404)
    tabela, entidade = DOCUMENTOS[tipo]
    destino = temporario = None
    try:
        arquivo = normalizar_anexo(request.files.get("arquivo"))
        arquivo_id = uuid.uuid4()
        object_key = nome_objeto_pdf(str(arquivo_id))
        destino = _caminho(object_key)
        destino.parent.mkdir(parents=True, exist_ok=True)
        temporario = destino.with_suffix(f".{uuid.uuid4().hex}.tmp")
        temporario.write_bytes(arquivo.conteudo)
        os.replace(temporario, destino)
        with get_engine().begin() as conn:
            registro = _registro(conn, tabela, registro_id, True)
            if not registro or registro["status"] not in EDITAVEIS:
                abort(409 if registro else 404)
            conn.execute(text("""
                INSERT INTO financeiro3_arquivos(id,storage_backend,object_key,nome_original,mime_original,
                    sha256_original,sha256_canonico,tamanho_original,tamanho_canonico,paginas,
                    compressao_aplicada,assinatura_digital_detectada,criado_por)
                VALUES (:id,'VOLUME',:key,:nome,:mime,:sha_o,:sha_c,:tam_o,:tam_c,:paginas,:compressao,:assinatura,:usuario)
            """), {"id": arquivo_id, "key": object_key, "nome": arquivo.nome_original,
                    "mime": arquivo.mime_original, "sha_o": arquivo.sha256_original,
                    "sha_c": arquivo.sha256_canonico, "tam_o": arquivo.tamanho_original,
                    "tam_c": arquivo.tamanho_canonico, "paginas": arquivo.paginas,
                    "compressao": arquivo.compressao_aplicada, "assinatura": arquivo.assinatura_digital_detectada,
                    "usuario": session.get("usuario_id")})
            vinculo = conn.execute(text("""
                INSERT INTO financeiro3_anexos(arquivo_id,entidade,entidade_id,categoria,criado_por)
                VALUES (:arquivo,:entidade,:registro,'COMPROVANTE',:usuario) RETURNING id
            """), {"arquivo": arquivo_id, "entidade": entidade, "registro": registro_id,
                    "usuario": session.get("usuario_id")}).scalar()
            registrar_evento(conn, entidade=f"{entidade}_ANEXO", entidade_id=vinculo, evento="ANEXADO", dados_novos={"arquivo_id": str(arquivo_id), "nome": arquivo.nome_original})
        flash("Documento convertido e anexado em PDF.", "sucesso")
    except AnexoInvalido as exc:
        flash(str(exc), "erro")
    except Exception:
        if destino and destino.exists():
            destino.unlink(missing_ok=True)
        raise
    finally:
        if temporario and temporario.exists():
            temporario.unlink(missing_ok=True)
    endpoints = {"om": ("financeiro_novo.om_detalhe", "om_id"), "rd": ("financeiro_novo.rd_detalhe", "rd_id"), "nd": ("financeiro_novo.nd_detalhe", "nd_id")}
    endpoint, parametro = endpoints[tipo]
    return redirect(url_for(endpoint, **{parametro: registro_id}))


@bp.get("/documentos/<tipo>/<int:registro_id>/anexos/<uuid:arquivo_id>")
@login_required
@permission_required("financeiro_novo", "visualizar")
def documento_anexo_baixar(tipo, registro_id, arquivo_id):
    if tipo not in DOCUMENTOS:
        abort(404)
    _, entidade = DOCUMENTOS[tipo]
    with get_engine().connect() as conn:
        arquivo = conn.execute(text("""
            SELECT ar.object_key,ar.nome_original FROM financeiro3_anexos a
            JOIN financeiro3_arquivos ar ON ar.id=a.arquivo_id
            WHERE a.entidade=:entidade AND a.entidade_id=:registro AND a.arquivo_id=:arquivo
              AND a.status='ATIVO' AND ar.status='ATIVO'
        """), {"entidade": entidade, "registro": registro_id, "arquivo": arquivo_id}).mappings().first()
        if not arquivo and tipo in {"om", "rd"}:
            _, tabela_pagamento, fk, _, _, _ = _pagamento_config(tipo)
            arquivo = conn.execute(text(f"""
                SELECT ar.object_key,ar.nome_original FROM {tabela_pagamento} pg
                JOIN financeiro3_anexos a ON a.entidade=:entidade_pagamento
                  AND a.entidade_id=pg.id AND a.status='ATIVO'
                JOIN financeiro3_arquivos ar ON ar.id=a.arquivo_id AND ar.status='ATIVO'
                WHERE pg.{fk}=:registro AND a.arquivo_id=:arquivo
            """), {"entidade_pagamento": f"{entidade}_PAGAMENTO",
                     "registro": registro_id, "arquivo": arquivo_id}).mappings().first()
    if not arquivo:
        abort(404)
    caminho = _caminho(arquivo["object_key"])
    if not caminho.is_file():
        abort(404)
    return send_file(caminho, mimetype="application/pdf", as_attachment=True,
                     download_name=f"{Path(arquivo['nome_original']).stem}.pdf")


@bp.get("/documentos/<tipo>/<int:registro_id>/pagamentos/<int:pagamento_id>/anexos/<uuid:arquivo_id>")
@login_required
@permission_required("financeiro_novo", "visualizar")
def pagamento_anexo_baixar(tipo, registro_id, pagamento_id, arquivo_id):
    if tipo not in {"om", "rd"}:
        abort(404)
    _, tabela, fk, entidade, _, _ = _pagamento_config(tipo)
    with get_engine().connect() as conn:
        arquivo = conn.execute(text(f"""
            SELECT ar.object_key,ar.nome_original FROM {tabela} pg
            JOIN financeiro3_anexos a ON a.entidade=:entidade
              AND a.entidade_id=pg.id AND a.status='ATIVO'
            JOIN financeiro3_arquivos ar ON ar.id=a.arquivo_id AND ar.status='ATIVO'
            WHERE pg.id=:pagamento AND pg.{fk}=:registro AND a.arquivo_id=:arquivo
        """), {"entidade": f"{entidade}_PAGAMENTO", "pagamento": pagamento_id,
                 "registro": registro_id, "arquivo": arquivo_id}).mappings().first()
    if not arquivo:
        abort(404)
    caminho = _caminho(arquivo["object_key"])
    if not caminho.is_file():
        abort(404)
    return send_file(caminho, mimetype="application/pdf", as_attachment=True,
                     download_name=f"{Path(arquivo['nome_original']).stem}.pdf")


@bp.post("/documentos/<tipo>/<int:registro_id>/anexos/<int:anexo_id>/remover")
@login_required
@permission_required("financeiro_novo", "editar")
def documento_anexo_remover(tipo, registro_id, anexo_id):
    if tipo not in DOCUMENTOS:
        abort(404)
    tabela, entidade = DOCUMENTOS[tipo]
    with get_engine().begin() as conn:
        registro = _registro(conn, tabela, registro_id, True)
        if not registro or registro["status"] not in EDITAVEIS:
            abort(409 if registro else 404)
        anterior = conn.execute(text("""
            SELECT * FROM financeiro3_anexos WHERE id=:anexo AND entidade=:entidade
              AND entidade_id=:registro AND status='ATIVO' FOR UPDATE
        """), {"anexo": anexo_id, "entidade": entidade, "registro": registro_id}).mappings().first()
        if not anterior:
            abort(404)
        novo = conn.execute(text("""
            UPDATE financeiro3_anexos SET status='REMOVIDO',removido_por=:usuario,
              removido_em=NOW(),motivo_remocao='Removido durante edição do rascunho'
            WHERE id=:id RETURNING *
        """), {"usuario": session.get("usuario_id"), "id": anexo_id}).mappings().one()
        registrar_evento(conn, entidade=f"{entidade}_ANEXO", entidade_id=anexo_id,
                         evento="REMOVIDO", dados_anteriores=dict(anterior), dados_novos=dict(novo))
    flash("Documento removido do lançamento e preservado na auditoria.", "sucesso")
    endpoints = {"om": ("financeiro_novo.om_detalhe", "om_id"), "rd": ("financeiro_novo.rd_detalhe", "rd_id"), "nd": ("financeiro_novo.nd_detalhe", "nd_id")}
    endpoint, parametro = endpoints[tipo]
    return redirect(url_for(endpoint, **{parametro: registro_id}))
